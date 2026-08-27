"""Parser del Excel Integrity (mayor del día).

Portado de reference/auditoria.py. Sin pandas.
- find_sheet(): identifica la hoja del mayor por su ENCABEZADO.
- parse_integrity_lines(): líneas crudas (grano de stg_integrity_line).
- aggregate_lines(): agrega por TCode (int_cr/int_db) para la reconciliación.
- parse_integrity(): atajo lines -> aggregate (compatibilidad).

TCode = regex `TCode(?:\\s+CXC)?:\\s*(\\d+)` sobre 'Referencia'.
Columnas USD reales: 'Créditos Dol' / 'Débitos Dol' (NO 'Col').

Ni el nombre de la hoja ni la fila del encabezado se dan por fijos: el reporte
se llamó 'Datos' y pasó a 'Asiento' (2026-08-27) sin avisar, y la fila 8 es una
convención del emisor, no un contrato. Ambos se detectan por contenido — el
mismo criterio que usa la ingesta para clasificar los archivos (§2.8).
"""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

TCODE_RE = re.compile(r"TCode(?:\s+CXC)?:\s*(\d+)")
# Columnas que identifican al mayor. Si están las dos, es la hoja correcta.
HEADER_HINTS = ("Cuenta", "Referencia")
HEADER_SCAN = 25  # cuántas filas mirar buscando el encabezado


def _header_index(rows: list[list]) -> int | None:
    """Fila del encabezado, ubicada por sus nombres de columna."""
    for i, r in enumerate(rows[:HEADER_SCAN]):
        vals = {str(c).strip() for c in r if c is not None}
        if all(h in vals for h in HEADER_HINTS):
            return i
    return None


def find_sheet(wb) -> str | None:
    """Nombre de la hoja del mayor dentro de un workbook ya abierto, o None.

    Lo usa también la clasificación de la ingesta para decidir si un .xlsx es
    un Integrity, sin depender de cómo se llame la hoja ni el archivo."""
    for name in wb.sheetnames:
        rows = []
        for i, r in enumerate(wb[name].iter_rows(values_only=True)):
            rows.append(list(r))
            if i >= HEADER_SCAN:
                break
        if _header_index(rows) is not None:
            return name
    return None


def _tcode(ref) -> str | None:
    if ref is None:
        return None
    m = TCODE_RE.search(str(ref))
    return m.group(1) if m else None


def _num(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def parse_integrity_lines(path: str | Path, sheet: str | None = None) -> list[dict]:
    """Líneas crudas del mayor (una por asiento), con tcode parseado de Referencia.

    `sheet=None` (lo normal) detecta la hoja sola."""
    wb = load_workbook(path, read_only=True, data_only=True)
    hoja = sheet or find_sheet(wb)
    if hoja is None or hoja not in wb.sheetnames:
        wb.close()
        return []
    rows = [list(r) for r in wb[hoja].iter_rows(values_only=True)]
    wb.close()
    hdr = _header_index(rows)
    if hdr is None or len(rows) <= hdr:
        return []
    header = [str(c).strip() if c is not None else "" for c in rows[hdr]]
    col = {n: i for i, n in enumerate(header)}

    def cell(r, name):
        i = col.get(name)
        return r[i] if i is not None and i < len(r) else None

    out = []
    for r in rows[hdr + 1:]:
        if not any(c not in (None, "") for c in r):
            continue
        cuenta = cell(r, "Cuenta")
        if cuenta is None:
            continue
        out.append({
            "cuenta": str(cuenta).strip(),
            "nombre_cuenta": (str(cell(r, "Nombre cuenta")).strip()
                              if cell(r, "Nombre cuenta") else None),
            "centro_costo": (str(cell(r, "Centro de costo")).strip()
                             if cell(r, "Centro de costo") else None),
            "referencia": (str(cell(r, "Referencia")).strip()
                           if cell(r, "Referencia") else None),
            "detalle": str(cell(r, "Detalle")).strip() if cell(r, "Detalle") else None,
            "moneda_fuente": str(cell(r, "Moneda")).strip() if cell(r, "Moneda") else None,
            "tc": _num(cell(r, "T.C.")) or None,
            "deb_col": _num(cell(r, "Débitos Col")),
            "cred_col": _num(cell(r, "Créditos Col")),
            "deb_usd": _num(cell(r, "Débitos Dol")),
            "cred_usd": _num(cell(r, "Créditos Dol")),
            "tcode": _tcode(cell(r, "Referencia")),
        })
    return out


def aggregate_lines(lines: list[dict]) -> list[dict]:
    """Agrega líneas por tcode (solo las que tienen tcode): int_cr, int_db, cuenta, nombre, tc."""
    agg: dict[str, dict] = {}
    for ln in lines:
        tc = ln.get("tcode")
        if not tc:
            continue
        a = agg.setdefault(tc, {
            "tcode": tc, "int_cr": 0.0, "int_db": 0.0,
            "cuentas": set(), "nombres": set(), "tc": None,
        })
        a["int_cr"] += ln["cred_usd"]
        a["int_db"] += ln["deb_usd"]
        if ln.get("cuenta"):
            a["cuentas"].add(ln["cuenta"])
        if ln.get("nombre_cuenta"):
            a["nombres"].add(ln["nombre_cuenta"])
        if a["tc"] is None:
            a["tc"] = ln.get("tc")
    return [{
        "tcode": a["tcode"],
        "int_cr": round(a["int_cr"], 2),
        "int_db": round(a["int_db"], 2),
        "cuenta": " | ".join(sorted(a["cuentas"])),
        "nombre": " | ".join(sorted(a["nombres"])),
        "tc": a["tc"],
    } for a in agg.values()]


def parse_integrity(path: str | Path, sheet: str | None = None) -> list[dict]:
    """Compatibilidad: líneas crudas -> agregado por tcode."""
    return aggregate_lines(parse_integrity_lines(path, sheet))
