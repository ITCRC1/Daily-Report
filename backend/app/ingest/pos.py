"""Parser del Excel de Ventas Simphony/POS (portado de reference/auditoria.py).

Lee 'Resumen Ejecutivo', 'Detalle de Checks' y 'Mapeo Simphony → Opera'.
Devuelve un dict con checks, resúmenes por forma de pago / empleado y room charges.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

RC_SHEET = "Mapeo Simphony → Opera"  # 'Mapeo Simphony → Opera'


def _num(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _sheet_rows(wb, name):
    ws = wb[name]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _find_header(rows, needle, max_scan=5):
    """Índice de la fila cuyo encabezado contiene `needle` (ej '# Check')."""
    for i, r in enumerate(rows[:max_scan + 1]):
        cells = [str(c).strip() if c is not None else "" for c in r]
        if needle in cells:
            return i, cells
    return None, None


def parse_pos_excel(path: str | Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    meta: dict = {
        "source": "excel",
        "all_checks": [], "by_payment": [], "by_employee": [], "rc_detail": [],
        "room_charge": 0.0,
    }

    # ── Resumen Ejecutivo: buscar valores por keyword ──────────────────
    if "Resumen Ejecutivo" in wb.sheetnames:
        rows = _sheet_rows(wb, "Resumen Ejecutivo")

        def find_val(keyword):
            for r in rows:
                vals = [v for v in r if v not in (None, "")]
                if any(keyword.lower() in str(v).lower() for v in vals):
                    nums = [v for v in vals if isinstance(v, (int, float))]
                    if nums:
                        return float(nums[0])
            return 0.0

        meta["ventas_netas"] = find_val("Ventas Netas")
        meta["sc"] = find_val("Cargos de Servicio")
        meta["total_dia"] = find_val("TOTAL VENTAS")
        meta["voids"] = abs(find_val("Anulaciones"))

    # ── Detalle de Checks: header en fila 1 o 2 ────────────────────────
    if "Detalle de Checks" in wb.sheetnames:
        rows = _sheet_rows(wb, "Detalle de Checks")
        hidx, header = _find_header(rows, "# Check")
        if header:
            col = {name: i for i, name in enumerate(header)}
            pay_tot: dict[str, list[float]] = {}
            emp_tot: dict[str, list[float]] = {}
            for r in rows[hidx + 1:]:
                cn = r[col["# Check"]] if col.get("# Check") is not None else None
                if cn is None or str(cn).strip() == "":
                    continue
                try:
                    check_num = str(int(cn))
                except (TypeError, ValueError):
                    continue
                monto = _num(r[col.get("Monto (USD)")]) if "Monto (USD)" in col else 0.0
                forma = str(r[col.get("Forma de Pago")]) if "Forma de Pago" in col else ""
                emp = str(r[col.get("Empleado")]) if "Empleado" in col else ""
                meta["all_checks"].append({
                    "restaurant": str(r[col.get("Restaurante")]) if "Restaurante" in col else "",
                    "employee": emp,
                    "check_num": check_num,
                    "hora": str(r[col.get("Hora Cierre")]) if "Hora Cierre" in col else "",
                    "forma_pago": forma,
                    "monto": monto,
                })
                pay_tot.setdefault(forma, []).append(monto)
                emp_tot.setdefault(emp, []).append(monto)
            meta["by_payment"] = [
                {"forma": k, "count": len(v), "total": round(sum(v), 2)} for k, v in pay_tot.items()
            ]
            meta["by_employee"] = [
                {"empleado": k, "count": len(v), "total": round(sum(v), 2)} for k, v in emp_tot.items()
            ]

    # ── Mapeo Simphony → Opera: room charges confirmados ───────────────
    if RC_SHEET in wb.sheetnames:
        rows = _sheet_rows(wb, RC_SHEET)
        hidx, header = _find_header(rows, "# Check")
        if header:
            col = {name: i for i, name in enumerate(header)}
            rc_total = 0.0
            for r in rows[hidx + 1:]:
                cn = r[col["# Check"]] if col.get("# Check") is not None else None
                if cn is None or str(cn).strip() == "":
                    continue
                try:
                    check_num = str(int(cn))
                except (TypeError, ValueError):
                    continue
                monto = _num(r[col.get("Monto Cargado (USD)")]) if "Monto Cargado (USD)" in col else 0.0
                rc_total += monto
                meta["rc_detail"].append({
                    "restaurant": str(r[col.get("Restaurante")]) if "Restaurante" in col else "",
                    "employee": str(r[col.get("Empleado")]) if "Empleado" in col else "",
                    "check_num": check_num,
                    "hora": str(r[col.get("Hora Cierre")]) if "Hora Cierre" in col else "",
                    "monto": monto,
                })
            meta["room_charge"] = round(rc_total, 2)

    wb.close()
    return meta
