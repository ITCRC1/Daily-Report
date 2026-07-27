"""Bootstrap de datos iniciales (Budget 2026 + Revenue Actual diario 2026).

Carga los 2 Excel versionados en `db/seed_data/` directamente contra la DB
(mismos servicios que usan los endpoints de upload, sin pasar por HTTP) --
así CUALQUIER servidor nuevo (hotel, staging, otro futuro) queda con los
datos reales sin tener que volver a subirlos a mano desde la UI.

Orden de arranque en un servidor nuevo:
    python -m app.seed                 # dim_property base
    python -m app.seed_from_goldens    # dim_department + dim_payment_map
    python -m app.bootstrap_initial_data

Idempotente: `upload_and_replace` (Budget) hace reset del año completo;
`upload_daily_grid` (Revenue Actual) reemplaza solo los días presentes en el
archivo. Volver a correr este script no duplica nada.

Si los Excel de origen cambian (nuevo presupuesto, más días de revenue real),
reemplazá los archivos en `db/seed_data/` y volvé a correr este script -- no
hace falta tocar código.
"""
import asyncio
import json
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.dialects.postgresql import insert as pg_insert

from app.config import get_settings
from app.db import SessionLocal
from app.models import CompStatMonthly, Department, Property, RevenueActualDaily, RoomStatOpening
from app.services import budget_service, revenue_actual_service

COMPS_SENTINEL = "Comps/In-House"

settings = get_settings()

SEED_DATA = Path(__file__).resolve().parents[2] / "db" / "seed_data"
BUDGET_FILE = SEED_DATA / "Budget_COWLCR_2026.xlsx"
REVENUE_ACTUAL_FILE = SEED_DATA / "RevenueActual_COWLCR_2026.xlsx"
COMPSTAT_FILE = SEED_DATA / "CompStat_COWLCR_2026.xlsx"
ROOM_STAT_OPENING_FILE = SEED_DATA / "RoomStatOpening_COWLCR_2026.json"
PL_BASELINE_FILE = SEED_DATA / "PL_Baseline_COWLCR_2026_H1.json"
YEAR = 2026


async def _property_id(s):
    return (await s.execute(
        select(Property.id).where(Property.code == settings.DEFAULT_PROPERTY)
    )).scalar_one()


async def _load_comps_monthly(s, pid) -> None:
    """comp_stat_monthly desde el Excel del owner (comps Ene-Jun por tipo).
    INSERT-only (DO NOTHING): nunca pisa lo que ya está cargado."""
    wb = load_workbook(COMPSTAT_FILE, data_only=True)
    ws = wb["STATISTIC"]

    def block(r0):
        return {ws.cell(r, 2).value: [ws.cell(r, c).value or 0 for c in range(3, 9)] for r in range(r0, r0 + 6)}

    rn, pax = block(29), block(41)  # Complimentary RN (29-34) y Pax (41-46)
    rows = [{"property_id": pid, "year": YEAR, "month": i + 1, "room_category": cat,
             "rn": rn[cat][i], "pax": pax[cat][i]} for cat in rn for i in range(6)]
    await s.execute(pg_insert(CompStatMonthly).values(rows)
                    .on_conflict_do_nothing(index_elements=["property_id", "year", "month", "room_category"]))


async def _load_room_stat_opening(s, pid) -> None:
    """room_stat_opening (ancla de Tab 6.6, categorías en GROSS + línea Comps)
    desde el JSON versionado. INSERT-only: no pisa un ancla ya editado."""
    data = json.loads(ROOM_STAT_OPENING_FILE.read_text(encoding="utf-8"))
    rows = [{"property_id": pid, "room_category": o["room_category"],
             "anchor_date": date.fromisoformat(o["anchor_date"]), "revenue": o["revenue"],
             "stay_rooms": o["stay_rooms"], "stay_persons": o["stay_persons"],
             "physical_rooms": o["physical_rooms"]} for o in data]
    await s.execute(pg_insert(RoomStatOpening).values(rows)
                    .on_conflict_do_nothing(index_elements=["property_id", "room_category"]))


async def _reconcile_rooms_to_anchor(s, pid) -> None:
    """Ajusta el room revenue diario del grid 6.4 en el tramo PRE-corte (Ene 1 ->
    anchor_date) para que su total = el total del ancla de Tab 6.6 (revenue por
    categoría, la fuente autoritativa confirmada por el owner). El grid diario y
    el ancla son dos inputs del owner que diferían $14,639.29 en Ene-Jun; acá se
    reconcilia el detalle al control total (escalado proporcional, residual al
    último día). Idempotente: si ya está reconciliado, factor=1, no cambia nada.
    Así 'Revenue – Rooms' (todos los reportes) = room stats total siempre."""
    anchors = (await s.execute(
        select(RoomStatOpening).where(RoomStatOpening.property_id == pid)
    )).scalars().all()
    cats = [a for a in anchors if a.room_category != COMPS_SENTINEL]
    if not cats:
        return
    target = sum(Decimal(str(a.revenue)) for a in cats)
    anchor_date = min(a.anchor_date for a in cats)
    dept = (await s.execute(
        select(Department.id).where(Department.property_id == pid).where(Department.cost_center == "0110")
    )).scalar_one_or_none()
    if dept is None or target <= 0:
        return
    rows = (await s.execute(
        select(RevenueActualDaily).where(RevenueActualDaily.property_id == pid)
        .where(RevenueActualDaily.dept_id == dept).where(RevenueActualDaily.date <= anchor_date)
        .order_by(RevenueActualDaily.date)
    )).scalars().all()
    cur = sum(Decimal(str(r.amount_usd)) for r in rows)
    if not rows or cur <= 0 or cur == target:
        return
    factor = target / cur
    acc = Decimal("0")
    for r in rows:
        nv = (Decimal(str(r.amount_usd)) * factor).quantize(Decimal("0.01"))
        r.amount_usd = nv
        acc += nv
    rows[-1].amount_usd = Decimal(str(rows[-1].amount_usd)) + (target - acc)
    print(f"Rooms grid pre-corte reconciliado al ancla: {cur} -> {target}")


async def _reconcile_categories_to_pl(s, pid) -> None:
    """Alinea el grid 6.4 Ene-Jun a la base de arranque del P&L del owner
    (PL_Baseline_*.json): escala cada grupo de cost_centers para que su total
    del período dé exacto el target (residual al último día). target=0 -> pone
    en 0 (ej. Innoceana). Rooms (0110) NO va acá (lo hace _reconcile_rooms_to_anchor).
    Idempotente: si ya está en el target, factor=1 / ya en 0, no cambia."""
    if not PL_BASELINE_FILE.exists():
        print(f"  ! falta {PL_BASELINE_FILE.name}; se omite reconciliación P&L")
        return
    cfg = json.loads(PL_BASELINE_FILE.read_text(encoding="utf-8"))
    d1 = date.fromisoformat(cfg["period"]["start"])
    d2 = date.fromisoformat(cfg["period"]["end"])
    depts = {d.cost_center: d.id for d in (await s.execute(
        select(Department).where(Department.property_id == pid))).scalars().all()}
    for g in cfg["groups"]:
        ids = [depts[c] for c in g["cost_centers"] if c in depts]
        if not ids:
            continue
        rows = (await s.execute(
            select(RevenueActualDaily).where(RevenueActualDaily.property_id == pid)
            .where(RevenueActualDaily.dept_id.in_(ids))
            .where(RevenueActualDaily.date >= d1).where(RevenueActualDaily.date <= d2)
            .order_by(RevenueActualDaily.date)
        )).scalars().all()
        if not rows:
            continue
        target = Decimal(str(g["target"]))
        cur = sum(Decimal(str(r.amount_usd)) for r in rows)
        if target == 0:
            for r in rows:
                r.amount_usd = 0
            continue
        if cur <= 0 or cur == target:
            continue
        f = target / cur
        acc = Decimal("0")
        for r in rows:
            nv = (Decimal(str(r.amount_usd)) * f).quantize(Decimal("0.01"))
            r.amount_usd = nv
            acc += nv
        rows[-1].amount_usd = Decimal(str(rows[-1].amount_usd)) + (target - acc)
    print("Base de arranque P&L (Ene-Jun) reconciliada por categoría")


async def bootstrap() -> None:
    async with SessionLocal() as s:
        if BUDGET_FILE.exists():
            result = await budget_service.upload_and_replace(
                s, YEAR, BUDGET_FILE.read_bytes(), property_code=settings.DEFAULT_PROPERTY)
            print(f"Budget {YEAR}: {result}")
        else:
            print(f"  ! falta {BUDGET_FILE.name}; se omite carga de presupuesto")

        if REVENUE_ACTUAL_FILE.exists():
            result = await revenue_actual_service.upload_daily_grid(
                s, REVENUE_ACTUAL_FILE.read_bytes(), property_code=settings.DEFAULT_PROPERTY)
            print(f"Revenue Actual: {result}")
        else:
            print(f"  ! falta {REVENUE_ACTUAL_FILE.name}; se omite carga de revenue actual")

        pid = await _property_id(s)
        if COMPSTAT_FILE.exists():
            await _load_comps_monthly(s, pid)
            print(f"Comps monthly (Tab 7.8): cargado desde {COMPSTAT_FILE.name}")
        else:
            print(f"  ! falta {COMPSTAT_FILE.name}; se omite carga de comps")

        if ROOM_STAT_OPENING_FILE.exists():
            await _load_room_stat_opening(s, pid)
            print(f"Room Stat Opening (ancla Tab 6.6): cargado desde {ROOM_STAT_OPENING_FILE.name}")
        else:
            print(f"  ! falta {ROOM_STAT_OPENING_FILE.name}; se omite carga del ancla")

        # Reconciliar el room revenue diario pre-corte al total del ancla (autoritativo).
        await _reconcile_rooms_to_anchor(s, pid)
        # Reconciliar las demás categorías Ene-Jun a la base de arranque del P&L.
        await _reconcile_categories_to_pl(s, pid)

        await s.commit()


if __name__ == "__main__":
    asyncio.run(bootstrap())
