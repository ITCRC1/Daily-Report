"""Seed de master data extraído de los golden files (§7).

Completa lo que el seed base dejó como TODO(bismark), leyendo directamente:
  - dim_department  <- hoja 'DEPT_MAP' (outlets 4-díg -> OutputColumn)
  - dim_payment_map <- hoja 'Mapping' del libro de Cash

Estrategia idempotente: borra las filas de la propiedad y reinserta (full refresh
de una tabla chica y controlada). NO inventa datos: si una hoja no está, avisa y sigue.

    python -m app.seed_from_goldens
"""
import asyncio
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete, select

from app.config import get_settings
from app.db import SessionLocal
from app.models import Department, PaymentMap, Property

settings = get_settings()

GOLDENS = Path(__file__).resolve().parents[2] / "goldens"
DAY31 = GOLDENS / "DAILY REV REP AS OF DAY 31.xlsm"
CASH = GOLDENS / "DAILY CASH POSITION MASTER FILE.xlsx"

# Deptos SINTÉTICOS (no son outlets físicos de ningún golden DEPT_MAP) usados
# por Tab 6.1 (Budget) y Tab 6.4 (Revenue Actual diario) para clasificar por
# NATURALEZA en vez de por outlet (§5.1a, misma apertura que engine/revenue.py
# WEEKLY_COLUMNS: Rooms/Rooms Others, F&B abierto en Food/Beverage/Misc,
# Sustainable Fee/Misc. Rev Others). Van hardcodeados acá (no en un Excel)
# para que viajen con el código a cualquier servidor/DB nueva -- si solo
# existieran como filas creadas a mano por API, un refresh de este seed en
# otro entorno los borraría sin dejar rastro.
SYNTHETIC_DEPARTMENTS = [
    {"cost_center": "FB-FOOD", "outlet_name": "F&B Food", "output_column": "F&B"},
    {"cost_center": "FB-BEV", "outlet_name": "F&B Beverage", "output_column": "F&B"},
    {"cost_center": "FB-MISC", "outlet_name": "F&B Misc", "output_column": "F&B"},
    {"cost_center": "ROOMS-OTH", "outlet_name": "Rooms Others", "output_column": "Rooms Others"},
    {"cost_center": "MISC-REV-OTH", "outlet_name": "Misc. Rev Others", "output_column": "Misc. Rev Others"},
    {"cost_center": "MISC-REV", "outlet_name": "Misc. Revenue", "output_column": "Otros"},
]


def read_dept_map() -> list[dict]:
    """Hoja DEPT_MAP: encuentra la fila header 'DeptCode' y lee las columnas."""
    if not DAY31.exists():
        print(f"  ! falta {DAY31.name}; se omite dim_department")
        return []
    wb = load_workbook(DAY31, read_only=True, data_only=True)
    ws = wb["DEPT_MAP"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    # localizar header
    hidx = next((i for i, r in enumerate(rows)
                 if r and any(str(c).strip() == "DeptCode" for c in r if c)), None)
    if hidx is None:
        print("  ! DEPT_MAP sin header 'DeptCode'")
        return []
    header = [str(c).strip() if c else "" for c in rows[hidx]]
    ci = {name: header.index(name) for name in ("DeptCode", "DeptName", "OutputColumn")
          if name in header}
    out = []
    for r in rows[hidx + 1:]:
        code = r[ci["DeptCode"]] if ci.get("DeptCode") is not None else None
        if code is None or str(code).strip() == "":
            continue
        out.append({
            "cost_center": str(code).strip(),
            "outlet_name": str(r[ci["DeptName"]]).strip() if ci.get("DeptName") is not None and r[ci["DeptName"]] else None,
            "output_column": str(r[ci["OutputColumn"]]).strip() if ci.get("OutputColumn") is not None and r[ci["OutputColumn"]] else None,
        })
    return out


def read_payment_map() -> list[dict]:
    """Hoja 'Mapping' del libro de Cash: header en la primera fila con 'Transaction Code'."""
    if not CASH.exists():
        print(f"  ! falta {CASH.name}; se omite dim_payment_map")
        return []
    wb = load_workbook(CASH, read_only=True, data_only=True)
    ws = wb["Mapping"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hidx = next((i for i, r in enumerate(rows)
                 if r and any(str(c).strip() == "Transaction Code" for c in r if c)), None)
    if hidx is None:
        print("  ! Mapping sin header 'Transaction Code'")
        return []
    header = [str(c).strip() if c else "" for c in rows[hidx]]

    # columnas del golden -> campos del modelo
    colmap = {
        "Code": "code", "Description": "description", "Transaction Code": "transaction_code",
        "Banco Código": "banco_codigo", "Banco Nombre": "banco_nombre", "Moneda": "moneda",
        "Tipo Pago": "tipo_pago", "Marca / Método": "marca_metodo", "Grupo": "grupo",
        "Cash Flow": "cash_flow", "Canal": "canal", "Report Bucket": "report_bucket",
    }
    idx = {header.index(k): v for k, v in colmap.items() if k in header}

    out, seen = [], set()
    for r in rows[hidx + 1:]:
        rec = {}
        for i, field in idx.items():
            val = r[i] if i < len(r) else None
            rec[field] = str(val).strip() if val is not None and str(val).strip() != "" else None
        tc = rec.get("transaction_code")
        if not tc:
            continue
        tc = str(tc).replace(".0", "")  # openpyxl a veces devuelve 3700.0
        rec["transaction_code"] = tc
        if tc in seen:
            print(f"  ! TCode duplicado en Mapping, se omite: {tc}")
            continue
        seen.add(tc)
        out.append(rec)
    return out


async def seed() -> None:
    depts = read_dept_map()
    pays = read_payment_map()

    golden_codes = {d["cost_center"] for d in depts}
    for synth in SYNTHETIC_DEPARTMENTS:
        if synth["cost_center"] in golden_codes:
            print(f"  ! {synth['cost_center']} ya existe en el golden DEPT_MAP, se omite el sintético")
            continue
        depts.append(synth)

    async with SessionLocal() as s:
        prop = (await s.execute(
            select(Property).where(Property.code == settings.DEFAULT_PROPERTY)
        )).scalar_one_or_none()
        if prop is None:
            raise SystemExit("Corré primero `python -m app.seed` (falta dim_property).")
        pid = prop.id

        # full refresh idempotente de estas dos dims para la propiedad
        await s.execute(delete(Department).where(Department.property_id == pid))
        await s.execute(delete(PaymentMap).where(PaymentMap.property_id == pid))

        for d in depts:
            s.add(Department(property_id=pid, cuenta_nature=None, **d))
        for p in pays:
            s.add(PaymentMap(property_id=pid, **p))

        await s.commit()

    print(f"Seed goldens OK: dim_department={len(depts)} outlets, dim_payment_map={len(pays)} TCodes.")


if __name__ == "__main__":
    asyncio.run(seed())
