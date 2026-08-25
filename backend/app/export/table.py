"""Export genérico a Excel — el motor que usan TODOS los tabs.

Un solo builder con el formato de la casa: cualquier tab manda la(s) tabla(s)
que está mostrando y recibe un .xlsx presentable, sin escribir un exportador
por pantalla. La alternativa era un exportador dedicado por tab (40+ tablas
entre tabs y sub-tabs), cada uno divergiendo de a poco del formato del resto.

Por qué el front manda la tabla en vez de que el back la recalcule: así el
Excel dice EXACTAMENTE lo que el usuario tenía en pantalla. Un exportador que
recalcula puede quedar desfasado de la vista y devolver dos cifras distintas
para lo mismo — el problema del Excel que este sistema vino a reemplazar. Acá
no se recalcula nada: se transcribe y se le da formato.

El formato numérico se aplica por columna (`type`), no celda por celda, así una
columna de dinero sale alineada y con el mismo formato en todas sus filas.
"""
from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# --- Paleta y formatos (una sola definición para todo el sistema) ------------
INK = "1F2430"
MUTED = "6B7280"
HEADER_BG = "2D3A5C"
GROUP_BG = "1F2942"
TOTAL_BG = "EEF1F6"
BAND_BG = "FAFAF9"
RULE = "D4D7DE"
NEGATIVE = "B4232C"

MONEY_FMT = '#,##0.00;[Red](#,##0.00)'
INT_FMT = '#,##0;[Red](#,##0)'
NUMBER_FMT = '#,##0.00;[Red](#,##0.00)'
PCT_FMT = '0.0%;[Red](0.0%)'
DATE_FMT = 'yyyy-mm-dd'

_FMT_BY_TYPE = {
    "money": MONEY_FMT,
    "int": INT_FMT,
    "number": NUMBER_FMT,
    "pct": PCT_FMT,
    "date": DATE_FMT,
    "text": None,
}
_NUMERIC_TYPES = {"money", "int", "number", "pct"}

_thin = Side(style="thin", color=RULE)
_medium = Side(style="medium", color=HEADER_BG)


def _sanitize_sheet_name(name: str, usados: set[str]) -> str:
    """Excel: máximo 31 caracteres, sin []:*?/\\ y sin nombres repetidos."""
    limpio = "".join("-" if ch in "[]:*?/\\" else ch for ch in (name or "Hoja")).strip() or "Hoja"
    limpio = limpio[:31]
    if limpio not in usados:
        usados.add(limpio)
        return limpio
    for i in range(2, 100):
        sufijo = f" ({i})"
        cand = limpio[: 31 - len(sufijo)] + sufijo
        if cand not in usados:
            usados.add(cand)
            return cand
    usados.add(limpio)
    return limpio


def _titulo(ws: Worksheet, row: int, title: str, subtitle: str | None, ancho: int) -> int:
    """Bloque de encabezado del reporte: título, subtítulo y sello de generación."""
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(bold=True, size=15, color=INK)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 22
    row += 1
    if subtitle:
        c = ws.cell(row=row, column=1, value=subtitle)
        c.font = Font(size=10, color=MUTED)
        row += 1
    sello = datetime.now(timezone.utc).strftime("Generado %Y-%m-%d %H:%M UTC · DAILY-OPS")
    ws.cell(row=row, column=1, value=sello).font = Font(size=8, italic=True, color=MUTED)
    return row + 2


def _grupos(ws: Worksheet, row: int, groups: list[dict], n_cols: int) -> int:
    """Fila de encabezados agrupados (ej. TODAY / MONTH TO DAY / FULL MONTH
    RESULT), con merge por span. Refleja el `colSpan` de la tabla en pantalla."""
    col = 1
    for g in groups:
        span = max(1, int(g.get("span") or 1))
        label = g.get("label") or ""
        fin = min(col + span - 1, n_cols)
        if span > 1:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=fin)
        for c_idx in range(col, fin + 1):
            cell = ws.cell(row=row, column=c_idx)
            cell.fill = PatternFill("solid", fgColor=GROUP_BG)
            cell.border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = Font(bold=True, size=9, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        col = fin + 1
        if col > n_cols:
            break
    ws.row_dimensions[row].height = 18
    return row + 1


def _encabezado(ws: Worksheet, row: int, columns: list[dict]) -> int:
    for i, col in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=i, value=col.get("label") or "")
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.font = Font(bold=True, size=9, color="FFFFFF")
        cell.alignment = Alignment(
            horizontal="right" if col.get("type") in _NUMERIC_TYPES else "left",
            vertical="center", wrap_text=True,
        )
        cell.border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    ws.row_dimensions[row].height = 26
    return row + 1


def _ancho(col: dict, valores: list) -> float:
    """Ancho por contenido, acotado. Si el tab lo manda explícito, se respeta."""
    if col.get("width"):
        return float(col["width"])
    largo = len(str(col.get("label") or ""))
    for v in valores[:200]:
        if v is None:
            continue
        texto = f"{v:,.2f}" if isinstance(v, (int, float)) else str(v)
        largo = max(largo, len(texto))
    return min(46.0, max(9.0, largo + 2.4))


def _hoja(wb: Workbook, spec: dict, title: str, subtitle: str | None, usados: set[str]) -> None:
    columns: list[dict] = spec.get("columns") or []
    rows: list[list] = spec.get("rows") or []
    if not columns:
        columns = [{"label": "", "type": "text"}]
    n_cols = len(columns)

    ws = wb.create_sheet(_sanitize_sheet_name(spec.get("name") or "Hoja", usados))
    ws.sheet_view.showGridLines = False

    row = _titulo(ws, 1, spec.get("title") or title, spec.get("subtitle") or subtitle, n_cols)

    caption = spec.get("caption")
    if caption:
        c = ws.cell(row=row, column=1, value=caption)
        c.font = Font(bold=True, size=11, color=INK)
        row += 1

    groups = spec.get("header_groups") or []
    if groups:
        row = _grupos(ws, row, groups, n_cols)
    fila_encabezado = row
    row = _encabezado(ws, row, columns)
    primera_fila = row

    totales = {int(i) for i in (spec.get("total_rows") or [])}
    for r_idx, fila in enumerate(rows):
        es_total = r_idx in totales
        for c_idx, col in enumerate(columns, start=1):
            valor = fila[c_idx - 1] if c_idx - 1 < len(fila) else None
            cell = ws.cell(row=row, column=c_idx, value=valor)
            tipo = col.get("type") or "text"
            fmt = _FMT_BY_TYPE.get(tipo)
            if fmt and isinstance(valor, (int, float)):
                cell.number_format = fmt
            cell.alignment = Alignment(
                horizontal="right" if tipo in _NUMERIC_TYPES else "left", vertical="center",
            )
            if es_total:
                cell.font = Font(bold=True, size=10, color=INK)
                cell.fill = PatternFill("solid", fgColor=TOTAL_BG)
                cell.border = Border(top=_medium, bottom=_thin)
            else:
                cell.font = Font(size=10, color=INK)
                if r_idx % 2 == 1:
                    cell.fill = PatternFill("solid", fgColor=BAND_BG)
                cell.border = Border(bottom=_thin)
        row += 1

    for i, col in enumerate(columns, start=1):
        valores = [f[i - 1] if i - 1 < len(f) else None for f in rows]
        ws.column_dimensions[get_column_letter(i)].width = _ancho(col, valores)

    # Congelar encabezado + la primera columna (la de etiquetas), que es lo que
    # hace legible una tabla ancha de 12 meses o 11 columnas de varianza.
    ws.freeze_panes = f"B{primera_fila}"

    notas = spec.get("notes") or []
    if notas:
        row += 1
        for nota in notas:
            ws.cell(row=row, column=1, value=nota).font = Font(size=8, italic=True, color=MUTED)
            row += 1

    # Impresión: horizontal, ajustado al ancho de la página y repitiendo el
    # encabezado en cada hoja — para que imprimir no requiera tocar nada.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{fila_encabezado}:{fila_encabezado}"
    ws.print_options.horizontalCentered = True


def build_workbook(payload: dict) -> bytes:
    """`payload`: {title, subtitle, sheets:[{name, columns, rows, ...}]}.

    `columns`: [{label, type: text|money|pct|int|number|date, width?}]
    `rows`: lista de listas, en el mismo orden que `columns`.
    `total_rows`: índices (base 0) de las filas a resaltar como totales.
    """
    title = payload.get("title") or "DAILY-OPS"
    subtitle = payload.get("subtitle")
    hojas = payload.get("sheets") or []

    wb = Workbook()
    wb.remove(wb.active)
    usados: set[str] = set()
    for spec in hojas:
        _hoja(wb, spec, title, subtitle, usados)
    if not wb.sheetnames:  # payload sin hojas: no devolver un archivo corrupto
        ws = wb.create_sheet("Sin datos")
        ws["A1"] = "No hay datos para exportar."

    wb.properties.title = title
    wb.properties.creator = "DAILY-OPS"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
