"""Export to Excel (.xlsx) for the day -- stage 8. Three sheets: Revenue, Cash, Audit.

Pure function: receives the payloads ALREADY computed by the services (revenue_service.
daily_report, cash_service.daily_cash, audit_service.get_audit) and builds the workbook.
Never recomputes anything -- single source of truth, the same one the pages see.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="2D3A5C")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FILL = PatternFill("solid", fgColor="1A7F4B")
TOTAL_FONT = Font(bold=True, color="FFFFFF")
MONEY_FMT = "#,##0.00;(#,##0.00)"
PCT_FMT = "0.0%"


def _header_row(ws, row: int, headers: list[str]):
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")


def _autowidth(ws, widths: list[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _kpi_block(ws, row: int, title: str, kpis: list[tuple[str, str]]):
    """`kpis`: [(label, formatted_value)]."""
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(bold=True, size=13)
    row += 1
    for label, value in kpis:
        ws.cell(row=row, column=1, value=label).font = Font(color="666666")
        ws.cell(row=row, column=2, value=value).font = Font(bold=True)
        row += 1
    return row + 1


def _revenue_sheet(wb: Workbook, revenue: dict, business_date: str):
    ws = wb.create_sheet("Revenue")
    ws.sheet_view.showGridLines = False
    row = _kpi_block(ws, 1, f"Daily Revenue Report — {business_date}", [
        ("Total Pax", str(revenue["kpis"]["pax"])),
        ("Rooms Occupied", str(revenue["kpis"]["rooms_occupied"])),
        ("Available", str(revenue["kpis"]["available_rooms"])),
        ("ADR", f"${revenue['kpis']['adr']:,.2f}"),
        ("Occupancy %", f"{revenue['kpis']['occupancy_pct'] * 100:.1f}%"),
    ])

    _header_row(ws, row, ["Revenue Center", "Today Actual", "Today Budget",
                          "MTD Actual", "MTD Budget"])
    row += 1
    start_data = row
    for c in revenue["centers"]:
        ws.cell(row=row, column=1, value=c["center"])
        ws.cell(row=row, column=2, value=c["today_actual"]).number_format = MONEY_FMT
        ws.cell(row=row, column=3, value=c["today_budget"]).number_format = MONEY_FMT
        ws.cell(row=row, column=4, value=c["mtd_actual"]).number_format = MONEY_FMT
        ws.cell(row=row, column=5, value=c["mtd_budget"]).number_format = MONEY_FMT
        row += 1
    ws.cell(row=row, column=1, value="GRAND TOTAL").font = TOTAL_FONT
    ws.cell(row=row, column=1).fill = TOTAL_FILL
    gt = revenue["grand_total"]
    for col, val in zip((2, 3, 4, 5), (gt["today_actual"], gt["today_budget"],
                                       gt["mtd_actual"], gt["mtd_budget"])):
        cell = ws.cell(row=row, column=col, value=val)
        cell.number_format = MONEY_FMT
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
    row += 2

    _header_row(ws, row, ["Category (Room Stats)", "RN", "Pax", "Available",
                          "Revenue", "Occ %", "ADR", "Yield Index"])
    row += 1
    for rc in revenue.get("room_categories", []):
        ws.cell(row=row, column=1, value=rc["category"])
        ws.cell(row=row, column=2, value=rc["stay_rooms"])
        ws.cell(row=row, column=3, value=rc["stay_persons"])
        ws.cell(row=row, column=4, value=rc["physical_rooms"])
        ws.cell(row=row, column=5, value=rc["revenue"]).number_format = MONEY_FMT
        ws.cell(row=row, column=6, value=rc["occupancy_pct"]).number_format = PCT_FMT
        ws.cell(row=row, column=7, value=rc["adr"]).number_format = MONEY_FMT
        ws.cell(row=row, column=8, value=rc["yield_index"])
        row += 1

    if revenue.get("otros"):
        row += 1
        ws.cell(row=row, column=1, value="⚠ Other (outside the canonical map)").font = Font(bold=True, color="C0392B")
        row += 1
        _header_row(ws, row, ["Account", "Name", "Amount"])
        row += 1
        for o in revenue["otros"]:
            ws.cell(row=row, column=1, value=o["cuenta"])
            ws.cell(row=row, column=2, value=o.get("nombre"))
            ws.cell(row=row, column=3, value=o["amount"]).number_format = MONEY_FMT
            row += 1

    _autowidth(ws, [26, 16, 16, 16, 16, 12, 12, 12])
    ws.freeze_panes = f"A{start_data}"


def _cash_sheet(wb: Workbook, cash: dict, business_date: str):
    ws = wb.create_sheet("Cash")
    ws.sheet_view.showGridLines = False
    today = cash["today"]
    row = _kpi_block(ws, 1, f"Daily Cash from Operation — {business_date}", [
        ("Real Cash", f"${today['real_cash']:,.2f}"),
        ("Non-Cash", f"${today['non_cash']:,.2f}"),
        ("Cash-relevant (broad)", f"${today['cash_relevant_total']:,.2f}"),
        ("Bank-only (strict)", f"${today['bank_only_total']:,.2f}"),
    ])

    for title, key in (("By Bucket", "by_bucket"), ("By Bank", "by_bank"),
                       ("By Brand/Method", "by_brand"), ("By Channel", "by_channel")):
        ws.cell(row=row, column=1, value=title).font = Font(bold=True)
        row += 1
        _header_row(ws, row, ["", "Amount"])
        row += 1
        for k, v in sorted(today[key].items(), key=lambda kv: -kv[1]):
            ws.cell(row=row, column=1, value=k)
            ws.cell(row=row, column=2, value=v).number_format = MONEY_FMT
            row += 1
        row += 1

    if cash.get("unmapped_today"):
        ws.cell(row=row, column=1, value="⚠ Unmapped payment TCodes (UNMAPPED)").font = Font(bold=True, color="C0392B")
        row += 1
        _header_row(ws, row, ["TCode", "Description", "Opera Total"])
        row += 1
        for u in cash["unmapped_today"]:
            ws.cell(row=row, column=1, value=u["tcode"])
            ws.cell(row=row, column=2, value=u.get("description"))
            ws.cell(row=row, column=3, value=u["opera_total"]).number_format = MONEY_FMT
            row += 1

    _autowidth(ws, [26, 16])


_STATUS_LABEL_EN = {"abierto": "Open", "cerrado": "Closed"}
_ESTADO_LABEL_EN = {
    "OK": "OK", "DISCREPANCIA": "DISCREPANCY", "FALTA EN INTEGRITY": "MISSING IN INTEGRITY",
    "FALTA EN OPERA": "MISSING IN OPERA", "INTERNO": "INTERNAL", "DIF_OPERATIVA": "OPERATIONAL DIFF",
}


def _audit_sheet(wb: Workbook, audit: dict, business_date: str):
    ws = wb.create_sheet("Audit")
    ws.sheet_view.showGridLines = False
    k = audit["kpis"]
    status = audit.get("status") or "abierto"
    row = _kpi_block(ws, 1, f"Daily Audit — {business_date}", [
        ("Status", _STATUS_LABEL_EN.get(status, status)),
        ("Reconciled", str(k["ok"])),
        ("Discrepancies", str(k["discrepancia"])),
        ("Missing", str(k["faltante"])),
        ("Internal", str(k["interno"])),
        ("Gate", audit["gate"]["reason"]),
    ])

    _header_row(ws, row, ["TCode", "Description", "Type", "Opera", "Integrity", "Difference", "Status"])
    row += 1
    start_data = row
    for r in audit["rows"]:
        ws.cell(row=row, column=1, value=r["tcode"])
        ws.cell(row=row, column=2, value=r["description"])
        ws.cell(row=row, column=3, value=r["type"])
        for col, val in ((4, r["opera"]), (5, r["integrity"]), (6, r["diferencia"])):
            cell = ws.cell(row=row, column=col, value=val)
            if val is not None:
                cell.number_format = MONEY_FMT
        estado_cell = ws.cell(row=row, column=7, value=_ESTADO_LABEL_EN.get(r["estado"], r["estado"]))
        if r["estado"] == "DISCREPANCIA":
            estado_cell.font = Font(color="C0392B", bold=True)
        elif r["estado"] == "OK":
            estado_cell.font = Font(color="1A7F4B")
        row += 1

    _autowidth(ws, [10, 34, 14, 14, 14, 14, 16])
    ws.freeze_panes = f"A{start_data}"


def build_daily_excel(revenue: dict, cash: dict, audit: dict,
                      business_date: str, property_code: str) -> bytes:
    """Arma el workbook de las 3 hojas y devuelve los bytes del .xlsx."""
    wb = Workbook()
    wb.remove(wb.active)  # la hoja default en blanco
    _revenue_sheet(wb, revenue, business_date)
    _cash_sheet(wb, cash, business_date)
    _audit_sheet(wb, audit, business_date)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
