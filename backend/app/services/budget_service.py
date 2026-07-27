"""Budget (Tab 6.1 Monthly by Department + 6.5 Daily derivado, §2 decisión
cerrada #2): descarga plantilla Excel (prellenada con lo ya cargado) →
Bismark la llena offline → sube de vuelta → reemplazo total del año
("reset anual") → se deriva `fact_budget` día por día (§3, engine/budget.py).

No se adoptó el layout viejo/transpuesto de '12 months Budget 06' (bloqueante
documentado) — esta es una plantilla nueva, propia de la app, mapeada 1:1 al
esquema de budget_monthly. No hace falta el archivo legado para esto.
"""
from __future__ import annotations

import calendar
from datetime import date as date_cls
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.engine.budget import derive_daily_amounts
from app.models import Budget, BudgetMonthly, Department, Property

HEADERS = ["Dept Code", "Dept Name", "Mes", "Amount USD", "Available Rooms",
          "Rooms Occupied", "Guests", "Occupancy %", "ADR", "Food", "Beverage", "Misc"]


async def _property_id(session: AsyncSession, code: str):
    pid = (await session.execute(
        select(Property.id).where(Property.code == code)
    )).scalar_one_or_none()
    if pid is None:
        raise ValueError(f"Propiedad '{code}' no existe.")
    return pid


async def _departments(session: AsyncSession, pid) -> list[Department]:
    return (await session.execute(
        select(Department).where(Department.property_id == pid).order_by(Department.cost_center)
    )).scalars().all()


def _n(v) -> float | None:
    return float(v) if v is not None else None


async def build_template(session: AsyncSession, year: int, property_code: str = "COWLCR") -> bytes:
    """Excel prellenado con lo que ya está cargado para `year` (o en cero si no
    hay nada) — un renglón por departamento × mes."""
    pid = await _property_id(session, property_code)
    depts = await _departments(session, pid)
    existing = (await session.execute(
        select(BudgetMonthly).where(BudgetMonthly.property_id == pid, BudgetMonthly.year == year)
    )).scalars().all()
    by_key = {(str(b.dept_id), b.month): b for b in existing}

    wb = Workbook()
    ws = wb.active
    ws.title = f"Budget {year}"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2D3A5C")

    for d in depts:
        for month in range(1, 13):
            b = by_key.get((str(d.id), month))
            ws.append([
                d.cost_center, d.outlet_name, month,
                float(b.amount_usd) if b else 0.0,
                _n(b.available_rooms) if b else None, _n(b.rooms_occupied) if b else None,
                _n(b.guests) if b else None, _n(b.occupancy_pct) if b else None,
                _n(b.adr) if b else None,
                float(b.food) if b else 0.0, float(b.beverage) if b else 0.0,
                float(b.misc) if b else 0.0,
            ])
    ws.freeze_panes = "A2"
    for i, w in enumerate([11, 28, 6, 13, 15, 15, 9, 12, 10, 11, 11, 11], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def upload_and_replace(session: AsyncSession, year: int, file_bytes: bytes,
                             property_code: str = "COWLCR") -> dict:
    """Sube la plantilla llena: reemplazo total del año (budget_monthly +
    fact_budget derivado) — 'reset anual', §2 decisión cerrada #2."""
    pid = await _property_id(session, property_code)
    depts = await _departments(session, pid)
    by_code = {d.cost_center: d for d in depts}

    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    await session.execute(delete(BudgetMonthly).where(
        BudgetMonthly.property_id == pid, BudgetMonthly.year == year))
    await session.execute(delete(Budget).where(
        Budget.property_id == pid,
        Budget.date >= date_cls(year, 1, 1), Budget.date <= date_cls(year, 12, 31)))

    loaded, skipped = 0, []
    for r in rows:
        if not r or r[0] is None:
            continue
        (dept_code, _name, month, amount_usd, avail, occ_rooms, guests,
         occ_pct, adr, food, bev, misc) = (list(r) + [None] * 12)[:12]
        dept = by_code.get(str(dept_code))
        if dept is None or month is None:
            skipped.append(str(dept_code))
            continue
        month = int(month)
        amount = Decimal(str(amount_usd or 0))

        session.add(BudgetMonthly(
            property_id=pid, year=year, month=month, dept_id=dept.id, amount_usd=amount,
            available_rooms=Decimal(str(avail)) if avail not in (None, "") else None,
            rooms_occupied=Decimal(str(occ_rooms)) if occ_rooms not in (None, "") else None,
            guests=Decimal(str(guests)) if guests not in (None, "") else None,
            occupancy_pct=Decimal(str(occ_pct)) if occ_pct not in (None, "") else None,
            adr=Decimal(str(adr)) if adr not in (None, "") else None,
            food=Decimal(str(food or 0)), beverage=Decimal(str(bev or 0)), misc=Decimal(str(misc or 0)),
        ))
        loaded += 1

        if amount:
            for day_idx, amt in enumerate(derive_daily_amounts(amount, year, month), start=1):
                session.add(Budget(
                    property_id=pid, date=date_cls(year, month, day_idx),
                    dept_id=dept.id, amount_usd=amt,
                ))

    await session.commit()
    return {"year": year, "rows_loaded": loaded, "dept_codes_no_reconocidos": sorted(set(skipped))}


async def monthly_summary(session: AsyncSession, year: int, property_code: str = "COWLCR") -> list[dict]:
    pid = await _property_id(session, property_code)
    depts = {d.id: d for d in await _departments(session, pid)}
    rows = (await session.execute(
        select(BudgetMonthly).where(BudgetMonthly.property_id == pid, BudgetMonthly.year == year)
        .order_by(BudgetMonthly.month)
    )).scalars().all()
    return [{
        "dept_code": depts[r.dept_id].cost_center if r.dept_id in depts else None,
        "dept_name": depts[r.dept_id].outlet_name if r.dept_id in depts else None,
        "month": r.month, "amount_usd": float(r.amount_usd),
        "available_rooms": _n(r.available_rooms), "rooms_occupied": _n(r.rooms_occupied),
        "guests": _n(r.guests), "occupancy_pct": _n(r.occupancy_pct), "adr": _n(r.adr),
        "food": float(r.food), "beverage": float(r.beverage), "misc": float(r.misc),
    } for r in rows]


async def daily_summary(session: AsyncSession, year: int, month: int,
                        property_code: str = "COWLCR") -> list[dict]:
    """6.5 — presupuesto diario derivado (fact_budget) de un mes."""
    pid = await _property_id(session, property_code)
    depts = {d.id: d for d in await _departments(session, pid)}
    days_in_month = calendar.monthrange(year, month)[1]
    rows = (await session.execute(
        select(Budget).where(
            Budget.property_id == pid,
            Budget.date >= date_cls(year, month, 1),
            Budget.date <= date_cls(year, month, days_in_month),
        ).order_by(Budget.date)
    )).scalars().all()
    return [{
        "date": r.date.isoformat(),
        "dept_code": depts[r.dept_id].cost_center if r.dept_id in depts else None,
        "dept_name": depts[r.dept_id].outlet_name if r.dept_id in depts else None,
        "amount_usd": float(r.amount_usd),
    } for r in rows]
