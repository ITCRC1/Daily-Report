"""Tab 9 · Daily Extendido — 9.1 Summary (réplica pág. 2 del formato).

Deriva de lo que Daily Ops ya tiene: rooms stats (occupancy_stat + capacidad
fija) y revenue por categoría (fact_revenue_actual_daily) + budget (fact_budget),
para Today y MTD. Forecast y Año anterior quedan en None (vienen de la carga).
"""
from __future__ import annotations

import calendar
from datetime import date as date_cls
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from sqlalchemy import delete, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import (
    AppConfig, Budget, Department, FbCovers, IntegrityLine, OccupancyStat, Property,
)
from app.services import market_code_service, revenue_service

NON_REVENUE_MARKET_CODES = {"COM", "INHOUSE"}

# cost_center -> categoría del formato (FS). ROOMS_OTH y las variantes van con su grupo.
CAT_OF = {
    "0110": "rooms", "ROOMS-OTH": "rooms",
    "FB-FOOD": "food", "FB-BEV": "beverage", "FB-MISC": "fb_misc",
    "0140": "spa",
    "0150": "activities", "0152": "activities",  # Guest Support & Activities (Tours + Transport)
    "0151": "misc", "0160": "misc", "0170": "misc", "0155": "misc", "0156": "misc",
    "MISC-REV": "misc", "MISC-REV-OTH": "misc",
}


async def _pid(session, code):
    pid = (await session.execute(select(Property.id).where(Property.code == code))).scalar_one_or_none()
    if pid is None:
        raise ValueError(f"Propiedad '{code}' no existe.")
    return pid


async def _total_rooms(session, pid) -> int:
    row = (await session.execute(
        select(AppConfig).where(AppConfig.property_id == pid, AppConfig.key == "total_rooms")
    )).scalar_one_or_none()
    try:
        return int(row.value) if row else 30
    except Exception:  # noqa: BLE001
        return 30


async def _dept_cat(session, pid) -> dict:
    depts = (await session.execute(select(Department).where(Department.property_id == pid))).scalars().all()
    return {d.id: CAT_OF.get(d.cost_center) for d in depts}


async def _budget_by_cat(session, pid, dept_cat, start, end) -> dict:
    """Budget por categoría (fact_budget, por dept_id)."""
    rows = (await session.execute(
        select(Budget.dept_id, func.coalesce(func.sum(Budget.amount_usd), 0))
        .where(Budget.property_id == pid, Budget.date >= start, Budget.date <= end)
        .group_by(Budget.dept_id)
    )).all()
    acc: dict[str, float] = {}
    for dept_id, amt in rows:
        cat = dept_cat.get(dept_id)
        if cat:
            acc[cat] = acc.get(cat, 0.0) + float(amt or 0)
    return acc


async def _actual_by_cat(session, pid, start, end) -> dict:
    """Revenue actual por categoría desde el motor sobre Integrity (fuente real,
    igual que Tab 3/4) con respaldo de la grilla 6.4. dept_code -> categoría FS."""
    rows = await revenue_service.merged_revenue_actual(session, pid, start, end)
    acc: dict[str, float] = {}
    for r in rows:
        cat = CAT_OF.get(r.get("dept_code"))
        if cat:
            acc[cat] = acc.get(cat, 0.0) + float(r.get("amount_usd") or 0)
    return acc


async def _rooms_stats(session, pid, start, end) -> dict:
    """Occupied (net de comps) + Pax del rango, de occupancy_stat."""
    rows = (await session.execute(
        select(OccupancyStat.market_code,
               func.coalesce(func.sum(OccupancyStat.rooms), 0),
               func.coalesce(func.sum(OccupancyStat.persons), 0))
        .where(OccupancyStat.property_id == pid,
               OccupancyStat.business_date >= start, OccupancyStat.business_date <= end)
        .group_by(OccupancyStat.market_code)
    )).all()
    occ = pax = 0
    for mc, r, p in rows:
        if (mc or "").upper() in NON_REVENUE_MARKET_CODES:
            continue
        occ += int(r or 0); pax += int(p or 0)
    return {"occupied": occ, "pax": pax}


def _r(v):
    return None if v is None else round(v, 2)


def _col(today, mtd_actual, mtd_budget=None, month_budget=None):
    """Columna de las tablas del Tab 9 (9.1 y 9.6).

    `mtd_budget` es el presupuesto PRORRATEADO a los días corridos del mes
    (comparable 1:1 contra `mtd_actual`); `month_budget` es el del MES COMPLETO,
    para ver cuánto falta para cerrarlo.

    `amount_to_budget` = MTD actual − budget del mes completo (negativo = falta),
    misma convención que el bloque "Full Month Result" de Tab 3.

    OJO: esta función estaba DUPLICADA en el módulo (una versión más abajo, con
    redondeo, que por ser la última pisaba silenciosamente a la de arriba para
    TODOS los llamadores). Se unificaron acá; no volver a definir otra.
    """
    d = {"today": _r(today), "mtd_actual": _r(mtd_actual), "mtd_budget": _r(mtd_budget),
         "mtd_forecast": None, "mtd_ly": None,
         "month_budget": _r(month_budget), "amount_to_budget": None, "monthly_var_pct": None}
    if month_budget is not None:
        a2b = round((mtd_actual or 0) - month_budget, 2)
        d["amount_to_budget"] = a2b
        d["monthly_var_pct"] = round(a2b / month_budget, 4) if month_budget else None
    return d


async def summary(session: AsyncSession, business_date: date_cls, property_code: str = "COWLCR") -> dict:
    pid = await _pid(session, property_code)
    total_rooms = await _total_rooms(session, pid)
    dept_cat = await _dept_cat(session, pid)

    month_start = date_cls(business_date.year, business_date.month, 1)
    days_mtd = business_date.day
    avail_today, avail_mtd = total_rooms, total_rooms * days_mtd

    rs_today = await _rooms_stats(session, pid, business_date, business_date)
    rs_mtd = await _rooms_stats(session, pid, month_start, business_date)

    rev_today = await _actual_by_cat(session, pid, business_date, business_date)
    rev_mtd = await _actual_by_cat(session, pid, month_start, business_date)
    bud_mtd = await _budget_by_cat(session, pid, dept_cat, month_start, business_date)
    bud_today = await _budget_by_cat(session, pid, dept_cat, business_date, business_date)  # noqa: F841
    # Budget del MES COMPLETO (1 -> último día): el MTD prorratea a los días
    # corridos, así que contra el mes entero se ve cuánto falta para cerrarlo.
    days_in_month = calendar.monthrange(business_date.year, business_date.month)[1]
    month_end = date_cls(business_date.year, business_date.month, days_in_month)
    bud_month = await _budget_by_cat(session, pid, dept_cat, month_start, month_end)

    def cat(d, k): return round(d.get(k, 0.0), 2)
    def fb(d): return round(d.get("food", 0) + d.get("beverage", 0) + d.get("fb_misc", 0), 2)
    def total_rev(d): return round(sum(d.values()), 2)

    def occ_pct(o, a): return round(o / a, 4) if a else 0.0
    def adr(rev, o): return round(rev / o, 2) if o else 0.0

    rooms_rev_today, rooms_rev_mtd = cat(rev_today, "rooms"), cat(rev_mtd, "rooms")
    o_t, o_m = rs_today["occupied"], rs_mtd["occupied"]

    rooms_stats = {
        "available": _col(avail_today, avail_mtd, avail_mtd, total_rooms * days_in_month),
        "occupied": _col(o_t, o_m, None),
        "occupancy": _col(occ_pct(o_t, avail_today), occ_pct(o_m, avail_mtd), None),
        "adr": _col(adr(rooms_rev_today, o_t), adr(rooms_rev_mtd, o_m), None),
        "revpar": _col(round(rooms_rev_today / avail_today, 2) if avail_today else 0,
                       round(rooms_rev_mtd / avail_mtd, 2) if avail_mtd else 0, None),
    }

    def revrow(k): return _col(cat(rev_today, k), cat(rev_mtd, k), cat(bud_mtd, k), cat(bud_month, k))
    revenue = {
        "rooms": _col(rooms_rev_today, rooms_rev_mtd, cat(bud_mtd, "rooms"), cat(bud_month, "rooms")),
        "food": revrow("food"), "beverage": revrow("beverage"), "fb_misc": revrow("fb_misc"),
        "total_fb": _col(fb(rev_today), fb(rev_mtd), fb(bud_mtd), fb(bud_month)),
        "spa": revrow("spa"),
        "telecom": _col(0.0, 0.0, 0.0, 0.0),
        "activities": revrow("activities"),
        "misc": revrow("misc"),
        "total": _col(total_rev(rev_today), total_rev(rev_mtd), total_rev(bud_mtd), total_rev(bud_month)),
    }

    add_stats = {
        "guests": _col(rs_today["pax"], rs_mtd["pax"], None),
        "double_occupancy": _col(round(rs_today["pax"] / o_t, 4) if o_t else 0,
                                 round(rs_mtd["pax"] / o_m, 4) if o_m else 0, None),
    }

    return {
        "business_date": business_date.isoformat(), "days_mtd": days_mtd,
        "rooms_stats": rooms_stats, "revenue": revenue, "add_stats": add_stats,
    }


# ---------------------------------------------------------------------------
# 9.3 Rooms by Segment (réplica pág. 5-6) — RN · Occ% · Rev · ADR por segmento
# ---------------------------------------------------------------------------
# Reutiliza el pivote del Tab 7.10 (market_code_service): RN y Pax del XML
# STATISTICS, Room Revenue (Accommodation) del XML Revenue. Aquí se agrega Occ%
# (RN del segmento / habitaciones disponibles del período) y ADR (Rev / RN), en
# columnas Today y MTD, agrupado por grupo de negocio (Direct/OTA/TA/Groups).

GROUP_MAP = market_code_service.GROUP_MAP
GROUP_ORDER = market_code_service.GROUP_ORDER


def _seg_metrics(rn: int, rev: float, avail: int) -> dict:
    return {
        "rn": rn,
        "rev": round(rev, 2),
        "occ": round(rn / avail, 4) if avail else 0.0,
        "adr": round(rev / rn, 2) if rn else 0.0,
    }


async def rooms_by_segment(session: AsyncSession, business_date: date_cls,
                           property_code: str = "COWLCR") -> dict:
    pid = await _pid(session, property_code)
    total_rooms = await _total_rooms(session, pid)

    month_start = date_cls(business_date.year, business_date.month, 1)
    days_mtd = business_date.day
    avail_today, avail_mtd = total_rooms, total_rooms * days_mtd

    rep_today = await market_code_service.market_code_report(
        session, property_code=property_code, date_from=business_date, date_to=business_date)
    rep_mtd = await market_code_service.market_code_report(
        session, property_code=property_code, date_from=month_start, date_to=business_date)

    # Índice por market_code, uniendo Today y MTD (§10: no se pierde ningún código).
    def index(rep):
        return {r["market_code"]: r for r in rep["rows"]}
    t_idx, m_idx = index(rep_today), index(rep_mtd)
    codes = set(t_idx) | set(m_idx)

    def group_of(code: str) -> str:
        r = m_idx.get(code) or t_idx.get(code) or {}
        # Prioriza el rollup de negocio (GROUP_MAP); si el código no está, usa el
        # kpi_group del catálogo; si tampoco, "Other".
        g = GROUP_MAP.get((code or "").upper())
        if g:
            return g
        mg = r.get("market_group")
        return mg if mg and mg != "Unmapped" else "Other"

    seg_rows: list[dict] = []
    for code in codes:
        t = t_idx.get(code, {})
        m = m_idx.get(code, {})
        seg_rows.append({
            "market_code": code,
            "description": (m.get("description") or t.get("description")),
            "group": group_of(code),
            "today": _seg_metrics(t.get("rooms", 0), t.get("room_revenue", 0.0), avail_today),
            "mtd": _seg_metrics(m.get("rooms", 0), m.get("room_revenue", 0.0), avail_mtd),
        })

    # Agrupar por grupo de negocio con subtotales.
    order = [*GROUP_ORDER, "Other"]
    groups = []
    for g in order:
        members = [s for s in seg_rows if s["group"] == g]
        if not members:
            continue
        members.sort(key=lambda x: (-x["mtd"]["rev"], x["market_code"]))
        t_rn = sum(s["today"]["rn"] for s in members)
        t_rev = sum(s["today"]["rev"] for s in members)
        m_rn = sum(s["mtd"]["rn"] for s in members)
        m_rev = sum(s["mtd"]["rev"] for s in members)
        groups.append({
            "group": g,
            "segments": members,
            "subtotal": {
                "today": _seg_metrics(t_rn, t_rev, avail_today),
                "mtd": _seg_metrics(m_rn, m_rev, avail_mtd),
            },
        })

    tot_t_rn = sum(s["today"]["rn"] for s in seg_rows)
    tot_t_rev = sum(s["today"]["rev"] for s in seg_rows)
    tot_m_rn = sum(s["mtd"]["rn"] for s in seg_rows)
    tot_m_rev = sum(s["mtd"]["rev"] for s in seg_rows)

    return {
        "business_date": business_date.isoformat(), "days_mtd": days_mtd,
        "available": {"today": avail_today, "mtd": avail_mtd},
        "groups": groups,
        "total": {
            "today": _seg_metrics(tot_t_rn, tot_t_rev, avail_today),
            "mtd": _seg_metrics(tot_m_rn, tot_m_rev, avail_mtd),
        },
    }


# ---------------------------------------------------------------------------
# 9.2 Revenue Detail (réplica pág. 3-4) — revenue por outlet, Today + MTD + Budget
# ---------------------------------------------------------------------------
# Detalle a nivel de OUTLET (dept_code) de lo que 9.1 muestra colapsado por
# categoría. Fuente: mismo motor sobre Integrity (`merged_revenue_actual`, igual
# que Tab 3/4/9.1) + budget por cost_center (`_budget_by_dept_cost_center`). Las
# secciones espejan los centros de ingreso del formato FS. Cualquier dept_code
# que devuelva el motor y no esté mapeado cae en "Other" (§10: no se pierde nada).

# (sección, [(dept_code, etiqueta)]) — orden y nombres del reporte.
REVENUE_DETAIL_SECTIONS: list[tuple[str, list[tuple[str, str]]]] = [
    ("Rooms", [
        ("0110", "Room Revenue"),
        ("ROOMS-OTH", "Other Rooms Revenue"),
    ]),
    ("Food & Beverage", [
        ("FB-FOOD", "Food"),
        ("FB-BEV", "Beverage"),
        ("FB-MISC", "F&B Other"),
    ]),
    ("Spa & Wellness", [
        ("0140", "Spa & Fitness"),
    ]),
    ("Guest Support & Activities", [
        ("0150", "Tours & Activities"),
        ("0152", "Transportation"),
    ]),
    ("Other Operating Departments", [
        ("0151", "Retail / Gift Shop"),
        ("0160", "Laundry"),
        ("0155", "Innoceana"),
        ("0156", "Crowther Lab"),
        ("0170", "Sustainable Fee"),
        ("MISC-REV", "Miscellaneous Revenue"),
        ("MISC-REV-OTH", "Misc. Revenue - Other"),
    ]),
]


async def _rev_by_dept(session, pid, start, end) -> dict[str, float]:
    """Revenue actual por dept_code (motor sobre Integrity), Σ del rango."""
    rows = await revenue_service.merged_revenue_actual(session, pid, start, end)
    acc: dict[str, float] = {}
    for r in rows:
        dc = r.get("dept_code")
        if dc:
            acc[dc] = acc.get(dc, 0.0) + float(r.get("amount_usd") or 0)
    return acc


async def revenue_detail(session: AsyncSession, business_date: date_cls,
                         property_code: str = "COWLCR") -> dict:
    pid = await _pid(session, property_code)
    month_start = date_cls(business_date.year, business_date.month, 1)
    days_mtd = business_date.day

    rev_today = await _rev_by_dept(session, pid, business_date, business_date)
    rev_mtd = await _rev_by_dept(session, pid, month_start, business_date)
    bud_mtd = await revenue_service._budget_by_dept_cost_center(session, pid, month_start, business_date)

    def line(dept_code: str, label: str) -> dict:
        return {
            "dept_code": dept_code, "label": label,
            "today": round(rev_today.get(dept_code, 0.0), 2),
            "mtd_actual": round(rev_mtd.get(dept_code, 0.0), 2),
            "mtd_budget": round(bud_mtd.get(dept_code, 0.0), 2),
        }

    mapped: set[str] = set()
    sections = []
    for name, members in REVENUE_DETAIL_SECTIONS:
        lines = []
        for dept_code, label in members:
            mapped.add(dept_code)
            lines.append(line(dept_code, label))
        sections.append({"name": name, "lines": lines})

    # §10 — dept_codes que el motor devolvió y no están en el mapa → "Other".
    seen = set(rev_today) | set(rev_mtd) | set(bud_mtd)
    extras = sorted(dc for dc in seen if dc not in mapped)
    if extras:
        sections.append({"name": "Other", "lines": [line(dc, dc) for dc in extras]})

    def subtotal(sec) -> dict:
        return {
            "today": round(sum(x["today"] for x in sec["lines"]), 2),
            "mtd_actual": round(sum(x["mtd_actual"] for x in sec["lines"]), 2),
            "mtd_budget": round(sum(x["mtd_budget"] for x in sec["lines"]), 2),
        }

    for sec in sections:
        sec["subtotal"] = subtotal(sec)

    total = {
        "today": round(sum(s["subtotal"]["today"] for s in sections), 2),
        "mtd_actual": round(sum(s["subtotal"]["mtd_actual"] for s in sections), 2),
        "mtd_budget": round(sum(s["subtotal"]["mtd_budget"] for s in sections), 2),
    }

    return {
        "business_date": business_date.isoformat(), "days_mtd": days_mtd,
        "sections": sections, "total": total,
    }


# ---------------------------------------------------------------------------
# 9.5 F&B by Meal Period (réplica pág. 9-11) — sub-departamento × meal period × Food/Bev
# ---------------------------------------------------------------------------
# Fuente: **Integrity** (misma que 9.1/9.2, revenue autoritativo — reconcilia).
# En el mapping de Integrity el F&B trae 3 dimensiones en la cuenta:
#   • outlet (`cuenta[5:9]`, 0123-0130) → SUB-DEPARTAMENTO (Vitrales, Terra Kitchen…)
#   • naturaleza (`cuenta[:4]`) → Food / Beverage / Misc (engine `fb_subcategory`)
#   • meal period → en el NOMBRE de la cuenta (BREAKFAST/LUNCH/DINER/ALL DAY)
# Filtramos con el mismo criterio del pivote (`weekly_output_column == 'F&B'`) para
# que el Total F&B de 9.5 cuadre 1:1 con 9.2/9.1.
#
# ⚠️ Requiere Integrity cargado del día (igual que 9.1/9.2 en su parte primaria);
# días con solo respaldo 6.4 no tienen detalle por sub-departamento. Covers/checks
# (POS, Tab 2.9) y Avg Check por cubierto vienen de la plantilla de carga.

from app.engine.revenue import fb_subcategory, weekly_output_column  # noqa: E402

# outlet code F&B (cuenta[5:9]) -> nombre de sub-departamento (fallback; el nombre
# real sale de dim_department.outlet_name cuando existe).
FB_OUTLETS = {
    "0123": "Vitrales Restaurant", "0124": "Sueños del Bosque Restaurant",
    "0125": "Pool", "0126": "Beach", "0127": "Room Service",
    "0128": "Private Bar", "0129": "Banquets & Events", "0130": "Terra Kitchen Restaurant",
}
FB_MEAL_ORDER = ["Breakfast", "Lunch", "Dinner", "All Day", "Other"]
# Plantilla de meal period POR outlet: los outlets "all-day" (sin servicio por
# tiempo de comida) muestran solo "All Day"; el resto muestra Breakfast/Lunch/
# Dinner como filas fijas (aunque en $0). Cualquier meal con dato fuera de la
# plantilla se muestra igual (§10, no se pierde plata).
FB_ALLDAY_OUTLETS = {"0125", "0126", "0127", "0128"}  # Pool, Beach, Room Service, Private Bar
FB_RESTAURANT_MEALS = ["Breakfast", "Lunch", "Dinner"]


def _fb_meal_template(outlet: str) -> list[str]:
    return ["All Day"] if outlet in FB_ALLDAY_OUTLETS else list(FB_RESTAURANT_MEALS)


def _fb_meal_from_name(nombre: str | None) -> str:
    n = (nombre or "").upper()
    if "BREAKFAST" in n:
        return "Breakfast"
    if "LUNCH" in n:
        return "Lunch"
    if "DINNER" in n or "DINER" in n:
        return "Dinner"
    if "ALL DAY" in n:
        return "All Day"
    return "Other"


def _fb_cell(d: dict) -> dict:
    food = round(d.get("food", 0.0), 2)
    bev = round(d.get("beverage", 0.0), 2)
    misc = round(d.get("misc", 0.0), 2)
    return {"food": food, "beverage": bev, "misc": misc, "total": round(food + bev + misc, 2)}


async def _fb_integrity_rows(session, pid, start, end):
    return (await session.execute(
        select(IntegrityLine.cuenta, IntegrityLine.nombre_cuenta,
               IntegrityLine.cred_usd, IntegrityLine.deb_usd)
        .where(IntegrityLine.property_id == pid,
               IntegrityLine.business_date >= start, IntegrityLine.business_date <= end)
    )).all()


def _fb_aggregate(rows, dept_names: dict) -> dict:
    """(meal, outlet) -> {food,beverage,misc,name}. Solo cuentas cuya columna de
    pivote es 'F&B' (mismo set que 9.1/9.2 → reconcilia)."""
    acc: dict[tuple[str, str], dict] = {}
    for cuenta, nombre, cred, deb in rows:
        c = str(cuenta or "")
        if weekly_output_column(c) != "F&B":
            continue
        outlet = c[5:9]
        meal = _fb_meal_from_name(nombre)
        sub = fb_subcategory(c).lower()  # 'food' | 'beverage' | 'misc'
        amt = float(cred or 0) - float(deb or 0)
        a = acc.setdefault((meal, outlet), {
            "food": 0.0, "beverage": 0.0, "misc": 0.0,
            "name": dept_names.get(outlet) or FB_OUTLETS.get(outlet) or outlet,
        })
        a[sub] = a.get(sub, 0.0) + amt
    return acc


async def _merged_fb(session, pid, start, end) -> dict:
    """F&B por naturaleza (food/bev/misc) desde `merged_revenue_actual` — la MISMA
    fuente de 9.1/9.2, que incluye el respaldo 6.4 para días sin Integrity."""
    by = await _rev_by_dept(session, pid, start, end)
    return {"food": by.get("FB-FOOD", 0.0), "beverage": by.get("FB-BEV", 0.0),
            "misc": by.get("FB-MISC", 0.0)}


# --- Contadores manuales por día (app_config `PREFIX:YYYY-MM-DD` → int) --------
# Datos que NO existen en el sistema y se capturan a mano (treatments de Spa,
# customers de F&B). Sin migración; el MTD suma los días del mes.
_FB_CUST_PREFIX = "fb_customers:"


async def _daycount_day(session, pid, prefix: str, d: date_cls) -> int:
    row = (await session.execute(
        select(AppConfig.value).where(
            AppConfig.property_id == pid, AppConfig.key == f"{prefix}{d.isoformat()}")
    )).scalar_one_or_none()
    try:
        return int(row) if row else 0
    except (TypeError, ValueError):
        return 0


async def _daycount_range(session, pid, prefix: str, start: date_cls, end: date_cls) -> int:
    rows = (await session.execute(
        select(AppConfig.key, AppConfig.value).where(
            AppConfig.property_id == pid, AppConfig.key.like(f"{prefix}%"))
    )).all()
    total = 0
    for k, v in rows:
        try:
            dd = date_cls.fromisoformat(k.split(":", 1)[1])
        except (ValueError, IndexError):
            continue
        if start <= dd <= end:
            try:
                total += int(v)
            except (TypeError, ValueError):
                pass
    return total


async def _set_daycount(session, pid, prefix: str, d: date_cls, count: int) -> None:
    key = f"{prefix}{d.isoformat()}"
    row = (await session.execute(
        select(AppConfig).where(AppConfig.property_id == pid, AppConfig.key == key)
    )).scalar_one_or_none()
    val = str(max(0, int(count)))
    if row:
        row.value = val
    else:
        session.add(AppConfig(property_id=pid, key=key, value=val))
    await session.commit()


# --- Cubiertos (comensales) por outlet x meal period — Tab 9.5 ----------------
# Captura manual (grilla editable o carga masiva): ni Integrity (mayor contable)
# ni Simphony POS (fact_pos_check = una fila por CHECK, sin comensales) tienen
# este dato. Es el divisor del Cheque Promedio = Revenue / cubiertos.

def _avg_check(rev: float, covers: int) -> float:
    return round(rev / covers, 2) if covers else 0.0


async def _covers_cells(session, pid, start, end) -> dict[tuple[str, str], int]:
    """(meal_period, outlet) -> suma de cubiertos en [start, end]."""
    rows = (await session.execute(
        select(FbCovers.meal_period, FbCovers.outlet, func.sum(FbCovers.covers))
        .where(FbCovers.property_id == pid,
               FbCovers.business_date >= start, FbCovers.business_date <= end)
        .group_by(FbCovers.meal_period, FbCovers.outlet)
    )).all()
    return {(m, o): int(c or 0) for m, o, c in rows}


async def _covers_by_day(session, pid, start, end) -> dict[date_cls, int]:
    rows = (await session.execute(
        select(FbCovers.business_date, func.sum(FbCovers.covers))
        .where(FbCovers.property_id == pid,
               FbCovers.business_date >= start, FbCovers.business_date <= end)
        .group_by(FbCovers.business_date)
    )).all()
    return {d: int(c or 0) for d, c in rows}


async def _legacy_customers_by_day(session, pid, start, end) -> dict[date_cls, int]:
    """Contador viejo (app_config `fb_customers:YYYY-MM-DD`): un total por día,
    sin desglose. Se conserva para no perder lo ya capturado."""
    rows = (await session.execute(
        select(AppConfig.key, AppConfig.value).where(
            AppConfig.property_id == pid, AppConfig.key.like(f"{_FB_CUST_PREFIX}%"))
    )).all()
    out: dict[date_cls, int] = {}
    for k, v in rows:
        try:
            dd = date_cls.fromisoformat(k.split(":", 1)[1])
        except (ValueError, IndexError):
            continue
        if start <= dd <= end:
            try:
                out[dd] = int(v)
            except (TypeError, ValueError):
                pass
    return out


async def _customers_total(session, pid, start, end) -> int:
    """Total de cubiertos del rango. Por DÍA usa el desglose por outlet/meal si
    existe; si ese día no tiene desglose, cae al contador viejo. Nunca suma las
    dos fuentes el mismo día (mismo criterio que el merge Integrity/respaldo 6.4)."""
    cells = await _covers_by_day(session, pid, start, end)
    legacy = await _legacy_customers_by_day(session, pid, start, end)
    return sum(cells.get(d) or legacy.get(d, 0) for d in set(cells) | set(legacy))


async def fb_by_meal_period(session: AsyncSession, business_date: date_cls,
                            property_code: str = "COWLCR") -> dict:
    pid = await _pid(session, property_code)
    month_start = date_cls(business_date.year, business_date.month, 1)
    days_mtd = business_date.day
    dept_names = await revenue_service._dept_names_by_cost_center(session, pid)

    t_acc = _fb_aggregate(await _fb_integrity_rows(session, pid, business_date, business_date), dept_names)
    m_acc = _fb_aggregate(await _fb_integrity_rows(session, pid, month_start, business_date), dept_names)
    # Cubiertos capturados por celda (outlet x meal period), Today y MTD.
    cov_t = await _covers_cells(session, pid, business_date, business_date)
    cov_m = await _covers_cells(session, pid, month_start, business_date)

    def _meal_row(meal, o, t_cell, m_cell) -> dict:
        ct, cm = cov_t.get((meal, o), 0), cov_m.get((meal, o), 0)
        return {
            "meal_period": meal, "today": t_cell, "mtd": m_cell,
            "covers": {"today": ct, "mtd": cm},
            "avg_check": {"today": _avg_check(t_cell["total"], ct),
                          "mtd": _avg_check(m_cell["total"], cm)},
        }

    def name_of(outlet):
        for a in (m_acc, t_acc):
            for (mm, oo), v in a.items():
                if oo == outlet:
                    return v["name"]
        return dept_names.get(outlet) or FB_OUTLETS.get(outlet) or outlet

    # Roster COMPLETO de sub-departamentos: los 8 outlets del catálogo SIEMPRE
    # aparecen (aunque estén en $0), en orden de código, + cualquier outlet extra
    # que aparezca en los datos y no esté en el catálogo (§10, no se pierde nada).
    seen_outlets = {o for (_, o) in t_acc} | {o for (_, o) in m_acc}
    roster = list(FB_OUTLETS.keys()) + sorted(seen_outlets - set(FB_OUTLETS))

    outlets = []
    for o in roster:
        active = any(oo == o for (_, oo) in t_acc) or any(oo == o for (_, oo) in m_acc)
        meals = []
        if active:
            template = _fb_meal_template(o)
            shown = set()
            # Plantilla del outlet SIEMPRE (fila fija, aunque $0): All Day para los
            # all-day; Breakfast/Lunch/Dinner para el resto.
            for meal in template:
                meals.append(_meal_row(meal, o, _fb_cell(t_acc.get((meal, o), {})),
                                       _fb_cell(m_acc.get((meal, o), {}))))
                shown.add(meal)
            # Cualquier otro meal con dato fuera de la plantilla se muestra igual (§10).
            for meal in FB_MEAL_ORDER:
                if meal in shown:
                    continue
                t = t_acc.get((meal, o))
                m = m_acc.get((meal, o))
                if t or m or cov_t.get((meal, o)) or cov_m.get((meal, o)):
                    meals.append(_meal_row(meal, o, _fb_cell(t or {}), _fb_cell(m or {})))

        def _sub(acc):
            return _fb_cell({
                "food": sum(v["food"] for (mm, oo), v in acc.items() if oo == o),
                "beverage": sum(v["beverage"] for (mm, oo), v in acc.items() if oo == o),
                "misc": sum(v["misc"] for (mm, oo), v in acc.items() if oo == o),
            })
        sub_t, sub_m = _sub(t_acc), _sub(m_acc)
        cov_sub_t = sum(c for (mm, oo), c in cov_t.items() if oo == o)
        cov_sub_m = sum(c for (mm, oo), c in cov_m.items() if oo == o)
        outlets.append({
            "outlet": o, "sub_department": name_of(o), "meals": meals,
            "subtotal": {"today": sub_t, "mtd": sub_m},
            "covers": {"today": cov_sub_t, "mtd": cov_sub_m},
            "avg_check": {"today": _avg_check(sub_t["total"], cov_sub_t),
                          "mtd": _avg_check(sub_m["total"], cov_sub_m)},
        })

    def grand(acc):
        return _fb_cell({
            "food": sum(v["food"] for v in acc.values()),
            "beverage": sum(v["beverage"] for v in acc.values()),
            "misc": sum(v["misc"] for v in acc.values()),
        })

    # El total autoritativo es el de 9.1/9.2 (merged_revenue_actual, con respaldo
    # 6.4). Lo que Integrity NO pudo atribuir a un sub-departamento (días con solo
    # respaldo 6.4) se muestra como un "outlet" explícito para que el total cuadre
    # 1:1 con 9.1/9.2 y no se pierda plata en silencio (§10).
    merged_t = await _merged_fb(session, pid, business_date, business_date)
    merged_m = await _merged_fb(session, pid, month_start, business_date)
    attr_t, attr_m = grand(t_acc), grand(m_acc)

    def _gap(merged, attr):
        return _fb_cell({
            "food": merged["food"] - attr["food"],
            "beverage": merged["beverage"] - attr["beverage"],
            "misc": merged["misc"] - attr["misc"],
        })
    gap_t, gap_m = _gap(merged_t, attr_t), _gap(merged_m, attr_m)

    if abs(gap_t["total"]) >= 0.01 or abs(gap_m["total"]) >= 0.01:
        outlets.append({
            "outlet": "—", "sub_department": "Sin detalle por sub-departamento (respaldo 6.4)",
            "meals": [], "subtotal": {"today": gap_t, "mtd": gap_m},
            # Sin cubiertos atribuibles: la fila existe solo para cuadrar el revenue.
            # Las claves van igual para que el payload sea uniforme (el frontend
            # lee outlet.covers/avg_check en TODAS las filas).
            "covers": {"today": 0, "mtd": 0},
            "avg_check": {"today": 0.0, "mtd": 0.0},
        })

    # Resúmenes bajo la tabla: Total Revenue (ya en `total`), Total Customers
    # (conteo manual) y Average Check (= Total Revenue / Customers).
    total_t, total_m = _fb_cell(merged_t), _fb_cell(merged_m)
    # Total de cubiertos: suma del desglose por celda; para los días que aún no
    # tienen desglose se respeta el contador viejo (nunca se suman los dos).
    cust_t = await _customers_total(session, pid, business_date, business_date)
    cust_m = await _customers_total(session, pid, month_start, business_date)

    return {
        "business_date": business_date.isoformat(), "days_mtd": days_mtd,
        "outlets": outlets,
        "total": {"today": total_t, "mtd": total_m},
        "customers": {"today": cust_t, "mtd": cust_m},
        "avg_check": {"today": _avg_check(total_t["total"], cust_t),
                      "mtd": _avg_check(total_m["total"], cust_m)},
    }


async def set_fb_customers(session: AsyncSession, business_date: date_cls,
                           customers: int, property_code: str = "COWLCR") -> dict:
    pid = await _pid(session, property_code)
    await _set_daycount(session, pid, _FB_CUST_PREFIX, business_date, customers)
    return await fb_by_meal_period(session, business_date, property_code=property_code)


async def set_fb_covers(session: AsyncSession, business_date: date_cls, outlet: str,
                        meal_period: str, covers: int,
                        property_code: str = "COWLCR") -> dict:
    """Upsert de UNA celda (día, outlet, meal period). Devuelve el 9.5 recalculado."""
    outlet = (outlet or "").strip()
    meal_period = (meal_period or "").strip()
    if not outlet:
        raise ValueError("El outlet es obligatorio.")
    if meal_period not in FB_MEAL_ORDER:
        raise ValueError(f"Meal period inválido '{meal_period}'. Válidos: {FB_MEAL_ORDER}")
    pid = await _pid(session, property_code)
    row = (await session.execute(
        select(FbCovers).where(
            FbCovers.property_id == pid, FbCovers.business_date == business_date,
            FbCovers.outlet == outlet, FbCovers.meal_period == meal_period)
    )).scalar_one_or_none()
    val = max(0, int(covers))
    if row is None:
        session.add(FbCovers(property_id=pid, business_date=business_date, outlet=outlet,
                             meal_period=meal_period, covers=val))
    else:
        row.covers = val
    await session.commit()
    return await fb_by_meal_period(session, business_date, property_code=property_code)


# Encabezados de la plantilla de carga masiva de cubiertos (orden fijo).
COVERS_TEMPLATE_COLS = ["Date", "Outlet Code", "Sub-department", "Meal Period", "Covers"]


async def build_covers_template(session: AsyncSession, year: int, month: int,
                                property_code: str = "COWLCR") -> bytes:
    """Plantilla del mes: una fila por (día × outlet × meal period) según la
    misma plantilla que usa 9.5, con los cubiertos ya cargados precargados."""
    pid = await _pid(session, property_code)
    dept_names = await revenue_service._dept_names_by_cost_center(session, pid)
    last = calendar.monthrange(year, month)[1]
    start, end = date_cls(year, month, 1), date_cls(year, month, last)

    existing = {(r.business_date, r.outlet, r.meal_period): r.covers for r in (
        await session.execute(select(FbCovers).where(
            FbCovers.property_id == pid,
            FbCovers.business_date >= start, FbCovers.business_date <= end))
    ).scalars().all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "Covers"
    ws.append(COVERS_TEMPLATE_COLS)
    for c in range(1, len(COVERS_TEMPLATE_COLS) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
    for day in range(1, last + 1):
        d = date_cls(year, month, day)
        for outlet in FB_OUTLETS:
            for meal in _fb_meal_template(outlet):
                ws.append([d.isoformat(), outlet,
                           dept_names.get(outlet) or FB_OUTLETS.get(outlet) or outlet,
                           meal, existing.get((d, outlet, meal), 0)])
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 14
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def upload_covers_grid(session: AsyncSession, file_bytes: bytes,
                             property_code: str = "COWLCR") -> dict:
    """Carga masiva. Reemplaza SOLO las celdas (día, outlet, meal) presentes en
    el archivo -- no borra el resto del histórico (mismo criterio que 6.4)."""
    pid = await _pid(session, property_code)
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb["Covers"] if "Covers" in wb.sheetnames else wb.active

    parsed: dict[tuple[date_cls, str, str], int] = {}
    bad_meal: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] in (None, ""):
            continue
        raw_date, outlet, _name, meal, covers = (list(row) + [None] * 5)[:5]
        d = raw_date.date() if hasattr(raw_date, "date") else None
        if d is None:
            try:
                d = date_cls.fromisoformat(str(raw_date).strip()[:10])
            except ValueError:
                continue
        outlet = str(outlet).strip() if outlet is not None else ""
        meal = str(meal).strip() if meal is not None else ""
        if not outlet or not meal:
            continue
        if meal not in FB_MEAL_ORDER:
            bad_meal.add(meal)
            continue
        try:
            val = max(0, int(float(covers or 0)))
        except (TypeError, ValueError):
            val = 0
        parsed[(d, outlet, meal)] = val

    for (d, outlet, meal), val in parsed.items():
        await session.execute(delete(FbCovers).where(
            FbCovers.property_id == pid, FbCovers.business_date == d,
            FbCovers.outlet == outlet, FbCovers.meal_period == meal))
        session.add(FbCovers(property_id=pid, business_date=d, outlet=outlet,
                             meal_period=meal, covers=val))
    await session.commit()

    days = sorted({d for (d, _o, _m) in parsed})
    return {
        "rows_loaded": len(parsed),
        "days": len(days),
        "date_from": days[0].isoformat() if days else None,
        "date_to": days[-1].isoformat() if days else None,
        "total_covers": sum(parsed.values()),
        "unrecognized_meal_periods": sorted(bad_meal),
    }


# ---------------------------------------------------------------------------
# 9.6 F&B Revenue Recap (formato FS) — 3 secciones (Food/Beverage/Misc) × outlet
# ---------------------------------------------------------------------------
# Réplica del "F&B Revenue" recap del formato: una sección por tipo de revenue
# (Food / Beverage / Misc), y dentro los OUTLETS como filas + Total. Columnas
# Today/MTD/Forecast/Budget/Año-ant (mismas que 9.1). Actual real de Integrity
# por outlet; Budget solo por naturaleza (fila Total, no por outlet); Forecast y
# Año anterior = None (vienen de la carga). Reconcilia con 9.1/9.2.

_FB_NATURE_DEPT = {"food": "FB-FOOD", "beverage": "FB-BEV", "misc": "FB-MISC"}
_FB_SECTION_LABEL = {"food": "Food Revenue", "beverage": "Beverage Revenue", "misc": "F&B Misc. Revenue"}


def _fb_collapse_by_outlet(acc: dict) -> dict:
    """(meal,outlet)->cell  ⇒  outlet->{food,beverage,misc,name} (colapsa meal)."""
    out: dict[str, dict] = {}
    for (_meal, o), v in acc.items():
        a = out.setdefault(o, {"food": 0.0, "beverage": 0.0, "misc": 0.0, "name": v["name"]})
        a["food"] += v["food"]; a["beverage"] += v["beverage"]; a["misc"] += v["misc"]
    return out


async def fb_revenue_recap(session: AsyncSession, business_date: date_cls,
                           property_code: str = "COWLCR") -> dict:
    pid = await _pid(session, property_code)
    month_start = date_cls(business_date.year, business_date.month, 1)
    days_mtd = business_date.day
    dept_names = await revenue_service._dept_names_by_cost_center(session, pid)

    t_out = _fb_collapse_by_outlet(
        _fb_aggregate(await _fb_integrity_rows(session, pid, business_date, business_date), dept_names))
    m_out = _fb_collapse_by_outlet(
        _fb_aggregate(await _fb_integrity_rows(session, pid, month_start, business_date), dept_names))

    merged_t = await _merged_fb(session, pid, business_date, business_date)
    merged_m = await _merged_fb(session, pid, month_start, business_date)
    bud = await revenue_service._budget_by_dept_cost_center(session, pid, month_start, business_date)

    seen = set(t_out) | set(m_out)
    roster = list(FB_OUTLETS.keys()) + sorted(seen - set(FB_OUTLETS))

    def name_of(o):
        return (m_out.get(o) or t_out.get(o) or {}).get("name") or dept_names.get(o) or FB_OUTLETS.get(o) or o

    sections = []
    for key in ("food", "beverage", "misc"):
        lines = []
        for o in roster:
            t = t_out.get(o, {}).get(key, 0.0)
            m = m_out.get(o, {}).get(key, 0.0)
            lines.append({"label": name_of(o), "outlet": o, "col": _col(t, m, None)})
        attr_t = sum(l["col"]["today"] for l in lines)
        attr_m = sum(l["col"]["mtd_actual"] for l in lines)
        # residual (respaldo 6.4, sin outlet) para reconciliar con el total nature.
        res_t = merged_t[key] - attr_t
        res_m = merged_m[key] - attr_m
        if abs(res_t) >= 0.01 or abs(res_m) >= 0.01:
            lines.append({"label": "Sin detalle por outlet (respaldo 6.4)",
                          "outlet": "—", "col": _col(res_t, res_m, None)})
        total = _col(merged_t[key], merged_m[key], bud.get(_FB_NATURE_DEPT[key], 0.0))
        sections.append({"key": key, "label": _FB_SECTION_LABEL[key],
                         "lines": lines, "total": total})

    grand = _col(
        merged_t["food"] + merged_t["beverage"] + merged_t["misc"],
        merged_m["food"] + merged_m["beverage"] + merged_m["misc"],
        bud.get("FB-FOOD", 0.0) + bud.get("FB-BEV", 0.0) + bud.get("FB-MISC", 0.0),
    )

    return {
        "business_date": business_date.isoformat(), "days_mtd": days_mtd,
        "sections": sections, "total": grand,
    }


# ---------------------------------------------------------------------------
# 9.8 Beverage Detail — desglose de Beverage Revenue por CONCEPTO (Beer/Wine/Liquors)
# ---------------------------------------------------------------------------
# El "Beverage Revenue" (naturalezas 4125/4130/4131) se abre por concepto según
# la naturaleza de la cuenta, con los outlets como filas (patrón 9.6). Reconcilia
# con el Beverage de 9.6/9.2. El NA Beverage (4120) es un memo informativo: el
# spec lo clasifica en Food, NO en Beverage (§5.1a), así que va aparte.

FB_BEV_CONCEPT = {"4125": "Beer", "4130": "Liquors", "4131": "Wine"}
FB_BEV_ORDER = ["Beer", "Wine", "Liquors"]
_FB_NA_BEV_NATURE = "4120"


def _bev_agg(rows, dept_names: dict) -> dict:
    """(concepto, outlet) -> {amount, name}. Solo naturalezas de beverage."""
    acc: dict[tuple[str, str], dict] = {}
    for cuenta, nombre, cred, deb in rows:
        c = str(cuenta or "")
        if weekly_output_column(c) != "F&B":
            continue
        concept = FB_BEV_CONCEPT.get(c[:4])
        if not concept:
            continue
        outlet = c[5:9]
        amt = float(cred or 0) - float(deb or 0)
        a = acc.setdefault((concept, outlet), {
            "amount": 0.0, "name": dept_names.get(outlet) or FB_OUTLETS.get(outlet) or outlet})
        a["amount"] += amt
    return acc


def _na_bev_total(rows) -> float:
    tot = 0.0
    for cuenta, nombre, cred, deb in rows:
        c = str(cuenta or "")
        if weekly_output_column(c) == "F&B" and c[:4] == _FB_NA_BEV_NATURE:
            tot += float(cred or 0) - float(deb or 0)
    return tot


async def beverage_detail(session: AsyncSession, business_date: date_cls,
                          property_code: str = "COWLCR") -> dict:
    pid = await _pid(session, property_code)
    month_start = date_cls(business_date.year, business_date.month, 1)
    days_mtd = business_date.day
    dept_names = await revenue_service._dept_names_by_cost_center(session, pid)

    rows_t = await _fb_integrity_rows(session, pid, business_date, business_date)
    rows_m = await _fb_integrity_rows(session, pid, month_start, business_date)
    t_acc = _bev_agg(rows_t, dept_names)
    m_acc = _bev_agg(rows_m, dept_names)

    def name_of(concept, o):
        return (m_acc.get((concept, o)) or t_acc.get((concept, o)) or {}).get("name") \
            or dept_names.get(o) or FB_OUTLETS.get(o) or o

    sections = []
    for concept in FB_BEV_ORDER:
        outlets = sorted({o for (cc, o) in t_acc if cc == concept} | {o for (cc, o) in m_acc if cc == concept})
        if not outlets:
            continue
        lines = []
        for o in outlets:
            t = t_acc.get((concept, o), {}).get("amount", 0.0)
            m = m_acc.get((concept, o), {}).get("amount", 0.0)
            lines.append({"label": name_of(concept, o), "outlet": o, "col": _col(t, m, None)})
        lines.sort(key=lambda x: (-x["col"]["mtd_actual"], x["label"]))
        sub = _col(sum(l["col"]["today"] for l in lines), sum(l["col"]["mtd_actual"] for l in lines), None)
        sections.append({"concept": concept, "lines": lines, "subtotal": sub})

    # Total Beverage autoritativo (merged, incl. respaldo 6.4) + residual.
    merged_t = (await _merged_fb(session, pid, business_date, business_date))["beverage"]
    merged_m = (await _merged_fb(session, pid, month_start, business_date))["beverage"]
    attr_t = sum(s["subtotal"]["today"] for s in sections)
    attr_m = sum(s["subtotal"]["mtd_actual"] for s in sections)
    res_t, res_m = merged_t - attr_t, merged_m - attr_m
    if abs(res_t) >= 0.01 or abs(res_m) >= 0.01:
        sections.append({
            "concept": "Sin detalle por concepto (respaldo 6.4)",
            "lines": [{"label": "Beverage sin ingesta Integrity", "outlet": "—", "col": _col(res_t, res_m, None)}],
            "subtotal": _col(res_t, res_m, None),
        })

    bud = await revenue_service._budget_by_dept_cost_center(session, pid, month_start, business_date)
    total = _col(merged_t, merged_m, bud.get("FB-BEV", 0.0))

    na_bev = {"today": round(_na_bev_total(rows_t), 2), "mtd": round(_na_bev_total(rows_m), 2)}

    return {
        "business_date": business_date.isoformat(), "days_mtd": days_mtd,
        "sections": sections, "total": total, "na_beverage": na_bev,
    }


# ---------------------------------------------------------------------------
# 9.7 Spa — Monto · Total Treatments · Average Rate (mínimo)
# ---------------------------------------------------------------------------
# Monto = revenue Spa (dept 0140) de Integrity (reconcilia con 9.1/9.2). El
# conteo de treatments NO existe en el sistema → se captura a mano y se guarda
# en `app_config` (key `spa_treatments:YYYY-MM-DD`, sin migración). Average Rate
# = Monto / Treatments. MTD suma los conteos diarios del mes.

_SPA_DEPT = "0140"
_SPA_TREAT_PREFIX = "spa_treatments:"


async def _spa_treatments_day(session, pid, d: date_cls) -> int:
    row = (await session.execute(
        select(AppConfig.value).where(
            AppConfig.property_id == pid, AppConfig.key == f"{_SPA_TREAT_PREFIX}{d.isoformat()}")
    )).scalar_one_or_none()
    try:
        return int(row) if row else 0
    except (TypeError, ValueError):
        return 0


async def _spa_treatments_range(session, pid, start: date_cls, end: date_cls) -> int:
    rows = (await session.execute(
        select(AppConfig.key, AppConfig.value).where(
            AppConfig.property_id == pid, AppConfig.key.like(f"{_SPA_TREAT_PREFIX}%"))
    )).all()
    total = 0
    for k, v in rows:
        try:
            dd = date_cls.fromisoformat(k.split(":", 1)[1])
        except (ValueError, IndexError):
            continue
        if start <= dd <= end:
            try:
                total += int(v)
            except (TypeError, ValueError):
                pass
    return total


async def spa_summary(session: AsyncSession, business_date: date_cls,
                      property_code: str = "COWLCR") -> dict:
    pid = await _pid(session, property_code)
    month_start = date_cls(business_date.year, business_date.month, 1)
    days_mtd = business_date.day

    rev_t = (await _rev_by_dept(session, pid, business_date, business_date)).get(_SPA_DEPT, 0.0)
    rev_m = (await _rev_by_dept(session, pid, month_start, business_date)).get(_SPA_DEPT, 0.0)
    tr_t = await _spa_treatments_day(session, pid, business_date)
    tr_m = await _spa_treatments_range(session, pid, month_start, business_date)

    def rate(rev, t):
        return round(rev / t, 2) if t else 0.0

    return {
        "business_date": business_date.isoformat(), "days_mtd": days_mtd,
        "revenue": {"today": round(rev_t, 2), "mtd": round(rev_m, 2)},
        "treatments": {"today": tr_t, "mtd": tr_m},
        "avg_rate": {"today": rate(rev_t, tr_t), "mtd": rate(rev_m, tr_m)},
    }


async def set_spa_treatments(session: AsyncSession, business_date: date_cls,
                             treatments: int, property_code: str = "COWLCR") -> dict:
    pid = await _pid(session, property_code)
    key = f"{_SPA_TREAT_PREFIX}{business_date.isoformat()}"
    row = (await session.execute(
        select(AppConfig).where(AppConfig.property_id == pid, AppConfig.key == key)
    )).scalar_one_or_none()
    val = str(max(0, int(treatments)))
    if row:
        row.value = val
    else:
        session.add(AppConfig(property_id=pid, key=key, value=val))
    await session.commit()
    return await spa_summary(session, business_date, property_code=property_code)
