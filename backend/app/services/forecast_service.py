"""Forecast (Tab 6.1.1) — gemelo del Budget (Tab 6.1, budget_service).

Mismo ciclo que el presupuesto: se descarga la plantilla Excel (prellenada con
lo ya cargado), se llena offline, se sube de vuelta y eso **reemplaza el año
completo** — y el diario se deriva solo (`fact_forecast`, mismo reparto que
`fact_budget`: mensual ÷ días del mes, con el residual de redondeo al último día
para que Σ diarios = mensual exacto).

Por qué tabla propia y no una columna `kind` sobre budget_monthly: el reemplazo
anual de uno no puede pisar al otro, y `fact_budget` — que ya consumen revenue,
Tab 9 y el OTB — no cambia de forma.
"""
from __future__ import annotations

import calendar
from datetime import date as date_cls
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.engine.budget import derive_daily_amounts
from app.models import Department, Forecast, ForecastMonthly, Property

HEADERS = ["Dept Code", "Dept Name", "Mes", "Amount USD", "Available Rooms",
          "Rooms Occupied", "Guests", "Occupancy %", "ADR", "Food", "Beverage", "Misc"]

# --- Layout "Working" (Ingresos_12M_Working_YYYY.xlsx) -----------------------
# El archivo con que se trabaja el forecast viene TRANSPUESTO: una fila por
# línea de ingreso y una columna por mes. Sus líneas ya existen como filas de
# dim_department, así que se cargan al mismo grano que el Budget (Tab 6.1) y
# quedan en el mismo orden (ordenado por cost_center).
#
# Las 8 salas de F&B (0123..0130) NO están en el archivo -- el forecast lleva
# F&B agregado en Food/Beverage/Misc. No se escriben filas para ellas: así el
# Tab 6.1.1 muestra sólo el F&B que tiene saldo, sin ocho renglones en cero.
# Dos líneas del archivo van a departamentos cuyo output_column NO tiene fila
# en el cuadro de Tab 3 ("Rooms Others" y "Misc. Rev Others"): si se mapean ahí,
# el GRAND TOTAL suma plata que no aparece en ninguna fila -- exactamente el
# descuadre que este sistema vino a eliminar. Por eso el no-show va dentro de
# Rooms y el "Other / Misc Revenue" a MISC-REV, que sí es la fila visible
# "Misc. Revenue". TODO(bismark): confirmar los dos destinos.
WORKING_LINE_TO_DEPT = {
    "Rooms Revenue-": "0110",
    "Rooms Revenue-No Show": "0110",
    "F&B Food": "FB-FOOD",
    "F&B Beverage": "FB-BEV",
    "F&B Miscellaneous": "FB-MISC",
    "SPA": "0140",
    "Tours": "0150",
    "Gift Shop": "0151",
    "Transportation": "0152",
    "Laundry": "0160",
    "Oinn": "0155",
    "Sustainability Fee": "0170",
    "Other / Misc Revenue": "MISC-REV",
}
_WORKING_HEADER = "Línea de ingreso"
_WORKING_TOTAL = "TOTAL INGRESOS"
# Los estadísticos son por MES (no por línea): se guardan en la fila de Rooms,
# igual que hace la plantilla del Budget.
_WORKING_STATS = {
    "Total available Rooms": "available_rooms",
    "Total Rooms Occupied": "rooms_occupied",
    "Total Guests": "guests",
    "% Occupancy": "occupancy_pct",
    "ADR": "adr",
}


def _is_working_layout(ws) -> bool:
    """Distingue el archivo Working del template propio (Dept Code / Mes)."""
    for row in ws.iter_rows(min_row=1, max_row=30, max_col=1, values_only=True):
        if row[0] and str(row[0]).strip() == _WORKING_HEADER:
            return True
    return False


def _parse_working(ws) -> tuple[dict, dict, list[str]]:
    """Devuelve (montos, estadísticos, líneas_no_reconocidas).

    montos:  {(dept_code, month): Decimal}
    stats:   {month: {campo: Decimal}}
    Una celda vacía NO es cero: se omite (no se inventa dato). Un 0 explícito
    sí se carga -- octubre viene en cero a propósito (operación cerrada).
    """
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_col=1, values_only=True), start=1):
        if row[0] and str(row[0]).strip() == _WORKING_HEADER:
            header_row = i
            break
    if header_row is None:
        raise ValueError(f"No se encontró la fila de encabezado '{_WORKING_HEADER}'.")

    amounts: dict = {}
    stats: dict = {}
    unmapped: list[str] = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        label = row[0].value
        if label is None:
            continue
        label = str(label).strip()

        if label in _WORKING_STATS:
            field = _WORKING_STATS[label]
            for month in range(1, 13):
                v = row[month].value
                if v is None or v == "":
                    continue
                stats.setdefault(month, {})[field] = Decimal(str(v))
            continue

        if row[0].row <= header_row or label == _WORKING_TOTAL:
            continue

        dept_code = WORKING_LINE_TO_DEPT.get(label)
        if dept_code is None:
            unmapped.append(label)
            continue
        for month in range(1, 13):
            v = row[month].value
            if v is None or v == "":
                continue
            # ACUMULA: dos líneas del archivo (Rooms Revenue- y su no-show)
            # caen en el mismo departamento; asignar en vez de sumar perdería
            # una de las dos sin avisar.
            key = (dept_code, month)
            amounts[key] = amounts.get(key, Decimal("0")) + Decimal(str(v)).quantize(Decimal("0.01"))

    return amounts, stats, unmapped


async def _property_id(session: AsyncSession, code: str):
    pid = (await session.execute(
        select(Property.id).where(Property.code == code)
    )).scalar_one_or_none()
    if pid is None:
        raise ValueError(f"Propiedad '{code}' no existe.")
    return pid


async def _departments(session: AsyncSession, pid) -> list[Department]:
    return (await session.execute(
        select(Department).where(Department.property_id == pid).order_by(Department.cost_center)
    )).scalars().all()


def _n(v) -> float | None:
    return float(v) if v is not None else None


async def build_template(session: AsyncSession, year: int, property_code: str = "COWLCR") -> bytes:
    """Excel prellenado con lo que ya está cargado para `year` (o en cero si no
    hay nada) — un renglón por departamento × mes."""
    pid = await _property_id(session, property_code)
    depts = await _departments(session, pid)
    existing = (await session.execute(
        select(ForecastMonthly).where(ForecastMonthly.property_id == pid, ForecastMonthly.year == year)
    )).scalars().all()
    by_key = {(str(f.dept_id), f.month): f for f in existing}

    wb = Workbook()
    ws = wb.active
    ws.title = f"Forecast {year}"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2D3A5C")

    for d in depts:
        for month in range(1, 13):
            f = by_key.get((str(d.id), month))
            ws.append([
                d.cost_center, d.outlet_name, month,
                float(f.amount_usd) if f else 0.0,
                _n(f.available_rooms) if f else None, _n(f.rooms_occupied) if f else None,
                _n(f.guests) if f else None, _n(f.occupancy_pct) if f else None,
                _n(f.adr) if f else None,
                float(f.food) if f else 0.0, float(f.beverage) if f else 0.0,
                float(f.misc) if f else 0.0,
            ])
    ws.freeze_panes = "A2"
    for i, w in enumerate([11, 28, 6, 13, 15, 15, 9, 12, 10, 11, 11, 11], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _replace_year(session: AsyncSession, pid, year: int) -> None:
    await session.execute(delete(ForecastMonthly).where(
        ForecastMonthly.property_id == pid, ForecastMonthly.year == year))
    await session.execute(delete(Forecast).where(
        Forecast.property_id == pid,
        Forecast.date >= date_cls(year, 1, 1), Forecast.date <= date_cls(year, 12, 31)))


async def _upload_working(session: AsyncSession, pid, year: int, ws, by_code: dict) -> dict:
    """Carga el layout Working (línea de ingreso × mes). Escribe SOLO las líneas
    que trae el archivo: las 8 salas de F&B sin saldo no generan renglón."""
    amounts, stats, unmapped = _parse_working(ws)

    faltantes = sorted({c for c, _ in amounts if c not in by_code})
    if faltantes:
        raise ValueError(
            "Estos códigos de departamento no existen en dim_department: "
            f"{', '.join(faltantes)}. Cargalos en Tab 6.3 antes de subir el forecast.")

    await _replace_year(session, pid, year)

    meses = sorted({m for _, m in amounts} | set(stats))
    loaded = 0
    for month in meses:
        st = stats.get(month, {})
        for code in sorted({c for c, m in amounts if m == month}):
            dept = by_code[code]
            amount = amounts[(code, month)]
            # Los estadísticos son del mes: van en la fila de Rooms, igual que
            # en la plantilla del Budget.
            es_rooms = code == "0110"
            session.add(ForecastMonthly(
                property_id=pid, year=year, month=month, dept_id=dept.id, amount_usd=amount,
                available_rooms=st.get("available_rooms") if es_rooms else None,
                rooms_occupied=st.get("rooms_occupied") if es_rooms else None,
                guests=st.get("guests") if es_rooms else None,
                occupancy_pct=st.get("occupancy_pct") if es_rooms else None,
                adr=st.get("adr") if es_rooms else None,
                food=Decimal("0"), beverage=Decimal("0"), misc=Decimal("0"),
            ))
            loaded += 1
            if amount:
                for day_idx, amt in enumerate(derive_daily_amounts(amount, year, month), start=1):
                    session.add(Forecast(
                        property_id=pid, date=date_cls(year, month, day_idx),
                        dept_id=dept.id, amount_usd=amt,
                    ))

    await session.commit()
    return {"year": year, "layout": "working", "rows_loaded": loaded,
            "lineas_no_reconocidas": sorted(set(unmapped)),
            "dept_codes_no_reconocidos": []}


async def upload_and_replace(session: AsyncSession, year: int, file_bytes: bytes,
                             property_code: str = "COWLCR") -> dict:
    """Reemplazo total del año (forecast_monthly + fact_forecast derivado).

    Acepta los dos layouts: la plantilla propia (Dept Code × Mes) y el archivo
    Working del forecast (línea de ingreso × mes), que se detecta por contenido.
    """
    pid = await _property_id(session, property_code)
    depts = await _departments(session, pid)
    by_code = {d.cost_center: d for d in depts}

    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    if _is_working_layout(ws):
        return await _upload_working(session, pid, year, ws, by_code)

    rows = list(ws.iter_rows(min_row=2, values_only=True))

    await session.execute(delete(ForecastMonthly).where(
        ForecastMonthly.property_id == pid, ForecastMonthly.year == year))
    await session.execute(delete(Forecast).where(
        Forecast.property_id == pid,
        Forecast.date >= date_cls(year, 1, 1), Forecast.date <= date_cls(year, 12, 31)))

    loaded, skipped = 0, []
    for r in rows:
        if not r or r[0] is None:
            continue
        (dept_code, _name, month, amount_usd, avail, occ_rooms, guests,
         occ_pct, adr, food, bev, misc) = (list(r) + [None] * 12)[:12]
        dept = by_code.get(str(dept_code))
        if dept is None or month is None:
            skipped.append(str(dept_code))
            continue
        month = int(month)
        amount = Decimal(str(amount_usd or 0))

        session.add(ForecastMonthly(
            property_id=pid, year=year, month=month, dept_id=dept.id, amount_usd=amount,
            available_rooms=Decimal(str(avail)) if avail not in (None, "") else None,
            rooms_occupied=Decimal(str(occ_rooms)) if occ_rooms not in (None, "") else None,
            guests=Decimal(str(guests)) if guests not in (None, "") else None,
            occupancy_pct=Decimal(str(occ_pct)) if occ_pct not in (None, "") else None,
            adr=Decimal(str(adr)) if adr not in (None, "") else None,
            food=Decimal(str(food or 0)), beverage=Decimal(str(bev or 0)), misc=Decimal(str(misc or 0)),
        ))
        loaded += 1

        if amount:
            for day_idx, amt in enumerate(derive_daily_amounts(amount, year, month), start=1):
                session.add(Forecast(
                    property_id=pid, date=date_cls(year, month, day_idx),
                    dept_id=dept.id, amount_usd=amt,
                ))

    await session.commit()
    return {"year": year, "layout": "template", "rows_loaded": loaded,
            "lineas_no_reconocidas": [], "dept_codes_no_reconocidos": sorted(set(skipped))}


async def monthly_summary(session: AsyncSession, year: int, property_code: str = "COWLCR") -> list[dict]:
    pid = await _property_id(session, property_code)
    depts = {d.id: d for d in await _departments(session, pid)}
    rows = (await session.execute(
        select(ForecastMonthly).where(ForecastMonthly.property_id == pid, ForecastMonthly.year == year)
        .order_by(ForecastMonthly.month)
    )).scalars().all()
    return [{
        "dept_code": depts[r.dept_id].cost_center if r.dept_id in depts else None,
        "dept_name": depts[r.dept_id].outlet_name if r.dept_id in depts else None,
        "month": r.month, "amount_usd": float(r.amount_usd),
        "available_rooms": _n(r.available_rooms), "rooms_occupied": _n(r.rooms_occupied),
        "guests": _n(r.guests), "occupancy_pct": _n(r.occupancy_pct), "adr": _n(r.adr),
        "food": float(r.food), "beverage": float(r.beverage), "misc": float(r.misc),
    } for r in rows]


async def daily_summary(session: AsyncSession, year: int, month: int,
                        property_code: str = "COWLCR") -> list[dict]:
    """Forecast diario derivado (fact_forecast) de un mes."""
    pid = await _property_id(session, property_code)
    depts = {d.id: d for d in await _departments(session, pid)}
    days_in_month = calendar.monthrange(year, month)[1]
    rows = (await session.execute(
        select(Forecast).where(
            Forecast.property_id == pid,
            Forecast.date >= date_cls(year, month, 1),
            Forecast.date <= date_cls(year, month, days_in_month),
        ).order_by(Forecast.date)
    )).scalars().all()
    return [{
        "date": r.date.isoformat(),
        "dept_code": depts[r.dept_id].cost_center if r.dept_id in depts else None,
        "dept_name": depts[r.dept_id].outlet_name if r.dept_id in depts else None,
        "amount_usd": float(r.amount_usd),
    } for r in rows]
