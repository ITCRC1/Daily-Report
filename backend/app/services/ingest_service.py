"""Ingesta por batch de un día: archivos crudos -> base (§2.5 reemplazo total por día).

Clasifica por contenido (§2.8): el Integrity se detecta por tener hoja 'Datos',
no por el nombre del archivo (el real se llama 'DAILY REVENUE REPORT ...').
"""
from __future__ import annotations

import re
from collections import defaultdict
from datetime import date as date_cls
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import get_settings
from app.ingest import bills as bills_parser
from app.ingest import integrity as integ
from app.ingest import opera
from app.ingest import pos as pos_parser
from app.models import (
    AppConfig,
    AuditRun,
    Bill,
    BillLine,
    CityLedgerInvoice,
    IngestBatch,
    IngestDayStatus,
    IntegrityLine,
    LedgerBalance,
    OccupancyStat,
    OperaTxn,
    OperaTxnDetail,
    Otb,
    OtbDaily,
    OtbMonthly,
    PosCheck,
    PosSummary,
    Property,
    RoomStat,
    TrialBalanceLine,
)

# Sheets que identifican el Excel de Ventas Simphony/POS (app/ingest/pos.py).
POS_SHEETS = {"Resumen Ejecutivo", "Detalle de Checks"}

INPUTS_ROOT = Path(__file__).resolve().parents[2].parent / "goldens" / "inputs"


def uploads_root() -> Path:
    """Raíz de subidas reales de Tab 1 (uploads/inputs) -- única fuente de verdad,
    compartida por api/ingest.py y audit_service.refresh_day(). Antes cada uno
    recalculaba esta misma ruta por separado, y quedó desincronizada más de una
    vez (bugs reales del 2026-07-02/03 que borraron datos de producción).

    `UPLOADS_DIR` (env var) la puede sobreescribir -- en producción apunta a un
    volumen persistente de Railway; sin ella, cae al mismo cálculo relativo de
    siempre (correcto solo en dev local, donde el repo completo vive en disco)."""
    override = get_settings().UPLOADS_DIR
    if override:
        return Path(override)
    return Path(__file__).resolve().parents[2].parent / "uploads" / "inputs"


def _is_forecast(name: str) -> bool:
    """¿Es un archivo de History/Forecast de Opera?

    Estos archivos (ej. 'OPERA_HistoryForecast_Total Revenue.XML',
    'history_forecast (Total + Rooms Only).XML') traen la palabra 'Revenue' en
    el nombre y por eso encajaban por accidente en el patrón `REVENUE.*\\.xml`
    del archivo de transacciones de Opera. Como `find_file` devuelve el PRIMER
    match según el orden del sistema de archivos, en Linux (Railway) el forecast
    ganaba antes que el REVENUE real -> Opera quedaba en 0 transacciones (bug
    real 2026-07-04, dependiente del orden de `iterdir()`; no se reproducía en
    Windows). Se normaliza a solo-letras para que 'HistoryForecast',
    'history_forecast' y 'History Forecast' colapsen al mismo token."""
    return "historyforecast" in re.sub(r"[^a-z]", "", name.lower())


def _is_roomtype_stat(name: str) -> bool:
    """¿Es el archivo de estadísticas POR CATEGORÍA de habitación (room stats)?

    Nombre automático de Opera: 'OPERA_StatisticsRoomType_YYYY-MM-DD.xml'
    (histórico, subido a mano: 'statroomtype.xml.XML'). Tiene DOS problemas si se
    clasifica ingenuamente: (1) el patrón viejo `statroomtype` NO matchea el
    nombre automático 'StatisticsRoomType'; (2) 'StatisticsRoomType' SÍ matchea
    `STATISTICS.*\\.xml` y por eso chocaba con el archivo de ocupación real
    (OPERA_GEN_XMLBO_STATISTICS) según el orden de `iterdir()` -- mismo tipo de
    bug de orden que el forecast. Se aísla en su propia categoría y se saca del
    pool de archivo único. Normalizado a solo-letras para tolerar variantes."""
    n = re.sub(r"[^a-z]", "", name.lower())
    return "statisticsroomtype" in n or "statroomtype" in n


def _classify(inputs_dir: Path) -> dict:
    """Clasifica los archivos crudos de un día por rol, de forma DETERMINISTA
    (no depende del orden de `iterdir()`), tolerando tanto los nombres
    automáticos de Opera (OPERA_GEN_XMLBO_*, OPERA_HistoryForecast_*,
    OPERA_StatisticsRoomType_*) como los renombrados a mano de subidas viejas."""
    files = list(inputs_dir.iterdir()) if inputs_dir.exists() else []
    xmls = [f for f in files if f.suffix.lower() == ".xml"]
    xlsx = [f for f in files if f.suffix.lower() in (".xlsx", ".xlsm")]

    # Fuentes con su PROPIA categoría, detectadas por nombre canónico:
    #  - forecast (OTB): OPERA_HistoryForecast_TotalRevenue + _Default (los 2).
    #    parse_history_forecast() los ordena por revenue (mayor=Total, menor=Solo
    #    Habitaciones) -> el mapeo Total/Rooms sale solo, sin depender del nombre.
    #  - room_stats: OPERA_StatisticsRoomType.
    forecast_files = [f for f in xmls if _is_forecast(f.name)]
    room_stat_file = next((f for f in xmls if _is_roomtype_stat(f.name)), None)

    # Pool de archivo ÚNICO (revenue/statistics/bills/customer): EXCLUYE forecast
    # y room-stats para que 'Revenue' y 'STATISTICS' nunca colisionen con
    # 'HistoryForecast...Revenue' ni con 'StatisticsRoomType' según el orden del
    # filesystem (los 2 bugs de orden ya vistos).
    single = [f for f in xmls if not _is_forecast(f.name) and not _is_roomtype_stat(f.name)]

    integrity_file = None
    pos_file = None
    for f in xlsx:
        try:
            wb = load_workbook(f, read_only=True)
            sheets = set(wb.sheetnames)
            if integrity_file is None and "Datos" in sheets:
                integrity_file = f
            elif pos_file is None and POS_SHEETS & sheets:
                pos_file = f
            wb.close()
        except Exception:
            continue

    return {
        "revenue": opera.find_file(single, r"REVENUE.*\.xml"),
        "statistics": opera.find_file(single, r"STATISTICS.*\.xml"),
        "history_forecast": [str(f) for f in forecast_files],
        "integrity": str(integrity_file) if integrity_file else None,
        "bills": opera.find_file(single, r"BILLS.*\.xml"),
        "customer": opera.find_file(single, r"CUSTOMER.*\.xml"),
        "room_stats": str(room_stat_file) if room_stat_file else None,
        "trial_balance": opera.find_file(single, r"TrialBalance.*\.xml"),
        "city_ledger": opera.find_file(single, r"CITYLEDGER.*\.xml"),
        "pos": str(pos_file) if pos_file else None,
    }


async def _property_id(session: AsyncSession, code: str):
    pid = (await session.execute(
        select(Property.id).where(Property.code == code)
    )).scalar_one_or_none()
    if pid is None:
        raise ValueError(f"Propiedad '{code}' no existe (corré el seed).")
    return pid


async def _set_status(session: AsyncSession, pid, bdate, sistema: str, estado: str):
    await session.execute(
        pg_insert(IngestDayStatus)
        .values(property_id=pid, business_date=bdate, sistema=sistema, estado=estado)
        .on_conflict_do_update(
            index_elements=["property_id", "business_date", "sistema"],
            set_={"estado": estado},
        )
    )


async def ingest_day(
    session: AsyncSession,
    business_date: date_cls,
    property_code: str = "COWLCR",
    inputs_dir: Path | None = None,
    uploaded_by=None,
) -> dict:
    """Ingesta un día completo. Reemplazo total: borra lo previo del día y recarga."""
    inputs_dir = inputs_dir or (INPUTS_ROOT / business_date.isoformat())
    pid = await _property_id(session, property_code)
    cls = _classify(inputs_dir)

    # Guardia de seguridad (§10, 2026-07-03): el reemplazo total de abajo borra
    # los datos existentes del día ANTES de reinsertar. Si `_classify` no
    # reconoció ni un solo archivo (carpeta vacía o ruta equivocada), abortar
    # acá -- de lo contrario se borra lo real sin nada que lo reemplace. Este
    # es el bug que ya causó pérdida de datos reales de producción (2026-07-02,
    # ver comentario en api/ingest.py::ingest_and_audit).
    if not any(cls.values()):
        raise ValueError(
            f"No se encontró ningún archivo reconocible en '{inputs_dir}'. "
            "Se abortó la ingesta para no borrar los datos existentes del día "
            "sin reemplazo. Verificá que los archivos estén subidos y que la "
            "ruta sea la correcta."
        )

    # Pre-flight "todo o nada" (pedido del owner 2026-07-06): una auditoría
    # depende de TODAS sus fuentes -- si una fuente OBLIGATORIA viene en cero
    # (archivo ausente, equivocado o vacío), NO cargar a medias. Se valida el
    # set mínimo (`gate_min_set`, por defecto opera+integrity) parseando sólo
    # esos archivos ANTES del borrado de abajo: si algo falla se aborta sin
    # tocar la DB, y el mensaje dice exactamente qué corregir. Esto es lo que
    # habría frenado el incidente del 07-04 (Opera clasificado al archivo
    # equivocado -> 0 transacciones, pero la auditoría igual dio "OK").
    required = await _gate_min_set(session, pid)
    problems: list[str] = []
    if "opera" in required:
        if not cls["revenue"]:
            problems.append(
                "Opera Revenue: no se encontró el archivo de transacciones "
                "(OPERA_..._REVENUE_*.xml)."
            )
        else:
            _hdrs, _ = opera.parse_revenue(cls["revenue"])
            if not _hdrs:
                problems.append(
                    f"Opera Revenue: el archivo '{Path(cls['revenue']).name}' se "
                    "leyó pero trae 0 transacciones (archivo equivocado o vacío)."
                )
    if "integrity" in required:
        if not cls["integrity"]:
            problems.append("Integrity: no se encontró el archivo (hoja 'Datos').")
        elif not integ.parse_integrity_lines(cls["integrity"]):
            problems.append(
                f"Integrity: el archivo '{Path(cls['integrity']).name}' se leyó "
                "pero trae 0 líneas."
            )
    if problems:
        raise ValueError(
            "Ingesta abortada — una o más fuentes obligatorias vinieron en cero. "
            "NO se borró ni modificó ningún dato del día; corregí los archivos y "
            "volvé a subir:\n- " + "\n- ".join(problems)
        )

    # batch nuevo
    batch = IngestBatch(property_id=pid, business_date=business_date, uploaded_by=uploaded_by)
    session.add(batch)
    await session.flush()

    # reemplazo total por día (§2.5)
    for model in (OperaTxn, OperaTxnDetail, IntegrityLine, Bill, BillLine,
                  OccupancyStat, Otb, RoomStat, PosCheck, PosSummary,
                  TrialBalanceLine, LedgerBalance, CityLedgerInvoice):
        await session.execute(
            delete(model).where(model.property_id == pid, model.business_date == business_date)
        )

    # OtbMonthly/OtbDaily usan snapshot_date (no business_date), se borran aparte
    for _m in (OtbMonthly, OtbDaily):
        await session.execute(
            delete(_m).where(_m.property_id == pid, _m.snapshot_date == business_date)
        )

    counts = {"opera_hdr": 0, "opera_det": 0, "integrity_lines": 0, "bills": 0,
              "occupancy_stats": 0, "otb": 0, "room_stats": 0, "pos_checks": 0,
              "trial_balance": 0, "ledger_balances": 0, "city_ledger": 0,
              "otb_monthly": 0, "otb_daily": 0}

    # --- Opera REVENUE -> fact_opera_txn + fact_opera_txn_detail ---
    if cls["revenue"]:
        headers, details = opera.parse_revenue(cls["revenue"])
        for h in headers:
            session.add(OperaTxn(
                property_id=pid, business_date=business_date, tcode=h.tcode,
                description=h.description, type=h.type, total=h.total,
                guest_ledger=h.guest_ledger, package_ledger=h.package_ledger,
                ar_ledger=h.ar_ledger, deposit_ledger=h.deposit_ledger,
            ))
        for d in details:
            session.add(OperaTxnDetail(
                property_id=pid, business_date=business_date, tcode=d.tcode,
                description=d.description, type=d.type, market_code=d.market_code,
                room_class=d.room_class, trx_amount=d.trx_amount,
                trx_guest_ledger=d.trx_guest_ledger, trx_package_ledger=d.trx_package_ledger,
            ))
        counts["opera_hdr"] = len(headers)
        counts["opera_det"] = len(details)
        await _set_status(session, pid, business_date, "opera", "Listo")

    # --- Integrity -> stg_integrity_line (grano línea) ---
    if cls["integrity"]:
        lines = integ.parse_integrity_lines(cls["integrity"])
        src = Path(cls["integrity"]).name
        for ln in lines:
            session.add(IntegrityLine(
                property_id=pid, business_date=business_date, source_file=src,
                ingest_batch_id=batch.id, **ln,
            ))
        counts["integrity_lines"] = len(lines)
        await _set_status(session, pid, business_date, "integrity", "Listo")

    # --- Opera STATISTICS -> fact_occupancy_stat (2.4: ocupación por market/clase/tipo) ---
    if cls["statistics"]:
        stats = opera.parse_statistics(cls["statistics"])
        for s in stats:
            session.add(OccupancyStat(
                property_id=pid, business_date=business_date,
                market_code=s.market_code, room_class=s.room_class, room_type=s.room_type,
                rooms=s.rooms, persons=s.persons,
                noshow_rooms=s.noshow, cancel_rooms=s.cancel,
            ))
        counts["occupancy_stats"] = len(stats)

    # --- history_forecast (x2) -> fact_otb (2.5: Total vs Rooms Only, §5.6) ---
    if cls["history_forecast"]:
        hf = opera.parse_history_forecast(cls["history_forecast"], business_date)
        for scope in ("total", "rooms"):
            if scope in hf:
                row = hf[scope]
                session.add(Otb(
                    property_id=pid, business_date=business_date, scope=scope,
                    source_file=Path(row["file"]).name, revenue=row["revenue"],
                    no_rooms=row["rooms"], no_persons=row["persons"],
                    inventory_rooms=row["inventory"], adr=row["adr"],
                    occupancy=row["occupancy"],
                ))
        counts["otb"] = len(hf)

        # Agregado MENSUAL del history_forecast (año completo) -> fact_otb_monthly
        # (On The Books, Tab 8). Snapshot = business_date; reemplazo total del día.
        monthly = opera.history_forecast_monthly(cls["history_forecast"])
        for (yr, mth), mv in monthly.items():
            session.add(OtbMonthly(
                property_id=pid, snapshot_date=business_date, year=yr, month=mth,
                total_revenue=mv["total_revenue"], rooms_only_revenue=mv["rooms_only_revenue"],
                rooms_only_history=mv["rooms_only_history"], rooms_only_forecast=mv["rooms_only_forecast"],
                rooms_occ=mv["rooms_occ"], guests=mv["guests"], rooms_avail=mv["rooms_avail"],
            ))
        counts["otb_monthly"] = len(monthly)

        # Ocupación DIARIA del OTB -> fact_otb_daily (heatmap, Tab 8.3).
        daily = opera.history_forecast_daily(cls["history_forecast"])
        for iso, dv in daily.items():
            session.add(OtbDaily(
                property_id=pid, snapshot_date=business_date,
                the_date=date_cls.fromisoformat(iso),
                rooms_sold=dv["rooms_sold"], rooms_avail=dv["rooms_avail"],
            ))
        counts["otb_daily"] = len(daily)

    # --- statroomtype -> fact_room_stat (etapa 5: ADR/Occ/Yield por categoría) ---
    # El archivo trae el año completo; se filtra al día del batch (§2.5/§2.8).
    if cls["room_stats"]:
        all_stats = opera.parse_statroomtype(cls["room_stats"])
        day_stats = [s for s in all_stats if s.business_date == business_date]
        for s in day_stats:
            session.add(RoomStat(
                property_id=pid, business_date=business_date,
                room_category=s.short_description, room_revenue=s.room_revenue,
                stay_rooms=s.stay_rooms, stay_persons=s.stay_persons,
                physical_rooms=s.physical_rooms,
            ))
        counts["room_stats"] = len(day_stats)

    # --- POS (Simphony) -> fact_pos_check (detalle) + fact_pos_summary (Resumen +
    # Mapeo Simphony→Opera) — etapa/sub-tab 2.9. Portado de reference/auditoria.py.
    if cls.get("pos"):
        meta = pos_parser.parse_pos_excel(cls["pos"])
        for chk in meta.get("all_checks", []):
            session.add(PosCheck(
                property_id=pid, business_date=business_date,
                restaurant=chk["restaurant"], employee=chk["employee"],
                check_num=chk["check_num"], hora=chk["hora"],
                forma_pago=chk["forma_pago"], monto=chk["monto"],
                is_room_charge="ROOM CHARGE" in chk["forma_pago"].upper(),
            ))
        session.add(PosSummary(
            property_id=pid, business_date=business_date,
            source_file=Path(cls["pos"]).name,
            ventas_netas=meta.get("ventas_netas", 0), cargos_servicio=meta.get("sc", 0),
            total_ventas=meta.get("total_dia", 0), voids=meta.get("voids", 0),
            room_charge_confirmado=meta.get("room_charge", 0),
        ))
        counts["pos_checks"] = len(meta.get("all_checks", []))
        await _set_status(session, pid, business_date, "pos", "Listo")

    # --- Bills (folios) -> fact_bill + fact_bill_line (detalle auxiliar Guest Ledger) ---
    if cls["bills"]:
        customers = bills_parser.parse_customers(cls["customer"]) if cls["customer"] else {}
        parsed = bills_parser.parse_bills(cls["bills"], customers)
        for b in parsed:
            session.add(Bill(
                property_id=pid, business_date=business_date, bill_no=b.bill_no,
                bill_type=b.bill_type, status=b.status, guest_internal_id=b.guest_internal_id,
                guest_name=b.guest_name, total_amount=b.total_amount,
            ))
            for ln in b.lines:
                session.add(BillLine(
                    property_id=pid, business_date=business_date, bill_no=b.bill_no,
                    trx_code=ln.trx_code, trx_date=ln.trx_date, net_amount=ln.net_amount,
                    debit_amount=ln.debit_amount, credit_amount=ln.credit_amount,
                ))
        counts["bills"] = len(parsed)

    # --- Trial Balance oficial -> fact_trial_balance (2.2) + fact_ledger_balance (2.3) ---
    if cls.get("trial_balance"):
        tb_lines, tb_balances = opera.parse_trial_balance(cls["trial_balance"])
        for ln in tb_lines:
            session.add(TrialBalanceLine(
                property_id=pid, business_date=business_date,
                trx_type=ln["trx_type"], trx_type_desc=ln["trx_type_desc"],
                tcode=ln["tcode"], description=ln["description"],
                tb_amount=ln["tb_amount"], net_amount=ln["net_amount"],
                guest_ledger=ln["guest_ledger"], package_ledger=ln["package_ledger"],
                ar_ledger=ln["ar_ledger"], deposit_ledger=ln["deposit_ledger"],
            ))
        for ledger, b in tb_balances.items():
            session.add(LedgerBalance(
                property_id=pid, business_date=business_date, ledger=ledger,
                opening=b["opening"], debit=b["debit"], credit=b["credit"],
                closing=b["closing"],
            ))
        counts["trial_balance"] = len(tb_lines)
        counts["ledger_balances"] = len(tb_balances)

    # --- City Ledger -> fact_city_ledger (detalle del movimiento del AR Ledger, 2.3) ---
    if cls.get("city_ledger"):
        invoices = opera.parse_city_ledger(cls["city_ledger"])
        for inv in invoices:
            session.add(CityLedgerInvoice(
                property_id=pid, business_date=business_date,
                trx_code=inv["trx_code"], transaction_type=inv["transaction_type"],
                bill_no=inv["bill_no"], invoice_no=inv["invoice_no"], amount=inv["amount"],
                customer_internal_id=inv["customer_internal_id"],
                account_name=inv["account_name"], account_number=inv["account_number"],
                confirmation_no=inv["confirmation_no"], arrival_date=inv["arrival_date"],
                departure_date=inv["departure_date"], guest_name=inv["guest_name"],
            ))
        counts["city_ledger"] = len(invoices)

    await session.commit()
    return {
        "business_date": business_date.isoformat(),
        "property": property_code,
        "batch_id": str(batch.id),
        "classified": {k: (v if isinstance(v, (str, list)) else None) for k, v in cls.items()},
        "counts": counts,
    }


async def _gate_min_set(session: AsyncSession, pid) -> set[str]:
    row = (await session.execute(
        select(AppConfig).where(AppConfig.property_id == pid, AppConfig.key == "gate_min_set")
    )).scalar_one_or_none()
    raw = row.value if row else "opera,integrity"
    return {s.strip() for s in raw.split(",") if s.strip()}


async def day_status_grid(session: AsyncSession, year: int, property_code: str = "COWLCR") -> dict:
    """Malla de 365 días × sistema (Tab 1, §4): Incompleto/Listo/Auditado/Cerrado.

    Solo devuelve los días que tienen algún rastro (ingest_day_status o
    audit_run) — un año casi vacío no manda 365 filas en blanco; el frontend
    pinta el calendario completo y rellena los huecos como 'sin datos'.
    """
    pid = await _property_id(session, property_code)
    min_set = await _gate_min_set(session, pid)
    year_start = date_cls(year, 1, 1)
    year_end = date_cls(year, 12, 31)

    status_rows = (await session.execute(
        select(IngestDayStatus).where(
            IngestDayStatus.property_id == pid,
            IngestDayStatus.business_date >= year_start,
            IngestDayStatus.business_date <= year_end,
        )
    )).scalars().all()
    by_day: dict[date_cls, dict[str, str]] = defaultdict(dict)
    for r in status_rows:
        by_day[r.business_date][r.sistema] = r.estado

    run_rows = (await session.execute(
        select(AuditRun).where(
            AuditRun.property_id == pid,
            AuditRun.business_date >= year_start,
            AuditRun.business_date <= year_end,
        )
    )).scalars().all()
    run_by_day = {r.business_date: r for r in run_rows}

    days = []
    for d in sorted(set(by_day) | set(run_by_day)):
        systems = by_day.get(d, {})
        run = run_by_day.get(d)
        if run and run.status == "cerrado":
            overall = "Cerrado"
        elif run:
            overall = "Auditado"
        elif min_set.issubset(systems.keys()):
            overall = "Listo"
        else:
            overall = "Incompleto"
        days.append({
            "business_date": d.isoformat(),
            "systems": systems,
            "overall": overall,
            "kpis": ({"ok": run.kpi_ok, "discrepancia": run.kpi_discrepancia,
                      "faltante": run.kpi_faltante} if run else None),
        })
    return {"year": year, "property": property_code, "gate_min_set": sorted(min_set), "days": days}
