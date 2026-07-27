"""Tab 6.4 -- Revenue real diario por depto ("Year to Date"), cargado en bloque
desde la grilla diaria ya agregada (hoja 'Actual' del workbook Weekly, §5.1a
naturaleza -- abre Rooms/Rooms Others y Sustainable Fee/Misc. Rev Others,
igual criterio que engine/revenue.py::weekly_output_column). No reemplaza la
ingesta real día por día (Tabs 1-2) -- es un atajo para poblar el histórico
cuando no hay los XML/Excel de cada día, solo el agregado.
"""
from __future__ import annotations

from datetime import date as date_cls
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Department, Property, RevenueActualDaily

# Columna del Excel (1-indexed) -> cost_center del depto correspondiente.
# "F&B" (col 4) y "Total Daily Revenue" (col 17) se ignoran -- se reconstruyen
# de Food/Beverage/F&B Misc y de la suma de todo lo demás, respectivamente.
COL_TO_DEPT = {
    2: "0110",       # Rooms
    3: "ROOMS-OTH",  # Rooms Others
    5: "FB-FOOD",    # Food
    6: "FB-BEV",     # Beverage
    7: "FB-MISC",    # F&B Misc
    8: "0140",       # SPA
    9: "0150",       # Tours
    10: "0151",      # Retail-Gift Shop
    11: "0152",      # Transportation
    12: "0160",      # Laundry
    13: "0155",      # Innoceana
    14: "0156",      # Crowther Lab
    15: "0170",      # Sustainable Fee
    16: "MISC-REV-OTH",  # Misc. Rev Others
}
COL_ROOMS_SOLD = 24  # "Rooms Sold (RN) Total"
COL_TOTAL_PAX = 25   # "Total Pax"
# Columnas 28 (ADR) y 29 (% Ocupación) no traen encabezado en el Excel fuente,
# pero están calculadas -- de ahí se recupera disponibilidad = RN / Occ% sin
# inventar el dato (verificado constante en 30 para todo el rango Ene-Jun 2026,
# consistente con "Rooms per Day" del presupuesto).
COL_OCC_PCT = 29


async def _property_id(session: AsyncSession, code: str):
    pid = (await session.execute(
        select(Property.id).where(Property.code == code)
    )).scalar_one_or_none()
    if pid is None:
        raise ValueError(f"Propiedad '{code}' no existe.")
    return pid


async def upload_daily_grid(session: AsyncSession, file_bytes: bytes,
                             property_code: str = "COWLCR") -> dict:
    """Sube la grilla diaria tal cual (Date + columnas por naturaleza).
    Reemplaza SOLO los días presentes en el archivo (no todo el año) -- así se
    puede re-subir un archivo actualizado "hasta hoy" sin perder rangos ya
    cargados que el archivo nuevo no vuelva a incluir por error."""
    pid = await _property_id(session, property_code)
    depts = (await session.execute(
        select(Department).where(Department.property_id == pid)
    )).scalars().all()
    by_code = {d.cost_center: d for d in depts}
    missing = [c for c in COL_TO_DEPT.values() if c not in by_code]
    if missing:
        raise ValueError(f"Deptos no encontrados en el catálogo: {missing}")

    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    days_with_data = []
    for r in rows:
        if not r or r[0] is None:
            continue
        d = r[0].date() if hasattr(r[0], "date") else r[0]
        total = sum(float(r[c - 1] or 0) for c in COL_TO_DEPT)
        if total == 0:
            continue  # días futuros / sin cargar en el Excel -- no borrar histórico por esto
        days_with_data.append(d)

    if days_with_data:
        await session.execute(delete(RevenueActualDaily).where(
            RevenueActualDaily.property_id == pid,
            RevenueActualDaily.date.in_(days_with_data),
        ))

    loaded = 0
    for r in rows:
        if not r or r[0] is None:
            continue
        d = r[0].date() if hasattr(r[0], "date") else r[0]
        if d not in days_with_data:
            continue
        rooms_sold = r[COL_ROOMS_SOLD - 1]
        total_pax = r[COL_TOTAL_PAX - 1]
        occ_pct = r[COL_OCC_PCT - 1]
        available_rooms = (
            round(float(rooms_sold) / float(occ_pct))
            if rooms_sold not in (None, "") and occ_pct not in (None, "", 0)
            else None
        )
        for col, dept_code in COL_TO_DEPT.items():
            amt = Decimal(str(r[col - 1] or 0))
            dept = by_code[dept_code]
            is_rooms = dept_code == "0110"
            session.add(RevenueActualDaily(
                property_id=pid, date=d, dept_id=dept.id, amount_usd=amt,
                rooms_sold=Decimal(str(rooms_sold)) if is_rooms and rooms_sold not in (None, "") else None,
                total_pax=Decimal(str(total_pax)) if is_rooms and total_pax not in (None, "") else None,
                available_rooms=Decimal(str(available_rooms)) if is_rooms and available_rooms is not None else None,
            ))
            loaded += 1

    await session.commit()
    return {
        "days_loaded": len(days_with_data),
        "rows_loaded": loaded,
        "date_range": [str(min(days_with_data)), str(max(days_with_data))] if days_with_data else None,
    }


async def daily_view(session: AsyncSession, property_code: str = "COWLCR",
                      date_from: date_cls | None = None, date_to: date_cls | None = None) -> list[dict]:
    pid = await _property_id(session, property_code)
    depts = {d.id: d for d in (await session.execute(
        select(Department).where(Department.property_id == pid)
    )).scalars().all()}
    q = select(RevenueActualDaily).where(RevenueActualDaily.property_id == pid)
    if date_from:
        q = q.where(RevenueActualDaily.date >= date_from)
    if date_to:
        q = q.where(RevenueActualDaily.date <= date_to)
    rows = (await session.execute(q.order_by(RevenueActualDaily.date))).scalars().all()
    return [{
        "date": r.date.isoformat(),
        "dept_code": depts[r.dept_id].cost_center if r.dept_id in depts else None,
        "dept_name": depts[r.dept_id].outlet_name if r.dept_id in depts else None,
        "amount_usd": float(r.amount_usd),
        "rooms_sold": float(r.rooms_sold) if r.rooms_sold is not None else None,
        "total_pax": float(r.total_pax) if r.total_pax is not None else None,
        "available_rooms": float(r.available_rooms) if r.available_rooms is not None else None,
    } for r in rows]
