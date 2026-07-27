"""Inspección de formatos reales (goldens + XMLs) para definir la ingesta.
Corre con el venv:  ./.venv/Scripts/python reference/inspect_inputs.py
"""
import xml.etree.ElementTree as ET
from pathlib import Path

from openpyxl import load_workbook

G = Path(__file__).resolve().parents[1].parent / "goldens"
INP = G / "inputs" / "2026-06-08"


def dump_xlsx(path, max_sheets=40):
    print("\n" + "=" * 70)
    print("XLSX:", path.name)
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print("  ERROR:", e)
        return
    for ws in wb.worksheets[:max_sheets]:
        print(f"  - hoja '{ws.title}'  max={ws.max_row}x{ws.max_column}")
    wb.close()


def dump_xlsx_head(path, sheet, rows=12):
    print(f"\n--- '{path.name}' :: hoja '{sheet}' (primeras {rows} filas) ---")
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= rows:
                break
            cells = [("" if v is None else str(v))[:18] for v in row[:12]]
            print(f"  r{i}: " + " | ".join(cells))
        wb.close()
    except Exception as e:
        print("  ERROR:", e)


def dump_xml_root(path):
    print("\n" + "=" * 70)
    print("XML:", path.name)
    try:
        root = ET.parse(path).getroot()
    except Exception as e:
        print("  ERROR:", e)
        return
    print("  root:", root.tag, dict(root.attrib))
    # primeros 2 hijos con sus subhijos
    for child in list(root)[:2]:
        print(f"  <{child.tag}> attrib={dict(child.attrib)}")
        for gc in list(child)[:14]:
            txt = (gc.text or "").strip()[:30]
            print(f"      <{gc.tag}> {dict(gc.attrib)} = {txt!r}")


if __name__ == "__main__":
    print("########## GOLDENS (estructura de hojas) ##########")
    for f in sorted(G.glob("*.xls*")):
        dump_xlsx(f)

    print("\n########## XMLs de Opera (2026-06-08) ##########")
    for f in sorted(INP.glob("*.xml")) + sorted(INP.glob("*.XML")):
        dump_xml_root(f)
