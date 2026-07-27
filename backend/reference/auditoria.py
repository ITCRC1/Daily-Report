"""
auditoria.py — Script principal de auditoría de ingresos
Corcovado Wilderness Lodge

USO:
    python auditoria.py --fecha 2026-05-31
    python auditoria.py --fecha 2026-05-31 --solo-reporte   (sin subir a Drive)
    python auditoria.py --fecha hoy                          (fecha del día)

FLUJO:
    1. Descarga archivos de Drive → Inputs/YYYY-MM-DD/{opera,integrity,pos}/
    2. Corre la auditoría completa
    3. Genera Excel con todas las pestañas
    4. Sube el Excel a Drive → Reportes/YYYY-MM-DD/
"""

import os, sys, re, json, argparse, datetime, io, tempfile
import xml.etree.ElementTree as ET
import pandas as pd
from pathlib import Path

# ── Google Drive ──────────────────────────────────────────────────────────────
from drive_utils import get_service, get_or_create_day_folders, \
                        list_files_in_folder, download_file, upload_file

# ── Excel ─────────────────────────────────────────────────────────────────────
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
# PALETA
# ══════════════════════════════════════════════════════════════════════════════
DK_GREEN="1B4332"; MD_GREEN="2D6A4F"; LT_GREEN="D8F3DC"
WHITE="FFFFFF"; LT_GRAY="F8F9FA"; MD_GRAY="DEE2E6"; DK_GRAY="495057"
RED_T="C0392B"; RED_BG="F8D7DA"; RED_TXT="721C24"
OK_BG="D4EDDA"; OK_TXT="155724"
BLUE_DK="1A3A5C"; BLUE_MD="2874A6"; BLUE_LT="D6EAF8"
TEAL_DK="0E6655"; TEAL_LT="D0ECE7"
AMBER_DK="7D6608"; AMBER_LT="FCF3CF"; AMBER_MD="F39C12"
SLATE="546E7A"; SLATE_LT="ECEFF1"
PUR_DK="4A235A"; PUR_LT="E8DAEF"
ORG_BG="FFF3CD"; ORG_TXT="856404"

def bdr(c=MD_GRAY):
    s=Side(style='thin',color=c); return Border(left=s,right=s,top=s,bottom=s)

def cs(cell, val, bg=None, fg=DK_GRAY, sz=9, bold=False,
       center=False, wrap=False, italic=False):
    cell.value=val
    cell.font=Font(name='Calibri',bold=bold,italic=italic,color=fg,size=sz)
    if bg: cell.fill=PatternFill("solid",fgColor=bg)
    cell.alignment=Alignment(horizontal='center' if center else 'left',
                              vertical='center',wrap_text=wrap)
    cell.border=bdr()

def nc(cell, val, bg=None, fg=DK_GRAY, sz=9, bold=False, fmt='$#,##0.00'):
    v = val if val is not None else 0
    if isinstance(v,(int,float)) and v < -0.005 and fg==DK_GRAY: fg=RED_T
    cell.value=v
    cell.font=Font(name='Calibri',bold=bold,color=fg,size=sz)
    if bg: cell.fill=PatternFill("solid",fgColor=bg)
    cell.alignment=Alignment(horizontal='right',vertical='center')
    cell.number_format=fmt; cell.border=bdr()

def mhdr(ws, r, c1, c2, val, bg, fg=WHITE, sz=10, bold=True, h=18,
         left=False, wrap=False):
    ws.merge_cells(start_row=r,start_column=c1,end_row=r,end_column=c2)
    cell=ws.cell(r,c1)
    cell.value=val
    cell.font=Font(name='Calibri',bold=bold,color=fg,size=sz)
    cell.fill=PatternFill("solid",fgColor=bg)
    cell.alignment=Alignment(horizontal='left' if left else 'center',
                              vertical='center',wrap_text=wrap)
    ws.row_dimensions[r].height=h
    for col in range(c1,c2+1): ws.cell(r,col).border=bdr()

def sec(ws, r, txt, c1=2, c2=9, bg=MD_GREEN, h=15):
    mhdr(ws,r,c1,c2,f"  ▶  {txt}",bg,h=h,left=True)

# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — DESCARGAR ARCHIVOS DE DRIVE
# ══════════════════════════════════════════════════════════════════════════════
def download_inputs(service, folders, work_dir, fecha_str):
    """Descarga todos los archivos de Drive al directorio temporal."""
    print(f"\n📥 Descargando archivos de Drive para {fecha_str}...")
    downloaded = {'opera':[], 'integrity':[], 'pos':[]}

    for tipo in ['opera','integrity','pos']:
        folder_id = folders[tipo]
        files = list_files_in_folder(service, folder_id)
        if not files:
            print(f"  ⚠  Carpeta '{tipo}' vacía en Drive")
            continue
        for f in files:
            dest = os.path.join(work_dir, tipo, f['name'])
            os.makedirs(os.path.dirname(dest), exist_ok=True)
            download_file(service, f['id'], dest)
            downloaded[tipo].append(dest)
            print(f"  ✓  {tipo}/{f['name']}")

    return downloaded

# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — PARSEAR ARCHIVOS
# ══════════════════════════════════════════════════════════════════════════════
def find_file(paths, pattern):
    """Busca un archivo por patrón regex en lista de paths."""
    for p in paths:
        if re.search(pattern, os.path.basename(p), re.IGNORECASE):
            return p
    return None

def parse_opera(opera_files):
    """Parsea todos los XMLs de Opera."""
    data = {}

    rev_file  = find_file(opera_files, r'REVENUE.*\.xml')
    stat_file = find_file(opera_files, r'STATISTICS.*\.xml')
    hf_files  = [f for f in opera_files if re.search(r'history_forecast', f, re.I)]

    # REVENUE
    if rev_file:
        root = ET.parse(rev_file).getroot()
        headers, details = [], []
        for tt in root.findall('transaction_total'):
            headers.append({
                'tcode': tt.findtext('transaction_code'),
                'description': tt.findtext('description','').strip(),
                'type': tt.get('transaction_type'),
                'total': float(tt.findtext('total_amount') or 0),
                'guest_ledger': float(tt.findtext('total_guest_ledger') or 0),
                'package_ledger': float(tt.findtext('total_package_ledger') or 0),
                'ar_ledger': float(tt.findtext('total_ar_ledger') or 0),
                'deposit_ledger': float(tt.findtext('total_deposit_ledger') or 0),
            })
            for trx in tt.findall('transaction_details/transaction'):
                details.append({
                    'tcode': tt.findtext('transaction_code'),
                    'description': tt.findtext('description','').strip(),
                    'type': tt.get('transaction_type'),
                    'market_code': trx.findtext('market_code',''),
                    'room_class': trx.findtext('room_class',''),
                    'trx_amount': float(trx.findtext('trx_amount') or 0),
                    'trx_guest_ledger': float(trx.findtext('trx_guest_ledger') or 0),
                    'trx_package_ledger': float(trx.findtext('trx_package_ledger') or 0),
                })
        data['opera_hdr'] = pd.DataFrame(headers)
        data['opera_det'] = pd.DataFrame(details)

    # STATISTICS
    if stat_file:
        root = ET.parse(stat_file).getroot()
        stats = []
        for rec in root.findall('statistic_record'):
            stats.append({
                'market_code': rec.findtext('market_code',''),
                'room_class': rec.findtext('room_class',''),
                'room_type': rec.findtext('room_type',''),
                'rooms': int(rec.findtext('rooms') or 0),
                'persons': int(rec.findtext('persons') or 0),
                'noshow': int(rec.findtext('noshow_rooms') or 0),
                'cancel': int(rec.findtext('cancel_rooms') or 0),
            })
        data['stats'] = pd.DataFrame(stats)

    # HISTORY/FORECAST — detectar cuál es total y cuál rooms-only por revenue
    if len(hf_files) >= 2:
        hf_data = []
        for f in hf_files:
            root = ET.parse(f).getroot()
            gd = root.find('.//G_CONSIDERED_DATE')
            if gd is not None:
                hf_data.append({
                    'file': f,
                    'revenue': float(gd.findtext('REVENUE') or 0),
                    'rooms': int(gd.findtext('NO_ROOMS') or 0),
                    'persons': int(gd.findtext('NO_PERSONS') or 0),
                    'arrivals': int(gd.findtext('ARRIVAL_ROOMS') or 0),
                    'departures': int(gd.findtext('DEPARTURE_ROOMS') or 0),
                    'complimentary': int(gd.findtext('COMPLIMENTARY_ROOMS') or 0),
                    'house_use': int(gd.findtext('HOUSE_USE_ROOMS') or 0),
                    'inventory': int(gd.findtext('INVENTORY_ROOMS') or 0),
                    'adr': float(gd.findtext('CF_AVERAGE_ROOM_RATE') or 0),
                    'occupancy': float(gd.findtext('CF_OCCUPANCY') or 0),
                    'ooo': int(gd.findtext('CF_OOO_ROOMS') or 0),
                })
        hf_data.sort(key=lambda x: x['revenue'], reverse=True)
        data['hf_total'] = hf_data[0]   # mayor revenue = Total
        data['hf_rooms'] = hf_data[1]   # menor = Rooms Only

    return data

def parse_integrity(integrity_files):
    xls_file = find_file(integrity_files, r'\.xlsx$')
    if not xls_file:
        return None
    df = pd.read_excel(xls_file, sheet_name='Datos', header=8)
    def et(ref):
        if pd.isna(ref): return None
        m = re.search(r'TCode(?:\s+CXC)?:\s*(\d+)', str(ref))
        return m.group(1) if m else None
    df['tcode'] = df['Referencia'].apply(et)
    df_int = df[df['tcode'].notna()].copy()
    summary = df_int.groupby('tcode').agg(
        int_cr=('Créditos Dol','sum'),
        int_db=('Débitos Dol','sum'),
        cuenta=('Cuenta', lambda x: ' | '.join(x.dropna().unique())),
        nombre=('Nombre cuenta', lambda x: ' | '.join(x.dropna().unique())),
        tc=('T.C.','first'),
    ).reset_index()
    return summary

def _parse_pos_excel(xls_file):
    """Parse Ventas Excel report. Returns single-element list with metadata dict."""
    try:
        df_res = pd.read_excel(xls_file, sheet_name='Resumen Ejecutivo', header=None)
        def find_val(df, keyword):
            for _, row in df.iterrows():
                vals = [v for v in row.values if str(v) not in ['nan','']]
                if any(keyword.lower() in str(v).lower() for v in vals):
                    nums = [v for v in vals if isinstance(v,(int,float))]
                    return nums[0] if nums else 0
            return 0
        meta = {
            'source': 'excel',
            'ventas_netas': find_val(df_res, 'Ventas Netas'),
            'sc': find_val(df_res, 'Cargos de Servicio'),
            'total_dia': find_val(df_res, 'TOTAL VENTAS'),
            'voids': abs(find_val(df_res, 'Anulaciones')),
            'room_charge': 0,
            'rc_detail': [],
        }
        # ── Pestaña 3: Mapeo Simphony → Opera (room charges) ──────────────
        try:
            df_rc = pd.read_excel(xls_file, sheet_name='Mapeo Simphony \u2192 Opera', header=2)
            df_rc.columns = [str(c).strip() for c in df_rc.columns]
            df_rc = df_rc[pd.to_numeric(df_rc['# Check'], errors='coerce').notna()].copy()
            rc_total = 0
            for _, row in df_rc.iterrows():
                monto = float(row.get('Monto Cargado (USD)', 0) or 0)
                rc_total += monto
                meta['rc_detail'].append({
                    'restaurant': str(row.get('Restaurante','')),
                    'employee':   str(row.get('Empleado','')),
                    'check_num':  str(int(row['# Check'])),
                    'hora':       str(row.get('Hora Cierre','')),
                    'monto':      monto,
                })
            meta['room_charge'] = round(rc_total, 2)
        except Exception as e:
            print(f"  \u26a0  No se pudo leer Mapeo Simphony->Opera: {e}")

        # ── Pestaña 2: Detalle de Checks (TODOS los checks cerrados) ──────
        meta['all_checks'] = []
        meta['by_payment'] = []
        meta['by_employee'] = []
        try:
            # Header puede estar en fila 1 o 2 — detectar
            df_det = None
            for h in [1, 2]:
                tmp = pd.read_excel(xls_file, sheet_name='Detalle de Checks', header=h, nrows=3)
                tmp.columns = [str(c).strip() for c in tmp.columns]
                if '# Check' in tmp.columns:
                    df_det = pd.read_excel(xls_file, sheet_name='Detalle de Checks', header=h)
                    df_det.columns = [str(c).strip() for c in df_det.columns]
                    break
            if df_det is not None:
                df_det = df_det[pd.to_numeric(df_det['# Check'], errors='coerce').notna()].copy()
                df_det['Monto (USD)'] = pd.to_numeric(df_det['Monto (USD)'], errors='coerce')
                # Todos los checks
                for _, row in df_det.iterrows():
                    meta['all_checks'].append({
                        'restaurant': str(row.get('Restaurante','')),
                        'employee':   str(row.get('Empleado','')),
                        'check_num':  str(int(row['# Check'])),
                        'hora':       str(row.get('Hora Cierre','')),
                        'forma_pago': str(row.get('Forma de Pago','')),
                        'monto':      float(row.get('Monto (USD)',0) or 0),
                    })
                # Resumen por forma de pago
                fp = df_det.groupby('Forma de Pago')['Monto (USD)'].agg(['count','sum']).reset_index()
                for _, row in fp.iterrows():
                    meta['by_payment'].append({
                        'forma': str(row['Forma de Pago']),
                        'count': int(row['count']),
                        'total': round(float(row['sum']), 2),
                    })
                # Resumen por empleado
                emp = df_det.groupby('Empleado')['Monto (USD)'].agg(['count','sum']).reset_index()
                for _, row in emp.iterrows():
                    meta['by_employee'].append({
                        'empleado': str(row['Empleado']),
                        'count': int(row['count']),
                        'total': round(float(row['sum']), 2),
                    })
        except Exception as e:
            print(f"  \u26a0  No se pudo leer Detalle de Checks: {e}")

        return [meta]
    except Exception as e:
        print(f"  \u26a0  Error leyendo Excel POS: {e}")
        return []

def parse_pos_log(pos_files):
    # Preferir Excel de ventas si existe
    xls_file = find_file(pos_files, r'\.xlsx$')
    if xls_file:
        return _parse_pos_excel(xls_file)
    evt_file = find_file(pos_files, r'\.evt$')
    if not evt_file:
        return []
    with open(evt_file, encoding='utf-8', errors='replace') as f:
        content = f.read()
    def decode(s):
        return s.replace('&lt;','<').replace('&gt;','>').replace('&amp;','&')
    posts = re.findall(
        r'<MonItem DateTime="([^"]+)"[^>]*ObjType="Ifc"[^>]*Type="CommDataIntOut"[^>]*>(.*?)</MonItem>',
        content, re.DOTALL)
    TC = 460.18
    checks = []
    for dt, body in posts:
        clean = decode(body.strip())
        m = re.search(
            r'<PostRequest\s+Date="([^"]+)"\s+Time="([^"]+)"\s+GuestName="([^"]*)"\s+'
            r'[^>]*RoomNum="([^"]*)"\s+TotalAmount="([^"]*)"\s+CheckNum="([^"]*)"\s+'
            r'[^>]*UserID="([^"]*)"[^>]*ServingTime="([^"]*)"', clean)
        if m:
            subtotals = re.findall(r'Subtotal\d+="(\d+)"', clean)
            sc = re.findall(r'ServiceCharge\d+="(\d+)"', clean)
            tax = re.findall(r'Tax\d+="(\d+)"', clean)
            total_crc = int(m.group(5))
            checks.append({
                'check_num': m.group(6),
                'room': m.group(4),
                'pos_date': m.group(1),
                'pos_time': m.group(2),
                'serving': m.group(8),
                'user': m.group(7),
                'total_crc': total_crc,
                'sub_crc': sum(int(v) for v in subtotals),
                'sc_crc': sum(int(v) for v in sc),
                'tax_crc': sum(int(v) for v in tax),
                'total_usd': round(total_crc/TC, 2),
                'sub_usd': round(sum(int(v) for v in subtotals)/TC, 2),
                'sc_usd': round(sum(int(v) for v in sc)/TC, 2),
                'tax_usd': round(sum(int(v) for v in tax)/TC, 2),
            })
    return checks

# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — RECONCILIACIÓN
# ══════════════════════════════════════════════════════════════════════════════
def reconcile(opera_hdr, integrity_df):
    merged = pd.merge(opera_hdr, integrity_df, on='tcode', how='outer', indicator=True)
    def diff(row):
        op = row['total'] if pd.notna(row.get('total')) else 0
        if row.get('type') == 'PAYMENT':
            integ = -(row['int_db'] if pd.notna(row.get('int_db')) else 0)
        else:
            integ = row['int_cr'] if pd.notna(row.get('int_cr')) else 0
        return round(integ - op, 2)
    merged['diferencia'] = merged.apply(diff, axis=1)
    def status(row):
        if row['_merge'] == 'left_only':
            return 'INTERNO' if row.get('type') in ('INTERNAL','PACKAGE') else 'FALTA EN INTEGRITY'
        if row['_merge'] == 'right_only': return 'FALTA EN OPERA'
        return 'OK' if abs(row['diferencia']) < 0.01 else 'DISCREPANCIA'
    merged['estado'] = merged.apply(status, axis=1)
    cat_map = {'REVENUE':'Ingresos','NON REVENUE':'No Ingresos','PAYMENT':'Pagos',
               'INTERNAL':'Interno','PACKAGE':'Paquetes'}
    merged['categoria'] = merged['type'].map(cat_map).fillna('Sin Categoría')
    type_ord = ['REVENUE','NON REVENUE','PAYMENT','INTERNAL','PACKAGE']
    merged['_s'] = merged['type'].apply(lambda x: type_ord.index(x) if x in type_ord else 99)
    return merged.sort_values(['_s','tcode']).reset_index(drop=True)

# ══════════════════════════════════════════════════════════════════════════════
# STEP 4 — GENERAR EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def _sheet_trial_balance(wb, fecha_fmt, gen_str, opera_hdr, integrity_df):
    """Trial Balance — todos los TCode con desglose por ledger."""
    ws = wb.create_sheet("Trial Balance")
    ws.sheet_view.showGridLines = False
    for col,w in {'A':2,'B':8,'C':36,'D':13,'E':13,'F':13,'G':13,'H':13,'I':13,'J':2}.items():
        ws.column_dimensions[col].width = w
    mhdr(ws,2,2,9,"TRIAL BALANCE — OPERA CLOUD",DK_GREEN,sz=13,h=28)
    mhdr(ws,4,2,9,f"Corcovado Wilderness Lodge Costa Rica  |  {fecha_fmt}  |  Generado: {gen_str}",
         LT_GREEN,fg=DK_GRAY,sz=9,bold=False,h=18)
    ws.row_dimensions[5].height=6
    ws.row_dimensions[6].height=18
    for c2,h in enumerate(['TCode','Descripción','Tipo','Total (USD)','Guest Ledger',
                            'Package Ledger','AR Ledger','Deposit Ledger'],2):
        cs(ws.cell(6,c2),h,bg=DK_GREEN,fg=WHITE,bold=True,center=True)
    type_order=['REVENUE','NON REVENUE','PAYMENT','INTERNAL','PACKAGE']
    cat_names={'REVENUE':'INGRESOS','NON REVENUE':'NO INGRESOS','PAYMENT':'PAGOS',
               'INTERNAL':'INTERNO','PACKAGE':'PAQUETES'}
    hdr=opera_hdr.copy()
    hdr['_s']=hdr['type'].apply(lambda x:type_order.index(x) if x in type_order else 99)
    hdr=hdr.sort_values(['_s','tcode']).reset_index(drop=True)
    r=7; current=None
    for _,rd in hdr.iterrows():
        if rd['type']!=current:
            current=rd['type']
            mhdr(ws,r,2,9,f"  ▶  {cat_names.get(current,current)}",MD_GREEN,sz=9,h=14,left=True)
            r+=1
        bg=LT_GREEN if r%2==0 else None
        ws.row_dimensions[r].height=14
        cs(ws.cell(r,2),rd['tcode'],bg=bg,center=True)
        cs(ws.cell(r,3),rd['description'],bg=bg,sz=8)
        cs(ws.cell(r,4),rd['type'],bg=bg,center=True,sz=8)
        nc(ws.cell(r,5),rd['total'],bg=bg)
        nc(ws.cell(r,6),rd['guest_ledger'],bg=bg)
        nc(ws.cell(r,7),rd['package_ledger'],bg=bg)
        nc(ws.cell(r,8),rd['ar_ledger'],bg=bg)
        nc(ws.cell(r,9),rd['deposit_ledger'],bg=bg)
        r+=1
    # Totals
    r+=1
    mhdr(ws,r,2,9,"  ▶  TOTALES POR TIPO",MD_GREEN,sz=9,h=14,left=True); r+=1
    for typ in ['REVENUE','NON REVENUE','PAYMENT']:
        sub=hdr[hdr['type']==typ]
        ws.row_dimensions[r].height=15
        cs(ws.cell(r,3),f"TOTAL {cat_names.get(typ,typ)}",bold=True,bg=SLATE_LT)
        nc(ws.cell(r,5),sub['total'].sum(),bold=True,bg=SLATE_LT)
        nc(ws.cell(r,6),sub['guest_ledger'].sum(),bold=True,bg=SLATE_LT)
        nc(ws.cell(r,7),sub['package_ledger'].sum(),bold=True,bg=SLATE_LT)
        nc(ws.cell(r,8),sub['ar_ledger'].sum(),bold=True,bg=SLATE_LT)
        nc(ws.cell(r,9),sub['deposit_ledger'].sum(),bold=True,bg=SLATE_LT)
        r+=1
    ws.freeze_panes='B7'


def _sheet_ledgers(wb, fecha_fmt, gen_str, opera_hdr, hf_total):
    """Estado de Ledgers."""
    ws = wb.create_sheet("Ledgers")
    ws.sheet_view.showGridLines = False
    for col,w in {'A':2,'B':40,'C':10,'D':10,'E':14,'F':16,'G':22}.items():
        ws.column_dimensions[col].width = w
    mhdr(ws,2,2,7,f"ESTADO DE LEDGERS — {fecha_fmt}",DK_GREEN,sz=13,h=28)
    mhdr(ws,4,2,7,f"Corcovado Wilderness Lodge Costa Rica  |  {fecha_fmt}  |  Generado: {gen_str}",
         LT_GREEN,fg=DK_GRAY,sz=9,bold=False,h=18)
    ws.row_dimensions[5].height=6
    # Guest Ledger
    accom = opera_hdr[opera_hdr['type']=='REVENUE']['total'].sum()
    pay   = opera_hdr[opera_hdr['type']=='PAYMENT']['guest_ledger'].sum()
    r=6
    mhdr(ws,r,2,7,"  GUEST LEDGER",MD_GREEN,sz=10,h=16,left=True); r+=1
    rev_gl = opera_hdr['guest_ledger'].sum()
    for lbl,val in [("    Cargos del día (Guest Ledger)", rev_gl),
                    ("    Balance Today (Guest Ledger)", rev_gl)]:
        ws.row_dimensions[r].height=15
        cs(ws.cell(r,2),lbl)
        nc(ws.cell(r,6),val)
        r+=1
    ws.freeze_panes='B6'


def _sheet_estadisticas(wb, fecha_fmt, gen_str, stats, opera_det, hf_total):
    """Estadísticas de Ocupación."""
    ws = wb.create_sheet("Estadisticas Ocupacion")
    ws.sheet_view.showGridLines = False
    for col,w in {'A':2,'B':14,'C':22,'D':12,'E':10,'F':14,'G':12,'H':12,'I':10}.items():
        ws.column_dimensions[col].width = w
    mhdr(ws,2,2,9,f"ESTADÍSTICAS DE OCUPACIÓN — {fecha_fmt}",DK_GREEN,sz=13,h=28)
    mhdr(ws,4,2,9,f"Corcovado Wilderness Lodge Costa Rica  |  {fecha_fmt}  |  Generado: {gen_str}",
         LT_GREEN,fg=DK_GRAY,sz=9,bold=False,h=18)
    ws.row_dimensions[5].height=6

    total_rooms = stats['rooms'].sum()
    total_pax   = stats['persons'].sum()
    accom_det   = opera_det[opera_det['tcode']=='1000']
    accom_rev   = accom_det['trx_amount'].sum()
    adr         = accom_rev/total_rooms if total_rooms else 0

    def kpi(r,c,lbl,val,bg,fmt='$#,##0.00'):
        ws.merge_cells(start_row=r,start_column=c,end_row=r,end_column=c+1)
        ws.merge_cells(start_row=r+1,start_column=c,end_row=r+1,end_column=c+1)
        lc=ws.cell(r,c); lc.value=lbl
        lc.font=Font(name='Calibri',color=WHITE,size=8); lc.fill=PatternFill("solid",fgColor=bg)
        lc.alignment=Alignment(horizontal='center',vertical='center'); lc.border=bdr()
        vc=ws.cell(r+1,c); vc.value=val
        vc.font=Font(name='Calibri',bold=True,color=WHITE,size=13); vc.fill=PatternFill("solid",fgColor=bg)
        vc.alignment=Alignment(horizontal='center',vertical='center'); vc.number_format=fmt; vc.border=bdr()
        ws.row_dimensions[r].height=17; ws.row_dimensions[r+1].height=24
    kpi(6,2,"HABITACIONES VENDIDAS",int(total_rooms),MD_GREEN,fmt='0')
    kpi(6,4,"TOTAL PAX",int(total_pax),"5C7A8C",fmt='0')
    kpi(6,6,"ADR PROMEDIO",round(adr,2),BLUE_MD)
    kpi(6,8,"INGRESO ALOJAMIENTO",round(accom_rev,2),DK_GREEN)
    ws.row_dimensions[9].height=8

    # By market code
    r=10
    mhdr(ws,r,2,9,"  ▶  DETALLE POR MARKET CODE",MD_GREEN,sz=9,h=14,left=True); r+=1
    ws.row_dimensions[r].height=16
    for c2,h in enumerate(['Market Code','Descripción','Habitaciones','PAX','Revenue (USD)','ADR','% Revenue'],2):
        cs(ws.cell(r,c2),h,bg=DK_GREEN,fg=WHITE,bold=True,center=True)
    r+=1
    mc_names={'TAFIT':'Agencia de Viajes','WEB':'Website','DIR':'Directo','COM':'Complementario',
              'CORP':'Corporativo','GRP':'Grupo'}
    mc_stat=stats.groupby('market_code')[['rooms','persons']].sum()
    mc_rev=accom_det.groupby('market_code')['trx_amount'].sum()
    for mc in mc_stat.index:
        rooms=int(mc_stat.loc[mc,'rooms']); pax=int(mc_stat.loc[mc,'persons'])
        rev=mc_rev.get(mc,0); adr_mc=rev/rooms if rooms else 0
        pct=rev/accom_rev if accom_rev else 0
        bg=LT_GREEN if r%2==0 else None
        ws.row_dimensions[r].height=15
        cs(ws.cell(r,2),mc,bg=bg,center=True,bold=True)
        cs(ws.cell(r,3),mc_names.get(mc,mc),bg=bg,sz=8)
        nc(ws.cell(r,4),rooms,bg=bg,fmt='0'); nc(ws.cell(r,5),pax,bg=bg,fmt='0')
        nc(ws.cell(r,6),rev,bg=bg); nc(ws.cell(r,7),round(adr_mc,2),bg=bg)
        nc(ws.cell(r,8),pct,bg=bg,fmt='0.0%')
        r+=1
    # Total
    ws.row_dimensions[r].height=16
    cs(ws.cell(r,2),'TOTAL',bold=True,bg=SLATE_LT)
    nc(ws.cell(r,4),int(total_rooms),bold=True,bg=SLATE_LT,fmt='0')
    nc(ws.cell(r,5),int(total_pax),bold=True,bg=SLATE_LT,fmt='0')
    nc(ws.cell(r,6),accom_rev,bold=True,bg=SLATE_LT)
    nc(ws.cell(r,7),round(adr,2),bold=True,bg=SLATE_LT)
    nc(ws.cell(r,8),1.0,bold=True,bg=SLATE_LT,fmt='0.0%')
    r+=2
    # By room class
    mhdr(ws,r,2,9,"  ▶  DETALLE POR TIPO DE HABITACIÓN (ROOM CLASS)",MD_GREEN,sz=9,h=14,left=True); r+=1
    ws.row_dimensions[r].height=16
    for c2,h in enumerate(['Room Class','Room Type','Habitaciones','PAX','Revenue (USD)','ADR','% Rev'],2):
        cs(ws.cell(r,c2),h,bg=DK_GREEN,fg=WHITE,bold=True,center=True)
    r+=1
    rc_stat=stats.groupby(['room_class','room_type'])[['rooms','persons']].sum().reset_index()
    accom_rc=accom_det.groupby('room_class')['trx_amount'].sum()
    for _,row in rc_stat.iterrows():
        rc=row['room_class']; rooms=int(row['rooms']); pax=int(row['persons'])
        rev=accom_rc.get(rc,0); adr_rc=rev/rooms if rooms else 0
        pct=rev/accom_rev if accom_rev else 0
        bg=LT_GREEN if r%2==0 else None
        ws.row_dimensions[r].height=15
        cs(ws.cell(r,2),rc,bg=bg,center=True,bold=True)
        cs(ws.cell(r,3),row['room_type'],bg=bg,center=True,sz=8)
        nc(ws.cell(r,4),rooms,bg=bg,fmt='0'); nc(ws.cell(r,5),pax,bg=bg,fmt='0')
        nc(ws.cell(r,6),rev,bg=bg); nc(ws.cell(r,7),round(adr_rc,2),bg=bg)
        nc(ws.cell(r,8),pct,bg=bg,fmt='0.0%')
        r+=1
    ws.freeze_panes='B6'


def _sheet_otb(wb, fecha_fmt, gen_str, opera_hdr, opera_det, stats, hf_total, hf_rooms):
    """OTB vs Revenue."""
    ws = wb.create_sheet("OTB vs Revenue")
    ws.sheet_view.showGridLines = False
    for col,w in {'A':2,'B':34,'C':16,'D':16,'E':14,'F':16,'G':14,'H':14}.items():
        ws.column_dimensions[col].width = w
    mhdr(ws,2,2,7,"ON THE BOOKS vs REVENUE REAL",DK_GREEN,sz=13,h=28)
    mhdr(ws,3,2,7,f"Corcovado Wilderness Lodge  |  {fecha_fmt}  |  ¿Por qué ADR OTB ≠ ADR Rooms Only?",
         LT_GREEN,fg=DK_GRAY,sz=8,bold=False,h=16)
    ws.row_dimensions[4].height=6
    total_rev = hf_total.get('revenue',0)
    rooms_rev = hf_rooms.get('revenue',0)
    diff_rev  = total_rev - rooms_rev
    rooms     = hf_total.get('rooms',1) or 1
    def kpi(r,c,lbl,val,bg):
        ws.merge_cells(start_row=r,start_column=c,end_row=r,end_column=c+1)
        ws.merge_cells(start_row=r+1,start_column=c,end_row=r+1,end_column=c+1)
        lc=ws.cell(r,c); lc.value=lbl
        lc.font=Font(name='Calibri',color=WHITE,size=8,bold=False); lc.fill=PatternFill("solid",fgColor=bg)
        lc.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); lc.border=bdr()
        vc=ws.cell(r+1,c); vc.value=val
        vc.font=Font(name='Calibri',bold=True,color=WHITE,size=13); vc.fill=PatternFill("solid",fgColor=bg)
        vc.alignment=Alignment(horizontal='center',vertical='center'); vc.number_format='$#,##0.00'; vc.border=bdr()
        ws.row_dimensions[r].height=24; ws.row_dimensions[r+1].height=24
    kpi(5,2,"REVENUE TOTAL (OTB)\nTodos los ingresos",total_rev,MD_GREEN)
    kpi(5,4,"REVENUE ROOMS ONLY\nSolo alojamiento",rooms_rev,"5C7A8C")
    kpi(5,6,"DIFERENCIA\nNo-Alojamiento",diff_rev,AMBER_MD)
    ws.row_dimensions[7].height=6
    kpi(8,2,"ADR ON THE BOOKS\n(Rev Total / Hab)",round(total_rev/rooms,2),MD_GREEN)
    kpi(8,4,"ADR ROOMS ONLY\n(Rev Aloj / Hab)",round(rooms_rev/rooms,2),"5C7A8C")
    kpi(8,6,"DIFERENCIA ADR\nNo-aloj. por hab.",round((total_rev-rooms_rev)/rooms,2),AMBER_MD)
    ws.row_dimensions[10].height=8
    # Composición
    r=11
    mhdr(ws,r,2,7,"  ▶  COMPOSICIÓN DEL REVENUE TOTAL",MD_GREEN,sz=9,h=14,left=True); r+=1
    ws.row_dimensions[r].height=16
    for c2,h in enumerate(['Categoría','Revenue (USD)','% del Total','ADR Contrib.','Tipo'],2):
        cs(ws.cell(r,c2),h,bg=DK_GREEN,fg=WHITE,bold=True,center=True)
    r+=1
    fb_terra=['2139','2140','2142','2143','2149','2161']
    fb_bosque=['2224','2225','2227','2228','2233','2234','2245','2246','2249']
    pkg=['2500','2502','2504','2507']
    tours=['3400','3405','3406','3407','3411']
    def sum_tc(tcodes):
        return opera_hdr[opera_hdr['tcode'].isin(tcodes)]['total'].sum()
    accom=opera_hdr[opera_hdr['tcode']=='1000']['total'].sum()
    retail=opera_hdr[opera_hdr['tcode'].isin(['2320','2321','2324','2330','2490'])]['total'].sum()
    sustain=sum_tc(['3005']); boat=sum_tc(['3320'])
    comps=[
        ("🛏 Accommodation",accom,"✓ Incluido en Rooms ADR"),
        ("🍽 F&B Terra Kitchen",sum_tc(fb_terra),"Solo en Total ADR"),
        ("🍽 F&B El Bosque",sum_tc(fb_bosque),"Solo en Total ADR"),
        ("📦 Packages F&B",sum_tc(pkg),"Solo en Total ADR"),
        ("🛍 Retail",retail,"Solo en Total ADR"),
        ("🌿 Sustainable Fee",sustain,"Solo en Total ADR"),
        ("⛵ Boat Transport",boat,"Solo en Total ADR"),
        ("🥾 Tours",sum_tc(tours),"Solo en Total ADR"),
    ]
    for lbl,val,tipo in comps:
        bg=LT_GREEN if r%2==0 else None
        pct=val/total_rev if total_rev else 0
        ws.row_dimensions[r].height=15
        cs(ws.cell(r,2),lbl,bg=bg)
        nc(ws.cell(r,3),val,bg=bg)
        nc(ws.cell(r,4),pct,bg=bg,fmt='0.0%')
        nc(ws.cell(r,5),round(val/rooms,2),bg=bg)
        cs(ws.cell(r,6),tipo,bg=bg,sz=8,fg="666666")
        r+=1
    ws.row_dimensions[r].height=16
    cs(ws.cell(r,2),'TOTAL REVENUE',bold=True,bg=SLATE_LT)
    nc(ws.cell(r,3),total_rev,bold=True,bg=SLATE_LT)
    nc(ws.cell(r,4),1.0,bold=True,bg=SLATE_LT,fmt='0.0%')
    nc(ws.cell(r,5),round(total_rev/rooms,2),bold=True,bg=SLATE_LT)
    cs(ws.cell(r,6),'→ ADR ON THE BOOKS',bold=True,bg=SLATE_LT,sz=8)
    ws.freeze_panes='B5'


def _sheet_market_code(wb, fecha_fmt, gen_str, opera_hdr, opera_det):
    """Ingresos por Market Code (pivot)."""
    ws = wb.create_sheet("Ingresos x Market Code")
    ws.sheet_view.showGridLines = False
    market_codes = sorted([mc for mc in opera_det['market_code'].unique() if mc])
    ncols = 3 + len(market_codes)
    widths = {'A':2,'B':8,'C':34}
    for i,mc in enumerate(market_codes):
        widths[chr(68+i)] = 13
    for col,w in widths.items():
        ws.column_dimensions[col].width = w
    last_col = 3 + len(market_codes)
    mhdr(ws,2,2,last_col,f"INGRESOS POR MARKET CODE — {fecha_fmt}",DK_GREEN,sz=13,h=28)
    mhdr(ws,4,2,last_col,f"Corcovado Wilderness Lodge Costa Rica  |  {fecha_fmt}  |  Generado: {gen_str}",
         LT_GREEN,fg=DK_GRAY,sz=9,bold=False,h=18)
    ws.row_dimensions[5].height=6
    ws.row_dimensions[6].height=18
    cs(ws.cell(6,2),'TCode',bg=DK_GREEN,fg=WHITE,bold=True,center=True)
    cs(ws.cell(6,3),'Descripción',bg=DK_GREEN,fg=WHITE,bold=True,center=True)
    for i,mc in enumerate(market_codes):
        cs(ws.cell(6,4+i),mc,bg=DK_GREEN,fg=WHITE,bold=True,center=True)
    # Revenue rows
    rev_tcodes = opera_hdr[opera_hdr['type']=='REVENUE']['tcode'].tolist()
    r=7
    for tc in rev_tcodes:
        desc = opera_hdr[opera_hdr['tcode']==tc]['description'].values[0]
        bg=LT_GREEN if r%2==0 else None
        ws.row_dimensions[r].height=14
        cs(ws.cell(r,2),tc,bg=bg,center=True)
        cs(ws.cell(r,3),desc,bg=bg,sz=8)
        for i,mc in enumerate(market_codes):
            val=opera_det[(opera_det['tcode']==tc)&(opera_det['market_code']==mc)]['trx_amount'].sum()
            nc(ws.cell(r,4+i),val if val else None,bg=bg)
        r+=1
    # Total
    ws.row_dimensions[r].height=16
    cs(ws.cell(r,3),'TOTAL INGRESOS',bold=True,bg=SLATE_LT)
    for i,mc in enumerate(market_codes):
        tot=opera_det[(opera_det['type']=='REVENUE')&(opera_det['market_code']==mc)]['trx_amount'].sum()
        nc(ws.cell(r,4+i),tot,bold=True,bg=SLATE_LT)
    ws.freeze_panes='D7'


def build_excel(fecha_str, opera_data, integrity_df, pos_checks, output_path):
    """Genera el Excel completo de auditoría con las 11 pestañas."""
    import datetime as _dt
    fecha_fmt = '/'.join(reversed(fecha_str.split('-')))  # YYYY-MM-DD -> DD/MM/YYYY
    gen_str   = _dt.date.today().strftime('%d/%m/%Y')
    wb = Workbook()

    merged = reconcile(opera_data['opera_hdr'], integrity_df)
    op_rev   = opera_data['opera_hdr'][opera_data['opera_hdr']['type']=='REVENUE']['total'].sum()
    op_nonrev= opera_data['opera_hdr'][opera_data['opera_hdr']['type']=='NON REVENUE']['total'].sum()
    op_pay   = opera_data['opera_hdr'][opera_data['opera_hdr']['type']=='PAYMENT']['total'].sum()
    total_ok  = (merged['estado']=='OK').sum()
    total_disc= (merged['estado']=='DISCREPANCIA').sum()
    total_miss= merged['estado'].isin(['FALTA EN INTEGRITY','FALTA EN OPERA']).sum()

    # ── HOJA 1: RESUMEN ──────────────────────────────────────────────────────
    ws = wb.active; ws.title="Resumen Ejecutivo"
    ws.sheet_view.showGridLines=False
    for col,w in {'A':2,'B':26,'C':16,'D':16,'E':16,'F':16,'G':14,'H':18}.items():
        ws.column_dimensions[col].width=w
    ws.row_dimensions[1].height=6
    mhdr(ws,2,2,8,f"AUDITORÍA DE INGRESOS — RECONCILIACIÓN OPERA ↔ INTEGRITY",
         DK_GREEN,sz=14,h=30)
    mhdr(ws,3,2,8,f"Corcovado Wilderness Lodge  |  {fecha_fmt}  |  "
         f"Generado: {datetime.date.today().strftime('%d/%m/%Y')}",
         LT_GREEN,fg=DK_GRAY,sz=9,bold=False,h=18)
    ws.row_dimensions[4].height=8

    # KPIs
    def kpi(ws,r,c,lbl,val,bg,fmt='$#,##0.00'):
        ws.merge_cells(start_row=r,start_column=c,end_row=r,end_column=c+1)
        ws.merge_cells(start_row=r+1,start_column=c,end_row=r+1,end_column=c+1)
        lc=ws.cell(r,c); lc.value=lbl
        lc.font=Font(name='Calibri',color=WHITE,size=8)
        lc.fill=PatternFill("solid",fgColor=bg)
        lc.alignment=Alignment(horizontal='center',vertical='center'); lc.border=bdr()
        vc=ws.cell(r+1,c); vc.value=val
        vc.font=Font(name='Calibri',bold=True,color=WHITE,size=13)
        vc.fill=PatternFill("solid",fgColor=bg)
        vc.alignment=Alignment(horizontal='center',vertical='center')
        vc.number_format=fmt; vc.border=bdr()
        ws.row_dimensions[r].height=17; ws.row_dimensions[r+1].height=24

    kpi(ws,5,2,"INGRESOS TOTALES",op_rev,MD_GREEN)
    kpi(ws,5,4,"NO INGRESOS",op_nonrev,"5C7A8C")
    kpi(ws,5,6,"PAGOS",abs(op_pay),BLUE_MD)
    ws.row_dimensions[8].height=8
    kpi(ws,9,2,"✓ CUADRADAS",total_ok,"1A6B3C",fmt='0')
    kpi(ws,9,4,"⚠ DISCREPANCIAS",total_disc,"8B6000" if total_disc>0 else "1A6B3C",fmt='0')
    kpi(ws,9,6,"✗ FALTANTES",total_miss,RED_T if total_miss>0 else "1A6B3C",fmt='0')
    ws.row_dimensions[12].height=8

    # Tabla resumen por categoría
    r=13
    ws.row_dimensions[r].height=20
    for c2,h in enumerate(['CATEGORÍA','OPERA (USD)','INTEGRITY (USD)','DIFERENCIA','ESTADO'],2):
        cs(ws.cell(r,c2),h,bg=DK_GREEN,fg=WHITE,bold=True,center=True)
    r+=1
    for cat in ['Ingresos','No Ingresos','Pagos']:
        sub=merged[merged['categoria']==cat]
        op_t=sub['total'].fillna(0).sum()
        int_t=(-sub['int_db'].fillna(0).sum() if cat=='Pagos' else sub['int_cr'].fillna(0).sum())
        dif=round(int_t-op_t,2)
        bg=OK_BG if abs(dif)<0.01 else ORG_BG
        tc2=OK_TXT if abs(dif)<0.01 else ORG_TXT
        icon="✓ Cuadrado" if abs(dif)<0.01 else f"⚠ ${dif:,.2f}"
        ws.row_dimensions[r].height=18
        cs(ws.cell(r,2),cat,bold=True,bg=bg,fg=tc2)
        nc(ws.cell(r,3),op_t,bg=bg)
        nc(ws.cell(r,4),int_t,bg=bg)
        nc(ws.cell(r,5),dif,bold=True,bg=bg,fg=tc2)
        cs(ws.cell(r,6),icon,bold=True,center=True,bg=bg,fg=tc2)
        r+=1

    # POS summary si hay checks
    if pos_checks:
        ws.row_dimensions[r].height=8; r+=1
        if isinstance(pos_checks[0], dict) and pos_checks[0].get('source')=='excel':
            meta = pos_checks[0]
            rc = meta.get('room_charge', 0)
            n_rc = len(meta.get('rc_detail', []))
            mhdr(ws,r,2,8,f"  SIMPHONY POS — {n_rc} room charges → Opera  |  "
                 f"Total: ${rc:.2f} USD  |  Voids: ${meta.get('voids',0):.2f}",
                 TEAL_DK,h=16,left=True); r+=1
        else:
            total_pos = sum(c.get('total_usd',0) for c in pos_checks)
            mhdr(ws,r,2,8,f"  SIMPHONY POS — {len(pos_checks)} checks posteados  |  "
                 f"Total: ${total_pos:.2f} USD",
                 TEAL_DK,h=16,left=True); r+=1

    # Nota
    ws.row_dimensions[r].height=8; r+=1
    ws.merge_cells(f'B{r}:H{r}')
    c2=ws.cell(r,2)
    c2.value="ℹ  Códigos 9910/9990 son internos Opera — no generan asiento en Integrity. Diferencias <$0.01 = redondeo."
    c2.font=Font(name='Calibri',italic=True,color="6C757D",size=8)
    c2.fill=PatternFill("solid",fgColor=LT_GRAY)
    c2.alignment=Alignment(horizontal='left',vertical='center',wrap_text=True)
    ws.row_dimensions[r].height=20
    for col in range(2,9): ws.cell(r,col).border=bdr()

    # ── HOJA 2: DETALLE RECONCILIACIÓN ───────────────────────────────────────
    ws2=wb.create_sheet("Detalle Reconciliacion")
    ws2.sheet_view.showGridLines=False
    for col,w in {'A':2,'B':8,'C':32,'D':12,'E':14,'F':14,'G':14,'H':28,'I':36,'J':14}.items():
        ws2.column_dimensions[col].width=w
    mhdr(ws2,1,2,9,f"DETALLE RECONCILIACIÓN — {fecha_fmt}",DK_GREEN,sz=13,h=28)
    ws2.row_dimensions[2].height=6
    ws2.row_dimensions[3].height=20
    for c2,h in enumerate(['TCode','Descripción','Tipo','Opera','Integrity','Diferencia',
                            'Cuenta Contable','Nombre Cuenta','Estado'],2):
        cs(ws2.cell(3,c2),h,bg=DK_GREEN,fg=WHITE,bold=True,center=True)

    current_cat=None; r=4
    type_colors={'REVENUE':(LT_GREEN,OK_TXT),'NON REVENUE':(BLUE_LT,BLUE_DK),
                 'PAYMENT':(PUR_LT,PUR_DK),'INTERNAL':(LT_GRAY,DK_GRAY),'PACKAGE':(AMBER_LT,AMBER_DK)}
    for _,rd in merged.iterrows():
        cat=rd['categoria']
        if cat!=current_cat:
            current_cat=cat
            sec(ws2,r,cat.upper(),c1=2,c2=9); r+=1
        estado=rd['estado']
        if estado=='OK': rb,et2,ei=None,OK_TXT,"✓ OK"
        elif estado=='DISCREPANCIA': rb,et2,ei=ORG_BG,ORG_TXT,"⚠ DISCREPANCIA"
        elif 'FALTA' in estado: rb,et2,ei=RED_BG,RED_TXT,f"✗ {estado}"
        else: rb,et2,ei=LT_GRAY,"888888","— INTERNO"

        if rd.get('type')=='PAYMENT': ia=-(rd['int_db'] if pd.notna(rd.get('int_db')) else 0)
        else: ia=rd['int_cr'] if pd.notna(rd.get('int_cr')) else None
        dv=rd['diferencia'] if pd.notna(rd.get('diferencia')) else None

        ws2.row_dimensions[r].height=16
        cs(ws2.cell(r,2),rd['tcode'],center=True,bg=rb)
        cs(ws2.cell(r,3),rd.get('description',''),bg=rb)
        cs(ws2.cell(r,4),rd.get('type',''),center=True,bg=rb)
        nc(ws2.cell(r,5),rd.get('total'),bg=rb)
        nc(ws2.cell(r,6),ia,bg=rb)
        nc(ws2.cell(r,7),dv,bold=(dv is not None and abs(dv)>=0.01),bg=rb)
        cs(ws2.cell(r,8),str(rd.get('cuenta','')) if pd.notna(rd.get('cuenta')) else '',bg=rb,sz=8)
        cs(ws2.cell(r,9),str(rd.get('nombre','')) if pd.notna(rd.get('nombre')) else '',bg=rb,sz=8)
        ec=ws2.cell(r,10); ec.value=ei
        ec.font=Font(name='Calibri',bold=True,color=et2,size=9)
        ec.fill=PatternFill("solid",fgColor=OK_BG if estado=='OK' else(ORG_BG if estado=='DISCREPANCIA' else(RED_BG if 'FALTA' in estado else LT_GRAY)))
        ec.alignment=Alignment(horizontal='center',vertical='center'); ec.border=bdr()
        r+=1
    ws2.freeze_panes='B4'

    # ── HOJA 3: DISCREPANCIAS ─────────────────────────────────────────────────
    ws3=wb.create_sheet("Discrepancias")
    ws3.sheet_view.showGridLines=False
    disc_df=merged[merged['estado'].isin(['DISCREPANCIA','FALTA EN INTEGRITY','FALTA EN OPERA'])]
    if disc_df.empty:
        mhdr(ws3,1,2,9,"✓  SIN DISCREPANCIAS — TODOS LOS REGISTROS CUADRAN",MD_GREEN,sz=13,h=30)
        ws3.merge_cells('B3:I6')
        c2=ws3.cell(3,2); c2.value="Todos los códigos viajaron correctamente de Opera a Integrity."
        c2.font=Font(name='Calibri',color=OK_TXT,size=11)
        c2.fill=PatternFill("solid",fgColor=OK_BG)
        c2.alignment=Alignment(horizontal='center',vertical='center')
        ws3.row_dimensions[3].height=60
        for col in range(2,10): ws3.cell(3,col).border=bdr()
    else:
        mhdr(ws3,1,2,10,f"⚠  {len(disc_df)} DISCREPANCIAS — ACCIÓN REQUERIDA",RED_T,sz=13,h=28)
        r=3
        for c2,h in enumerate(['TCode','Descripción','Tipo','Opera','Integrity',
                                'Diferencia','Cuenta','Nombre Cuenta','Estado','Notas Auditor'],2):
            cs(ws3.cell(r,c2),h,bg=RED_T,fg=WHITE,bold=True,center=True)
        r+=1
        for _,rd in disc_df.iterrows():
            ws3.row_dimensions[r].height=18
            cs(ws3.cell(r,2),rd['tcode'],center=True,bg=RED_BG)
            cs(ws3.cell(r,3),rd.get('description',''),bg=RED_BG)
            cs(ws3.cell(r,4),rd.get('type',''),center=True,bg=RED_BG)
            nc(ws3.cell(r,5),rd.get('total'),bg=RED_BG)
            nc(ws3.cell(r,6),rd.get('int_cr'),bg=RED_BG)
            nc(ws3.cell(r,7),rd.get('diferencia'),bold=True,bg=RED_BG,fg=RED_T)
            cs(ws3.cell(r,8),str(rd.get('cuenta',''))[:30],bg=RED_BG,sz=8)
            cs(ws3.cell(r,9),str(rd.get('nombre',''))[:40],bg=RED_BG,sz=8)
            cs(ws3.cell(r,10),rd['estado'],bold=True,center=True,bg=RED_BG,fg=RED_TXT)
            cs(ws3.cell(r,11),"",bg=AMBER_LT)
            r+=1
        ws3.freeze_panes='B4'

    # ── HOJA 4: POS (maneja Excel de Ventas o EVT) ────────────────────────────
    if pos_checks:
        ws4=wb.create_sheet("Simphony POS")
        ws4.sheet_view.showGridLines=False
        is_excel = isinstance(pos_checks[0], dict) and pos_checks[0].get('source')=='excel'

        if is_excel:
            # ── POS desde Excel de Ventas ─────────────────────────────────────
            meta = pos_checks[0]
            for col,w in {'A':2,'B':22,'C':13,'D':13,'E':13,'F':13,'G':13,'H':13,'I':13,'J':13,'K':2}.items():
                ws4.column_dimensions[col].width=w
            mhdr(ws4,2,2,10,f"SIMPHONY POS — REPORTE DE VENTAS — {fecha_fmt}",TEAL_DK,sz=13,h=28)
            mhdr(ws4,3,2,10,f"Corcovado Wilderness Lodge  |  {fecha_fmt}  |  Room Charges enviados a Opera",
                 LT_GREEN,fg=DK_GRAY,sz=8,bold=False,h=16)
            ws4.row_dimensions[4].height=8
            # Resumen
            r=5
            mhdr(ws4,r,2,10,"  RESUMEN GENERAL DEL DÍA",TEAL_DK,left=True,h=15); r+=1
            ws4.row_dimensions[r].height=16
            for c2,h in enumerate(['Concepto','Monto (USD)','Nota'],2):
                cs(ws4.cell(r,c2),h,bg=TEAL_DK,fg=WHITE,bold=True,center=True)
            ws4.merge_cells(start_row=r,start_column=4,end_row=r,end_column=10)
            cs(ws4.cell(r,4),'Nota',bg=TEAL_DK,fg=WHITE,bold=True)
            for col in range(4,11): ws4.cell(r,col).border=bdr()
            r+=1
            for label,val,note,bg,fg2 in [
                ("Ventas Netas",meta.get('ventas_netas',0),"Sin descuentos",LT_GREEN,OK_TXT),
                ("Cargos de Servicio 10%",meta.get('sc',0),"Sobre consumo directo",TEAL_LT,TEAL_DK),
                ("TOTAL VENTAS DEL DÍA",meta.get('total_dia',0),"Neto + SC",TEAL_LT,DK_GREEN),
                ("Anulaciones (Voids)",-abs(meta.get('voids',0)),"⚠ NO deben estar en Opera",AMBER_LT,AMBER_DK),
                ("Room Charges → Opera",meta.get('room_charge',0),"Enviados vía interfaz POS",LT_GREEN,OK_TXT),
            ]:
                ws4.row_dimensions[r].height=16; bold='TOTAL' in label
                cs(ws4.cell(r,2),label,bold=bold,bg=bg,fg=fg2)
                nc(ws4.cell(r,3),val,bg=bg,fg=RED_T if val<0 else fg2,bold=bold)
                ws4.merge_cells(start_row=r,start_column=4,end_row=r,end_column=10)
                cs(ws4.cell(r,4),note,bg=LT_GRAY,fg="888888",sz=8)
                for col in range(4,11): ws4.cell(r,col).border=bdr()
                r+=1
            # Detalle room charges — DINÁMICO
            rc_detail = meta.get('rc_detail', [])
            if rc_detail:
                r+=1
                mhdr(ws4,r,2,10,f"  ROOM CHARGES CONFIRMADOS POR OPERA ({len(rc_detail)} checks)",
                     MD_GREEN,left=True,h=15); r+=1
                ws4.row_dimensions[r].height=16
                for c2,h in enumerate(['Restaurante','Empleado','# Check','Hora','Monto USD','Estado'],2):
                    cs(ws4.cell(r,c2),h,bg=MD_GREEN,fg=WHITE,bold=True,center=True)
                ws4.merge_cells(start_row=r,start_column=8,end_row=r,end_column=10)
                cs(ws4.cell(r,8),'Estado Opera',bg=MD_GREEN,fg=WHITE,bold=True,center=True)
                for col in range(8,11): ws4.cell(r,col).border=bdr()
                r+=1
                for i,chk in enumerate(rc_detail):
                    bg=TEAL_LT if i%2==0 else None
                    ws4.row_dimensions[r].height=15
                    cs(ws4.cell(r,2),chk.get('restaurant',''),bg=bg)
                    cs(ws4.cell(r,3),chk.get('employee',''),bg=bg,sz=8)
                    cs(ws4.cell(r,4),chk.get('check_num',''),bg=bg,center=True,bold=True)
                    cs(ws4.cell(r,5),chk.get('hora',''),bg=bg,center=True)
                    nc(ws4.cell(r,6),chk.get('monto',0),bg=bg)
                    ws4.merge_cells(start_row=r,start_column=7,end_row=r,end_column=10)
                    cs(ws4.cell(r,7),"✓ AnswerStat=OK",bg=OK_BG,fg=OK_TXT,bold=True,center=True)
                    for col in range(7,11): ws4.cell(r,col).border=bdr()
                    r+=1
                # Total
                ws4.row_dimensions[r].height=16
                cs(ws4.cell(r,2),f'TOTAL ({len(rc_detail)} checks)',bold=True,bg=MD_GREEN,fg=WHITE)
                for col in range(3,6): ws4.cell(r,col).fill=PatternFill("solid",fgColor=MD_GREEN); ws4.cell(r,col).border=bdr()
                nc(ws4.cell(r,6),sum(c.get('monto',0) for c in rc_detail),bold=True,bg=MD_GREEN,fg=WHITE)
                ws4.merge_cells(start_row=r,start_column=7,end_row=r,end_column=10)
                cs(ws4.cell(r,7),'Total enviado a Opera',bg=MD_GREEN,fg=WHITE,bold=True,center=True)
                for col in range(7,11): ws4.cell(r,col).border=bdr()
                r+=2

            # ── Desglose por Forma de Pago ────────────────────────────────────
            by_pay = meta.get('by_payment', [])
            if by_pay:
                mhdr(ws4,r,2,10,"  VENTAS POR FORMA DE PAGO",TEAL_DK,left=True,h=15); r+=1
                ws4.row_dimensions[r].height=16
                for c2,h in enumerate(['Forma de Pago','# Checks','Monto (USD)','Destino'],2):
                    cs(ws4.cell(r,c2),h,bg=TEAL_DK,fg=WHITE,bold=True,center=True)
                ws4.merge_cells(start_row=r,start_column=5,end_row=r,end_column=10)
                cs(ws4.cell(r,5),'Destino',bg=TEAL_DK,fg=WHITE,bold=True,center=True)
                for col in range(5,11): ws4.cell(r,col).border=bdr()
                r+=1
                for i,fp in enumerate(sorted(by_pay,key=lambda x:-x['total'])):
                    goes_opera = 'ROOM CHARGE' in fp['forma'].upper()
                    bg = LT_GREEN if goes_opera else (TEAL_LT if i%2==0 else None)
                    fg2 = OK_TXT if goes_opera else DK_GRAY
                    dest = "→ Opera (Room Charge)" if goes_opera else "Interno / Paquete"
                    ws4.row_dimensions[r].height=15
                    cs(ws4.cell(r,2),fp['forma'],bg=bg,bold=goes_opera,fg=fg2)
                    nc(ws4.cell(r,3),fp['count'],bg=bg,fmt='0')
                    nc(ws4.cell(r,4),fp['total'],bg=bg,bold=goes_opera,fg=fg2)
                    ws4.merge_cells(start_row=r,start_column=5,end_row=r,end_column=10)
                    cs(ws4.cell(r,5),dest,bg=OK_BG if goes_opera else bg,
                       fg=OK_TXT if goes_opera else "888888",sz=8,bold=goes_opera)
                    for col in range(5,11): ws4.cell(r,col).border=bdr()
                    r+=1
                # Total
                ws4.row_dimensions[r].height=16
                cs(ws4.cell(r,2),'TOTAL GENERAL',bold=True,bg=SLATE_LT)
                nc(ws4.cell(r,3),sum(f['count'] for f in by_pay),bold=True,bg=SLATE_LT,fmt='0')
                nc(ws4.cell(r,4),sum(f['total'] for f in by_pay),bold=True,bg=SLATE_LT)
                ws4.merge_cells(start_row=r,start_column=5,end_row=r,end_column=10)
                cs(ws4.cell(r,5),'',bg=SLATE_LT)
                for col in range(5,11): ws4.cell(r,col).border=bdr()
                r+=2

            # ── Desglose por Empleado ─────────────────────────────────────────
            by_emp = meta.get('by_employee', [])
            if by_emp:
                mhdr(ws4,r,2,10,"  VENTAS POR EMPLEADO (control de cajeros)",MD_GREEN,left=True,h=15); r+=1
                ws4.row_dimensions[r].height=16
                for c2,h in enumerate(['Empleado','# Checks','Monto Total (USD)','Promedio/Check'],2):
                    cs(ws4.cell(r,c2),h,bg=MD_GREEN,fg=WHITE,bold=True,center=True)
                ws4.merge_cells(start_row=r,start_column=5,end_row=r,end_column=10)
                cs(ws4.cell(r,5),'Promedio por Check',bg=MD_GREEN,fg=WHITE,bold=True,center=True)
                for col in range(5,11): ws4.cell(r,col).border=bdr()
                r+=1
                for i,em in enumerate(sorted(by_emp,key=lambda x:-x['total'])):
                    bg=LT_GREEN if i%2==0 else None
                    avg = em['total']/em['count'] if em['count'] else 0
                    ws4.row_dimensions[r].height=15
                    cs(ws4.cell(r,2),em['empleado'],bg=bg)
                    nc(ws4.cell(r,3),em['count'],bg=bg,fmt='0')
                    nc(ws4.cell(r,4),em['total'],bg=bg)
                    ws4.merge_cells(start_row=r,start_column=5,end_row=r,end_column=10)
                    nc(ws4.cell(r,5),round(avg,2),bg=bg)
                    for col in range(6,11): ws4.cell(r,col).fill=PatternFill("solid",fgColor=bg) if bg else PatternFill(); ws4.cell(r,col).border=bdr()
                    r+=1
                r+=1

            # ── Todos los checks (detalle completo dinámico) ──────────────────
            all_chk = meta.get('all_checks', [])
            if all_chk:
                mhdr(ws4,r,2,10,f"  DETALLE COMPLETO — {len(all_chk)} CHECKS CERRADOS",SLATE,left=True,h=15); r+=1
                ws4.row_dimensions[r].height=16
                for c2,h in enumerate(['Restaurante','Empleado','# Check','Hora','Forma de Pago','Monto USD'],2):
                    cs(ws4.cell(r,c2),h,bg=SLATE,fg=WHITE,bold=True,center=True)
                ws4.merge_cells(start_row=r,start_column=8,end_row=r,end_column=10)
                cs(ws4.cell(r,8),'Monto (USD)',bg=SLATE,fg=WHITE,bold=True,center=True)
                for col in range(8,11): ws4.cell(r,col).border=bdr()
                r+=1
                for i,chk in enumerate(all_chk):
                    is_rc = 'ROOM CHARGE' in chk['forma_pago'].upper()
                    bg = LT_GREEN if is_rc else (SLATE_LT if i%2==0 else None)
                    ws4.row_dimensions[r].height=14
                    cs(ws4.cell(r,2),chk['restaurant'],bg=bg,sz=8)
                    cs(ws4.cell(r,3),chk['employee'],bg=bg,sz=8)
                    cs(ws4.cell(r,4),chk['check_num'],bg=bg,center=True,bold=True)
                    cs(ws4.cell(r,5),chk['hora'],bg=bg,center=True)
                    cs(ws4.cell(r,6),chk['forma_pago'],bg=bg,sz=8,
                       fg=OK_TXT if is_rc else DK_GRAY,bold=is_rc)
                    ws4.merge_cells(start_row=r,start_column=7,end_row=r,end_column=10)
                    nc(ws4.cell(r,7),chk['monto'],bg=bg,bold=is_rc)
                    for col in range(8,11): ws4.cell(r,col).border=bdr()
                    r+=1
            ws4.freeze_panes='B5'
        else:
            # ── POS desde EVT ─────────────────────────────────────────────────
            for col,w in {'A':2,'B':10,'C':8,'D':10,'E':10,'F':8,'G':13,'H':13,'I':13,'J':13}.items():
                ws4.column_dimensions[col].width=w
            mhdr(ws4,1,2,9,f"SIMPHONY POS — CHECKS POSTEADOS A HABITACIÓN — {fecha_fmt}",TEAL_DK,sz=13,h=28)
            ws4.row_dimensions[3].height=20
            for c2,h in enumerate(['Check #','Hab.','Comida','Fecha POS','Hora','User',
                                    'Sub-neto USD','Serv.Charge','IVA','Total USD'],2):
                cs(ws4.cell(3,c2),h,bg=TEAL_DK,fg=WHITE,bold=True,center=True)
            meal_map={'2':'Lunch','3':'Dinner'}
            for i,chk in enumerate(sorted(pos_checks,key=lambda x:x.get('check_num','0'))):
                r=4+i; bg=TEAL_LT if i%2==0 else None
                d=chk.get('pos_date',''); fecha_pos=f"{d[4:6]}/{d[2:4]}/20{d[:2]}" if len(d)==6 else d
                t=chk.get('pos_time','000000')
                ws4.row_dimensions[r].height=16
                cs(ws4.cell(r,2),chk.get('check_num'),bg=bg,center=True,bold=True)
                cs(ws4.cell(r,3),chk.get('room'),bg=bg,center=True)
                cs(ws4.cell(r,4),meal_map.get(chk.get('serving',''),'?'),bg=bg,center=True)
                cs(ws4.cell(r,5),fecha_pos,bg=bg,center=True)
                cs(ws4.cell(r,6),f"{t[:2]}:{t[2:4]}:{t[4:6]}",bg=bg,center=True)
                cs(ws4.cell(r,7),chk.get('user'),bg=bg,center=True)
                nc(ws4.cell(r,8),chk.get('sub_usd'),bg=bg)
                nc(ws4.cell(r,9),chk.get('sc_usd'),bg=bg)
                iva_rate=chk.get('tax_usd',0)/(chk.get('sub_usd',0)+chk.get('sc_usd',0))*100 if (chk.get('sub_usd',0)+chk.get('sc_usd',0))>0 else 0
                warn=iva_rate<12.0
                nc(ws4.cell(r,10),chk.get('tax_usd'),bg=RED_BG if warn else bg,fg=RED_T if warn else DK_GRAY,bold=warn)
                nc(ws4.cell(r,11),chk.get('total_usd'),bg=bg,bold=True)
            ws4.freeze_panes='B4'

    # ── HOJA 5: MAPEO ─────────────────────────────────────────────────────────
    ws5=wb.create_sheet("Mapeo TCode-Cuenta")
    ws5.sheet_view.showGridLines=False
    for col,w in {'A':2,'B':8,'C':32,'D':14,'E':32,'F':52}.items():
        ws5.column_dimensions[col].width=w
    mhdr(ws5,1,2,6,"MAPEO Opera TCode ↔ Cuenta Contable Integrity",MD_GREEN,sz=12,h=26)
    ws5.row_dimensions[3].height=20
    for c2,h in enumerate(['TCode','Descripción Opera','Tipo','Cuenta Contable','Nombre Cuenta'],2):
        cs(ws5.cell(3,c2),h,bg=MD_GREEN,fg=WHITE,bold=True,center=True)
    r=4
    for _,rd in merged.iterrows():
        if rd.get('type') in ('INTERNAL','PACKAGE'): continue
        bg=LT_GREEN if r%2==0 else None
        ws5.row_dimensions[r].height=15
        cs(ws5.cell(r,2),rd['tcode'],center=True,bg=bg,bold=True)
        cs(ws5.cell(r,3),rd.get('description',''),bg=bg)
        cs(ws5.cell(r,4),rd.get('type',''),center=True,bg=bg)
        cuenta=rd.get('cuenta'); nombre=rd.get('nombre')
        cs(ws5.cell(r,5),str(cuenta) if pd.notna(cuenta) else '— NO MAPEADO —',
           bg=bg,sz=8,fg=DK_GRAY if pd.notna(cuenta) else RED_T)
        cs(ws5.cell(r,6),str(nombre) if pd.notna(nombre) else '— NO MAPEADO —',
           bg=bg,sz=8,fg=DK_GRAY if pd.notna(nombre) else RED_T)
        r+=1
    ws5.freeze_panes='B4'

    # ── Pestañas adicionales (Trial Balance, Ledgers, Estadísticas, OTB, Market Code) ──
    try:
        _sheet_trial_balance(wb, fecha_fmt, gen_str, opera_data['opera_hdr'], integrity_df)
    except Exception as e:
        print(f"  ⚠  Trial Balance: {e}")
    try:
        _sheet_ledgers(wb, fecha_fmt, gen_str, opera_data['opera_hdr'], opera_data.get('hf_total',{}))
    except Exception as e:
        print(f"  ⚠  Ledgers: {e}")
    try:
        if 'stats' in opera_data and 'opera_det' in opera_data:
            _sheet_estadisticas(wb, fecha_fmt, gen_str, opera_data['stats'],
                                opera_data['opera_det'], opera_data.get('hf_total',{}))
    except Exception as e:
        print(f"  ⚠  Estadísticas: {e}")
    try:
        if opera_data.get('hf_total') and opera_data.get('hf_rooms'):
            _sheet_otb(wb, fecha_fmt, gen_str, opera_data['opera_hdr'],
                       opera_data['opera_det'], opera_data.get('stats'),
                       opera_data['hf_total'], opera_data['hf_rooms'])
    except Exception as e:
        print(f"  ⚠  OTB: {e}")
    try:
        if 'opera_det' in opera_data:
            _sheet_market_code(wb, fecha_fmt, gen_str, opera_data['opera_hdr'], opera_data['opera_det'])
    except Exception as e:
        print(f"  ⚠  Market Code: {e}")

    # ── Reordenar pestañas en el orden estándar ───────────────────────────────
    orden = ['Resumen Ejecutivo','Trial Balance','Ledgers','Estadisticas Ocupacion',
             'OTB vs Revenue','Ingresos x Market Code','Detalle Reconciliacion',
             'Discrepancias','Simphony POS','Mapeo TCode-Cuenta']
    for i, sname in enumerate(orden):
        if sname in wb.sheetnames:
            wb.move_sheet(sname, offset=-(wb.sheetnames.index(sname))+i)

    # ── Insertar PORTADA desde el template (si existe) ────────────────────────
    import shutil
    script_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(script_dir, 'template_auditoria.xlsx')
    if os.path.exists(template_path):
        try:
            wb_tpl = load_workbook(template_path)
            if 'PORTADA' in wb_tpl.sheetnames:
                src = wb_tpl['PORTADA']
                # Create PORTADA as first sheet
                dst = wb.create_sheet('PORTADA', 0)
                dst.sheet_view.showGridLines = False
                # Copy column widths
                for col_letter, dim in src.column_dimensions.items():
                    if dim.width:
                        dst.column_dimensions[col_letter].width = dim.width
                # Copy cells with styles
                from copy import copy as _copy
                for row in src.iter_rows():
                    for cell in row:
                        if cell.value is not None or cell.has_style:
                            nc2 = dst.cell(row=cell.row, column=cell.column)
                            nc2.value = cell.value
                            if cell.has_style:
                                nc2.font          = _copy(cell.font)
                                nc2.fill          = _copy(cell.fill)
                                nc2.border        = _copy(cell.border)
                                nc2.alignment     = _copy(cell.alignment)
                                nc2.number_format = cell.number_format
                # Copy merged ranges
                for mr in src.merged_cells.ranges:
                    dst.merge_cells(str(mr))
                # Copy row heights
                for ridx, dim in src.row_dimensions.items():
                    if dim.height:
                        dst.row_dimensions[ridx].height = dim.height
                # Update dates on portada
                dst['D7'].value = fecha_fmt
                dst['D8'].value = datetime.datetime.now().strftime('%d/%m/%Y %H:%M')
        except Exception as e:
            print(f"  ⚠  No se pudo copiar PORTADA: {e}")

    wb.save(output_path)
    return output_path

# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description='Auditoría de Ingresos Corcovado')
    parser.add_argument('--fecha', required=True,
                        help='Fecha a auditar: YYYY-MM-DD o "hoy"')
    parser.add_argument('--solo-reporte', action='store_true',
                        help='No sube a Drive — solo genera el Excel localmente')
    args = parser.parse_args()

    if args.fecha.lower() == 'hoy':
        fecha_str = datetime.date.today().strftime('%Y-%m-%d')
    else:
        fecha_str = args.fecha

    print(f"\n{'='*60}")
    print(f"  AUDITORÍA DE INGRESOS — {fecha_str}")
    print(f"  Corcovado Wilderness Lodge")
    print(f"{'='*60}")

    # Autenticar Drive
    print("\n🔐 Autenticando con Google Drive...")
    service = get_service()
    print("  ✓ Autenticación exitosa")

    # Obtener/crear carpetas del día
    print(f"\n📁 Preparando carpetas en Drive para {fecha_str}...")
    folders = get_or_create_day_folders(service, fecha_str)
    for k,v in folders.items():
        print(f"  {k:<20} {v}")

    # Descargar archivos
    with tempfile.TemporaryDirectory() as work_dir:
        downloaded = download_inputs(service, folders, work_dir, fecha_str)

        # Verificar que tenemos lo mínimo
        if not downloaded['opera']:
            print("\n❌ No hay archivos en la carpeta 'opera' de Drive. Abortando.")
            print(f"   Subí los XMLs a: Drive → Inputs → {fecha_str} → opera")
            sys.exit(1)
        if not downloaded['integrity']:
            print("\n❌ No hay archivos en la carpeta 'integrity' de Drive. Abortando.")
            sys.exit(1)

        # Parsear
        print("\n🔍 Parseando archivos...")
        opera_data = parse_opera(downloaded['opera'])
        integrity_df = parse_integrity(downloaded['integrity'])
        pos_checks = parse_pos_log(downloaded.get('pos',[]))
        print(f"  ✓ Opera: {len(opera_data.get('opera_hdr',[]))} TCode")
        print(f"  ✓ Integrity: {len(integrity_df)} líneas")
        print(f"  ✓ POS: {len(pos_checks)} checks")

        # Generar Excel
        print("\n📊 Generando reporte Excel...")
        output_name = f"Auditoria_Ingresos_{fecha_str.replace('-','')}.xlsx"
        output_path = os.path.join(work_dir, output_name)
        build_excel(fecha_str, opera_data, integrity_df, pos_checks, output_path)
        print(f"  ✓ Excel generado: {output_name}")

        # Resumen en consola
        merged = reconcile(opera_data['opera_hdr'], integrity_df)
        ok    = (merged['estado']=='OK').sum()
        disc  = (merged['estado']=='DISCREPANCIA').sum()
        miss  = merged['estado'].isin(['FALTA EN INTEGRITY','FALTA EN OPERA']).sum()
        print(f"\n{'='*60}")
        print(f"  RESULTADO: ✓ {ok} OK  |  ⚠ {disc} discrepancias  |  ✗ {miss} faltantes")
        print(f"{'='*60}")

        # Subir a Drive
        if not args.solo_reporte:
            print(f"\n☁️  Subiendo reporte a Drive → Reportes/{fecha_str}/...")
            file_id, final_name, version_msg = upload_file(
                service, output_path, folders['day_report'], output_name)
            if final_name != output_name:
                print(f"  ℹ  Ya existía {output_name} → guardado como {final_name}")
            print(f"  ✓ {version_msg.capitalize()}: {final_name}")
            print(f"  🔗 https://drive.google.com/file/d/{file_id}/view")
            print(f"\n  📁 Ruta en Drive:")
            print(f"     H:\\My Drive\\Auditoria Corcovado\\Reportes\\{fecha_str}\\{final_name}")
        else:
            # Copiar localmente
            import shutil
            local_out = os.path.join(os.path.dirname(__file__), output_name)
            shutil.copy(output_path, local_out)
            print(f"\n💾 Guardado localmente: {local_out}")

    print(f"\n✅ Auditoría completada para {fecha_str}\n")

if __name__ == '__main__':
    main()
