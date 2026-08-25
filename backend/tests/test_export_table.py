"""Export genérico a Excel (`app/export/table.py`) — el motor que usan todos
los tabs. Lo que se fija acá es lo que rompería el archivo para el usuario:
que los números lleguen como números, que los totales se distingan, que los
grupos de encabezado queden combinados y que un nombre de hoja imposible no
haga que Excel no abra el archivo.
"""
from io import BytesIO

from openpyxl import load_workbook

from app.export.table import build_workbook


def _abrir(payload: dict):
    return load_workbook(BytesIO(build_workbook(payload)))


def _payload(**extra) -> dict:
    hoja = {
        "name": "Summary",
        "columns": [
            {"label": "Revenue Center", "type": "text"},
            {"label": "Actual", "type": "money"},
            {"label": "Var %", "type": "pct"},
            {"label": "RN", "type": "int"},
        ],
        "rows": [
            ["Rooms", 5122.41, 0.2136, 13],
            ["F&B", 1696.31, -0.0392, 0],
            ["GRAND TOTAL", 6818.72, 0.1744, 13],
        ],
        "total_rows": [2],
    }
    hoja.update(extra)
    return {"title": "Tab 3 · Daily Revenue", "subtitle": "COWLCR · 2026-08-24", "sheets": [hoja]}


def test_los_numeros_van_como_numeros_no_como_texto():
    """Si salieran como texto, en Excel no se pueden sumar ni filtrar -- que es
    la única razón por la que alguien baja un reporte a Excel."""
    ws = _abrir(_payload()).active
    fila = next(r for r in ws.iter_rows(values_only=True) if r and r[0] == "Rooms")
    assert isinstance(fila[1], float)
    assert isinstance(fila[3], int)


def test_cada_tipo_de_columna_lleva_su_formato():
    ws = _abrir(_payload()).active
    encabezado = next(i for i, r in enumerate(ws.iter_rows(values_only=True), start=1)
                      if r and r[0] == "Revenue Center")
    datos = encabezado + 1
    assert "#,##0.00" in ws.cell(row=datos, column=2).number_format
    assert "0.0%" in ws.cell(row=datos, column=3).number_format
    assert ws.cell(row=datos, column=1).number_format == "General"


def test_la_fila_de_total_se_distingue():
    ws = _abrir(_payload()).active
    fila = next(i for i, r in enumerate(ws.iter_rows(values_only=True), start=1)
                if r and r[0] == "GRAND TOTAL")
    celda = ws.cell(row=fila, column=1)
    assert celda.font.bold is True
    assert celda.border.top.style == "medium"


def test_los_grupos_de_encabezado_quedan_combinados():
    """TODAY / MONTH TO DAY / FULL MONTH RESULT: el colSpan de la pantalla tiene
    que verse como celda combinada, no como una etiqueta suelta."""
    p = _payload(header_groups=[{"label": "", "span": 1},
                                {"label": "TODAY", "span": 2},
                                {"label": "STATS", "span": 1}])
    ws = _abrir(p).active
    assert any(str(r) .startswith("B") and ":" in str(r) for r in ws.merged_cells.ranges)


def test_un_grupo_mas_ancho_que_la_tabla_no_desborda():
    p = _payload(header_groups=[{"label": "TODO", "span": 99}])
    ws = _abrir(p).active
    assert ws.max_column == 4


def test_encabezado_congelado_y_titulos_de_impresion():
    """Una tabla de 12 meses es ilegible sin el encabezado y la primera columna
    fijos; y al imprimir el encabezado tiene que repetirse en cada página."""
    ws = _abrir(_payload()).active
    assert ws.freeze_panes.startswith("B")
    assert ws.print_title_rows
    assert ws.page_setup.orientation == "landscape"


def test_nombre_de_hoja_saneado_y_sin_repetir():
    """Excel no acepta []:*?/\\ ni más de 31 caracteres, y dos hojas no pueden
    llamarse igual -- cualquiera de las tres cosas deja el archivo sin abrir."""
    largo = "Room Statistics by Category [§5.2] / ADR & Occupancy detail"
    wb = _abrir({"sheets": [{"name": largo, "columns": [{"label": "a", "type": "text"}], "rows": [["x"]]},
                            {"name": largo, "columns": [{"label": "a", "type": "text"}], "rows": [["y"]]}]})
    for nombre in wb.sheetnames:
        assert len(nombre) <= 31
        assert not set(nombre) & set("[]:*?/\\")
    assert len(set(wb.sheetnames)) == 2


def test_varias_hojas_en_un_archivo():
    wb = _abrir({"sheets": [
        {"name": "Summary", "columns": [{"label": "a", "type": "text"}], "rows": [["x"]]},
        {"name": "On-Property", "columns": [{"label": "b", "type": "money"}], "rows": [[1.5]]},
    ]})
    assert wb.sheetnames == ["Summary", "On-Property"]


def test_payload_sin_hojas_devuelve_un_archivo_valido():
    """Mejor un archivo que dice 'sin datos' que un .xlsx corrupto."""
    wb = _abrir({"sheets": []})
    assert wb.sheetnames
    assert wb.active["A1"].value


def test_fila_mas_corta_que_las_columnas_no_explota():
    p = _payload(rows=[["Solo la etiqueta"]])
    ws = _abrir(p).active
    assert ws.max_column == 4
