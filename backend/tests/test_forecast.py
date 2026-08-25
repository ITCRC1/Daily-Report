"""Forecast (Tab 6.1.1) — contrato con el Budget (Tab 6.1).

El Forecast es el gemelo del Budget: misma plantilla, mismo grano y el mismo
motor de derivación diaria. Estos tests fijan justamente eso, que es lo que se
rompe solo: si alguien le agrega una columna a la plantilla del budget y se
olvida del forecast, los dos Excel dejan de ser intercambiables y el upload
carga los campos corridos sin fallar.
"""
from decimal import Decimal

from app.engine.budget import derive_daily_amounts
from app.main import app
from app.services import budget_service, forecast_service


def test_plantilla_identica_a_la_del_budget():
    """Mismas columnas y en el mismo orden: un archivo llenado para uno sirve
    para el otro, y el parser posicional del upload no se corre."""
    assert forecast_service.HEADERS == budget_service.HEADERS


def test_las_cuatro_rutas_quedan_registradas():
    rutas = {r.path: getattr(r, "methods", set()) for r in app.routes}
    assert "GET" in rutas["/master-data/forecast"]
    assert "GET" in rutas["/master-data/forecast/daily"]
    assert "GET" in rutas["/master-data/forecast/template"]
    assert "POST" in rutas["/master-data/forecast/upload"]


def test_no_pisa_las_rutas_del_budget():
    """Aditivo: el Tab 6.1 sigue en pie tal como estaba."""
    rutas = {r.path for r in app.routes}
    for p in ("/master-data/budget", "/master-data/budget/daily",
              "/master-data/budget/template", "/master-data/budget/upload"):
        assert p in rutas


def test_deriva_el_diario_con_el_mismo_motor_que_el_budget():
    """El forecast diario se reparte igual que fact_budget: mensual ÷ días del
    mes con el residual al último día, de modo que Σ diarios = mensual exacto."""
    monthly = Decimal("121219.07")
    daily = derive_daily_amounts(monthly, 2026, 5)
    assert sum(daily) == monthly
    assert len(daily) == 31


def test_tablas_propias_no_las_del_budget():
    """Tabla aparte a propósito (§ migración f2a3b4c5d6e7): el reemplazo anual
    de uno no puede pisar al otro."""
    from app.models import Budget, BudgetMonthly, Forecast, ForecastMonthly

    assert ForecastMonthly.__tablename__ == "forecast_monthly"
    assert Forecast.__tablename__ == "fact_forecast"
    assert BudgetMonthly.__tablename__ == "budget_monthly"
    assert Budget.__tablename__ == "fact_budget"


# --- Layout "Working" (Ingresos_12M_Working_YYYY.xlsx) -----------------------
def _working_wb():
    """Mini archivo con la misma forma que el Working real: bloque de
    estadísticos arriba, fila de encabezado, líneas de ingreso y TOTAL."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Corcovado Wilderness Lodge"
    ws["A4"] = "ESTADÍSTICOS"
    ws.append([])
    for label, vals in (
        ("Total available Rooms", [930] + [0] * 11),
        ("Total Rooms Occupied", [637] + [0] * 11),
        ("Total Guests", [1181] + [0] * 11),
        ("% Occupancy", [0.6849] + [0] * 11),
        ("ADR", [610.56] + [0] * 11),
    ):
        ws.append([label] + vals)
    ws.append(["Línea de ingreso"] + [f"M{i}" for i in range(1, 13)])
    ws.append(["Rooms Revenue-", 417011.88] + [0] * 11)
    ws.append(["Rooms Revenue-No Show", 10246.62] + [0] * 11)
    ws.append(["F&B Food", 99072.98] + [0] * 11)
    ws.append(["Oinn"] + [None] * 12)
    ws.append(["TOTAL INGRESOS", 526331.48] + [0] * 11)
    return wb


def test_detecta_el_layout_working():
    from app.services.forecast_service import _is_working_layout
    assert _is_working_layout(_working_wb().active) is True


def test_suma_las_dos_lineas_de_rooms_en_el_mismo_depto():
    """`Rooms Revenue-` y su no-show van los dos al 0110: si se asignara en vez
    de acumular, una de las dos desaparecería sin aviso."""
    from app.services.forecast_service import _parse_working

    amounts, stats, unmapped = _parse_working(_working_wb().active)
    assert not unmapped
    assert amounts[("0110", 1)] == Decimal("417011.88") + Decimal("10246.62")
    assert amounts[("FB-FOOD", 1)] == Decimal("99072.98")


def test_no_escribe_la_fila_de_total_ni_las_lineas_vacias():
    """TOTAL INGRESOS no es una línea de ingreso, y una línea toda en blanco
    (Oinn) no se carga como cero: celda vacía ≠ 0."""
    from app.services.forecast_service import _parse_working

    amounts, _stats, _unmapped = _parse_working(_working_wb().active)
    assert not any(code == "0155" for code, _m in amounts)  # Oinn, todo vacío
    assert all(code in ("0110", "FB-FOOD") for code, _m in amounts)


def test_los_estadisticos_salen_del_bloque_de_arriba():
    from app.services.forecast_service import _parse_working

    _amounts, stats, _unmapped = _parse_working(_working_wb().active)
    assert stats[1]["available_rooms"] == Decimal("930")
    assert stats[1]["rooms_occupied"] == Decimal("637")
    assert stats[1]["adr"] == Decimal("610.56")


def test_una_linea_desconocida_se_reporta_no_se_descarta():
    """§10: nunca descartar en silencio."""
    from app.services.forecast_service import _parse_working

    wb = _working_wb()
    wb.active.append(["Línea Nueva Que Nadie Mapeó", 500.0] + [0] * 11)
    _amounts, _stats, unmapped = _parse_working(wb.active)
    assert unmapped == ["Línea Nueva Que Nadie Mapeó"]


# output_column real de cada departamento destino (dim_department, Tab 6.3).
# Los códigos sintéticos (FB-*, MISC-REV) no están en OUTLET_TO_COLUMN: su
# columna vive en la tabla, así que acá se declara explícita.
_DEPT_OUTPUT_COLUMN = {
    "0110": "Rooms", "0140": "SPA", "0150": "Tours", "0151": "Retail-Gift Shop",
    "0152": "Transportation", "0155": "Innoceana", "0160": "Laundry",
    "0170": "Sustainable Fee", "FB-FOOD": "F&B", "FB-BEV": "F&B",
    "FB-MISC": "F&B", "MISC-REV": "Otros",
    # las dos que NO tienen fila en el cuadro de Tab 3:
    "ROOMS-OTH": "Rooms Others", "MISC-REV-OTH": "Misc. Rev Others",
}


def test_todas_las_lineas_mapean_a_una_fila_visible_del_tab3():
    """Cada destino tiene que caer en una fila que el cuadro FULL MONTH RESULT
    realmente muestre; si no, el GRAND TOTAL sumaría plata que no aparece en
    ninguna fila -- el descuadre que este sistema vino a eliminar."""
    from app.engine.revenue import OUTPUT_COLUMNS
    from app.services.forecast_service import WORKING_LINE_TO_DEPT

    visibles = set(OUTPUT_COLUMNS) | {"Otros"}  # "Otros" = la fila Misc. Revenue
    for linea, code in WORKING_LINE_TO_DEPT.items():
        col = _DEPT_OUTPUT_COLUMN[code]
        assert col in visibles, f"'{linea}' -> {code} ({col}) no tiene fila en Tab 3"


def test_el_mapeo_no_usa_los_deptos_sin_fila():
    """Guardarraíl explícito de la decisión: el no-show va dentro de Rooms y
    'Other / Misc Revenue' a MISC-REV, no a los *-OTH."""
    from app.services.forecast_service import WORKING_LINE_TO_DEPT

    destinos = set(WORKING_LINE_TO_DEPT.values())
    assert "ROOMS-OTH" not in destinos
    assert "MISC-REV-OTH" not in destinos
    assert WORKING_LINE_TO_DEPT["Rooms Revenue-No Show"] == "0110"
    assert WORKING_LINE_TO_DEPT["Other / Misc Revenue"] == "MISC-REV"
