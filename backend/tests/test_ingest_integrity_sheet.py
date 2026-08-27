"""La hoja del Integrity se ubica por su encabezado, no por su nombre.

El emisor renombró la hoja de 'Datos' a 'Asiento' (2026-08-27) sin aviso y la
ingesta dejó de reconocer el archivo. Estos tests fijan que el nombre no
importa: lo que identifica al mayor son sus columnas.

Los libros se arman en memoria a propósito, para que el test no dependa de
`goldens/inputs/` (que no se versiona).
"""
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from app.ingest import integrity as integ

COLUMNAS = ["Cuenta", "Nombre cuenta", "Centro de costo", "Referencia", "Detalle",
            "Moneda", "T.C.", "Débitos Col", "Créditos Col", "Débitos Dol", "Créditos Dol"]


def _libro(nombre_hoja: str, filas_preambulo: int = 8) -> BytesIO:
    """Un Integrity mínimo: preámbulo, encabezado y dos asientos."""
    wb = Workbook()
    ws = wb.active
    ws.title = nombre_hoja
    for i in range(filas_preambulo):
        ws.append([f"linea de preambulo {i}"])
    ws.append(COLUMNAS)
    ws.append(["4000-0110-001", "Habitaciones", "0110", "TCode: 1000", "ACCOMMODATION",
               "DOL", 500.0, 0, 0, 0, 250.00])
    ws.append(["1000-0001-002", "BAC", "0000", "TCode: 3722", "PAYMENT BAC",
               "DOL", 500.0, 0, 0, 100.00, 0])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@pytest.mark.parametrize("nombre", ["Datos", "Asiento", "Hoja1", "Cualquier Cosa"])
def test_encuentra_la_hoja_sin_importar_el_nombre(nombre: str) -> None:
    wb = load_workbook(_libro(nombre), read_only=True)
    assert integ.find_sheet(wb) == nombre


@pytest.mark.parametrize("nombre", ["Datos", "Asiento"])
def test_parsea_igual_con_cualquier_nombre(nombre: str) -> None:
    lineas = integ.parse_integrity_lines(_libro(nombre))
    assert len(lineas) == 2
    assert [l["tcode"] for l in lineas] == ["1000", "3722"]
    assert lineas[0]["cred_usd"] == 250.00
    assert lineas[1]["deb_usd"] == 100.00


def test_encabezado_en_otra_fila() -> None:
    """La fila 8 es costumbre del emisor, no un contrato: si mueve el
    encabezado, se sigue encontrando."""
    lineas = integ.parse_integrity_lines(_libro("Asiento", filas_preambulo=3))
    assert len(lineas) == 2
    assert lineas[0]["cuenta"] == "4000-0110-001"


def test_libro_sin_mayor_no_es_integrity() -> None:
    """Un .xlsx que no tiene las columnas del mayor no debe clasificarse como
    Integrity -- si no, se llevaría por delante al POS."""
    wb = Workbook()
    wb.active.title = "Resumen Ejecutivo"
    wb.active.append(["Restaurante", "Total"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    assert integ.find_sheet(load_workbook(buf, read_only=True)) is None
